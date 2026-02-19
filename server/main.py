"""
FastAPI приложение - главный файл
REST API для многопользовательской CRM
"""
import logging
import os
import asyncio
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from sqlalchemy.exc import IntegrityError
from datetime import datetime, timedelta
from typing import List, Optional
import io
import json
import threading

# Блокировка для предотвращения параллельных сканирований одного договора
_scanning_contracts_lock = threading.Lock()
_scanning_contracts = set()

# Brute-force защита логина: персистентная (в БД) + in-memory fallback
from collections import defaultdict
_login_attempts = defaultdict(list)
_LOGIN_MAX_ATTEMPTS = 5
_LOGIN_BLOCK_MINUTES = 15

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

from config import get_settings
from database import (
    get_db, init_db,
    Employee, Client, Contract, Notification, UserSession, ActivityLog,
    CRMCard, StageExecutor, SupervisionCard, SupervisionProjectHistory,
    Payment, Rate, Salary, ProjectFile, ActionHistory, ConcurrentEdit,
    ApprovalStageDeadline, ProjectTimelineEntry, SupervisionTimelineEntry,
    StageWorkflowState,
    MessengerChat, MessengerChatMember, MessengerScript, MessengerSetting, MessengerMessageLog
)
from schemas import (
    LoginRequest, TokenResponse,
    EmployeeResponse, EmployeeCreate, EmployeeUpdate,
    ClientResponse, ClientCreate, ClientUpdate,
    ContractResponse, ContractCreate, ContractUpdate,
    NotificationResponse, SyncRequest, SyncResponse,
    # CRM
    CRMCardCreate, CRMCardUpdate, CRMCardResponse,
    ColumnMoveRequest, StageExecutorCreate, StageExecutorUpdate, StageExecutorResponse,
    # Supervision
    SupervisionCardCreate, SupervisionCardUpdate, SupervisionCardResponse,
    SupervisionColumnMoveRequest, SupervisionPauseRequest, SupervisionHistoryCreate, SupervisionHistoryResponse,
    # Payments
    PaymentCreate, PaymentUpdate, PaymentResponse,
    # Rates
    RateCreate, RateUpdate, RateResponse,
    # Salaries
    SalaryCreate, SalaryUpdate, SalaryResponse,
    # Files
    ProjectFileCreate, ProjectFileResponse,
    # Action History
    ActionHistoryCreate, ActionHistoryResponse,
    # Extended Request Schemas
    PaymentManualUpdateRequest, TemplateRateRequest, IndividualRateRequest,
    SupervisionRateRequest, SurveyorRateRequest, ContractFilesUpdate,
    # CRM Approval/Executor Request Schemas
    CompleteApprovalStageRequest, StageExecutorDeadlineRequest,
    CompleteStageExecutorRequest, ManagerAcceptanceRequest,
    # Timeline
    TimelineEntryCreate, TimelineEntryUpdate, TimelineEntryResponse, TimelineInitRequest,
    SupervisionTimelineCreate, SupervisionTimelineUpdate, SupervisionTimelineResponse
)
from messenger_schemas import (
    MessengerChatCreate, MessengerChatBind, MessengerChatResponse, MessengerChatDetailResponse,
    ChatMemberInput, ChatMemberResponse,
    MessengerScriptCreate, MessengerScriptUpdate, MessengerScriptResponse,
    MessengerSettingUpdate, MessengerSettingResponse, MessengerSettingsBulkUpdate,
    SendMessageRequest, SendFilesRequest, SendScriptMessageRequest, MessageLogResponse,
    SendInvitesRequest
)
from telegram_service import get_telegram_service, TelegramService, PYROGRAM_AVAILABLE
from email_service import get_email_service, EmailService
from auth import (
    verify_password, get_password_hash, create_access_token,
    create_refresh_token, verify_refresh_token, get_current_user
)
from permissions import (
    require_permission, check_permission as perm_check,
    PERMISSION_NAMES, DEFAULT_ROLE_PERMISSIONS, SUPERUSER_ROLES,
    seed_permissions, invalidate_cache as invalidate_perm_cache,
    get_employee_permissions, set_employee_permissions, reset_to_defaults,
)
from schemas import PermissionSetRequest, PermissionResponse, PermissionDefinition, LockRequest

settings = get_settings()

# Создание приложения
app = FastAPI(
    title=settings.app_name,
    version=settings.app_version,
    description="REST API для многопользовательской CRM Interior Studio"
)

# CORS middleware — ЗАПРЕЩЁН wildcard "*" при allow_credentials=True
_allowed_origins = os.environ.get("ALLOWED_ORIGINS", "").strip()
if not _allowed_origins or _allowed_origins == "*":
    # По умолчанию разрешаем только localhost (для разработки)
    _cors_origins = ["http://localhost:3000", "http://localhost:8080"]
    logger.warning("CORS: ALLOWED_ORIGINS не задан, используются origins для разработки")
else:
    _cors_origins = [o.strip() for o in _allowed_origins.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization", "X-Requested-With"],
)


# Security headers middleware
@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    """Добавление защитных HTTP заголовков"""
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    # HSTS — включать только при наличии HTTPS
    if request.url.scheme == "https":
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response


@app.on_event("startup")
async def startup_event():
    """Инициализация при запуске"""
    print(f"[INFO] Запуск {settings.app_name} v{settings.app_version}")
    init_db()
    print("[OK] База данных инициализирована")

    # Миграция таблицы user_permissions: переименование колонок
    from database import engine, UserPermission
    from sqlalchemy import inspect, text
    try:
        insp = inspect(engine)
        if insp.has_table("user_permissions"):
            columns = [c["name"] for c in insp.get_columns("user_permissions")]
            if "permission_type" in columns and "permission_name" not in columns:
                with engine.begin() as conn:
                    conn.execute(text("ALTER TABLE user_permissions RENAME COLUMN permission_type TO permission_name"))
                    logger.info("Migrated user_permissions: permission_type -> permission_name")
            if "target" in columns:
                with engine.begin() as conn:
                    conn.execute(text("ALTER TABLE user_permissions DROP COLUMN target"))
                    logger.info("Migrated user_permissions: dropped column target")
    except Exception as e:
        logger.warning(f"user_permissions migration note: {e}")

    # Seed дефолтных прав
    from database import SessionLocal
    db = SessionLocal()
    try:
        seed_permissions(db)
        print("[OK] Permissions seeded")

        # Инициализация Telegram и Email сервисов из настроек БД
        try:
            messenger_settings = {}
            ms_rows = db.query(MessengerSetting).all()
            for row in ms_rows:
                messenger_settings[row.setting_key] = row.setting_value or ""

            tg_service = get_telegram_service()
            tg_service.configure(messenger_settings)
            print(f"[OK] Telegram: bot={'да' if tg_service.bot_available else 'нет'}, mtproto={'да' if tg_service.mtproto_available else 'нет'}")

            email_svc = get_email_service()
            email_svc.configure(messenger_settings)
            print(f"[OK] Email: {'настроен' if email_svc.available else 'не настроен'}")
        except Exception as e:
            logger.warning(f"Messenger services init: {e}")

        # Seed дефолтных скриптов мессенджера
        try:
            _seed_default_messenger_scripts(db)
        except Exception as e:
            logger.warning(f"Messenger scripts seed: {e}")
    finally:
        db.close()


def _seed_default_messenger_scripts(db: Session):
    """Заполнить дефолтные скрипты если таблица пуста"""
    existing = db.query(MessengerScript).count()
    if existing > 0:
        return

    defaults = [
        MessengerScript(
            script_type="project_start",
            project_type=None,
            stage_name=None,
            message_template=(
                "Всем доброго дня!\n"
                "Приветствую в рабочем чате!\n"
                "Прикрепляем памятку, это кратко описанные основные пункты договора простым языком, "
                "которые важны в первую очередь. Пишите сюда любые вопросы, мы будем сами распределять "
                "кто на какие вопросы отвечает ввиду своих внутренних распорядков и компетенций.\n"
                "{senior_manager} отвечает за организационную часть проекта. "
                "{sdp} у нас старший дизайнер - она будет вести с вами весь проект. "
                "{director} руководитель студии, с ним вы заключали договор."
            ),
            use_auto_deadline=False,
            is_enabled=True,
            sort_order=0,
        ),
        MessengerScript(
            script_type="stage_complete",
            project_type=None,
            stage_name=None,
            message_template=(
                "Добрый день!\n"
                "Прошу рассмотреть {stage_name}, собрать для нас вопросы/правки/замечания "
                "и при необходимости назначить время-день для обсуждения (видео-созвон), когда Вам удобно.\n"
                "Рассмотреть необходимо до {deadline} или напишите, пожалуйста, "
                "срок необходимый Вам для согласования этапа."
            ),
            use_auto_deadline=True,
            is_enabled=True,
            sort_order=1,
        ),
        MessengerScript(
            script_type="project_end",
            project_type=None,
            stage_name=None,
            message_template=(
                "Добрый день!\n"
                "Благодарим за сотрудничество, надеемся вы остались довольны нашей работой.\n"
                "Желаем успехов в ремонте!\n"
                "Прикрепляем памятку о дальнейших этапах реализации проекта. "
                "Обращайтесь если будет необходима помощь.\n"
                "Очень просим Вас оставить отзыв о нашей работе: "
                "https://yandex.ru/maps/org/festival_color/21058411145/"
            ),
            use_auto_deadline=False,
            is_enabled=True,
            sort_order=2,
        ),
    ]

    for script in defaults:
        db.add(script)
    db.commit()
    print(f"[OK] Создано {len(defaults)} дефолтных скриптов мессенджера")


@app.get("/")
async def root():
    """Корневой эндпоинт"""
    return {
        "app": settings.app_name,
        "version": settings.app_version,
        "status": "running"
    }


@app.get("/health")
async def health_check():
    """Проверка здоровья сервиса"""
    return {"status": "healthy"}


@app.get("/api/version")
async def get_app_version():
    """Получить текущую версию серверного приложения для сверки клиентами"""
    return {
        "version": settings.app_version,
        "app": settings.app_name
    }


# =========================
# ГЛОБАЛЬНЫЙ ПОИСК
# =========================

@app.get("/api/search")
async def global_search(
    q: str,
    limit: int = 50,
    entity_types: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Полнотекстовый поиск по клиентам, договорам и CRM карточкам.
    entity_types — через запятую: clients,contracts,crm_cards
    """
    if not q or len(q.strip()) < 2:
        return {"results": [], "total": 0, "query": q}

    query_text = q.strip()
    search_pattern = f"%{query_text}%"
    results = []
    types_filter = entity_types.split(",") if entity_types else ["clients", "contracts", "crm_cards"]

    # Поиск по клиентам
    if "clients" in types_filter:
        from database import Client as ClientModel
        clients = db.query(ClientModel).filter(
            or_(
                ClientModel.full_name.ilike(search_pattern),
                ClientModel.phone.ilike(search_pattern),
                ClientModel.email.ilike(search_pattern),
                ClientModel.organization_name.ilike(search_pattern)
            )
        ).limit(limit).all()
        for c in clients:
            results.append({
                "type": "client",
                "id": c.id,
                "title": c.full_name or "",
                "subtitle": c.phone or c.email or "",
            })

    # Поиск по договорам
    if "contracts" in types_filter:
        from database import Contract as ContractModel
        contracts = db.query(ContractModel).filter(
            or_(
                ContractModel.contract_number.ilike(search_pattern),
                ContractModel.address.ilike(search_pattern),
            )
        ).limit(limit).all()
        for ct in contracts:
            results.append({
                "type": "contract",
                "id": ct.id,
                "title": ct.contract_number or "",
                "subtitle": ct.address or "",
            })

    # Поиск по CRM карточкам (через join с договором)
    if "crm_cards" in types_filter:
        from database import CRMCard as CRMCardModel, Contract as ContractModel2
        cards = db.query(CRMCardModel).join(
            ContractModel2, CRMCardModel.contract_id == ContractModel2.id
        ).filter(
            or_(
                ContractModel2.address.ilike(search_pattern),
                ContractModel2.contract_number.ilike(search_pattern),
            )
        ).limit(limit).all()
        for card in cards:
            contract = db.query(ContractModel2).filter(ContractModel2.id == card.contract_id).first()
            results.append({
                "type": "crm_card",
                "id": card.id,
                "title": f"Проект #{card.id}",
                "subtitle": f"{contract.address if contract else ''} ({card.column_name})",
            })

    return {
        "results": results[:limit],
        "total": len(results),
        "query": q
    }


# =========================
# АУТЕНТИФИКАЦИЯ
# =========================

@app.post("/api/auth/login")
async def login(
    request: Request,
    form_data: OAuth2PasswordRequestForm = Depends(),
    db: Session = Depends(get_db)
):
    """Вход в систему - возвращает access_token и refresh_token"""
    # Brute-force защита: in-memory + персистентная через ActivityLog
    client_ip = request.client.host if request.client else "unknown"
    now = datetime.utcnow()
    cutoff = now - timedelta(minutes=_LOGIN_BLOCK_MINUTES)

    # In-memory счётчик (быстрый)
    _login_attempts[client_ip] = [
        t for t in _login_attempts[client_ip] if t > cutoff
    ]

    # Дополнительно проверяем в БД (персистентно, переживает рестарт)
    db_failed_count = db.query(ActivityLog).filter(
        ActivityLog.action_type == "login_failed",
        ActivityLog.entity_type == "auth",
        ActivityLog.action_date > cutoff
    ).count()

    total_attempts = max(len(_login_attempts[client_ip]), db_failed_count)
    if total_attempts >= _LOGIN_MAX_ATTEMPTS:
        logger.warning(f"Brute-force заблокирован: IP={client_ip}, попыток={total_attempts}")
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail=f"Слишком много попыток входа. Повторите через {_LOGIN_BLOCK_MINUTES} минут",
        )

    # Поиск сотрудника
    employee = db.query(Employee).filter(Employee.login == form_data.username).first()
    if not employee or not verify_password(form_data.password, employee.password_hash):
        _login_attempts[client_ip].append(now)
        # Логируем неудачную попытку в БД (персистентно)
        failed_log = ActivityLog(
            employee_id=employee.id if employee else 0,
            action_type="login_failed",
            entity_type="auth",
            entity_id=0
        )
        db.add(failed_log)
        db.commit()
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Неверный логин или пароль",
            headers={"WWW-Authenticate": "Bearer"},
        )
    # Успешный вход — очищаем счётчик
    _login_attempts.pop(client_ip, None)

    # Создание токенов
    access_token = create_access_token(data={"sub": str(employee.id)})
    refresh_token = create_refresh_token(data={"sub": str(employee.id)})

    # Обновление статуса
    employee.last_login = datetime.utcnow()
    employee.is_online = True
    employee.current_session_token = access_token

    # Создание сессии с refresh_token
    session = UserSession(
        employee_id=employee.id,
        session_token=access_token,
        refresh_token=refresh_token,
        login_time=datetime.utcnow()
    )
    db.add(session)

    # Лог активности
    log = ActivityLog(
        employee_id=employee.id,
        session_id=session.id,
        action_type="login",
        entity_type="auth",
        entity_id=employee.id
    )
    db.add(log)

    db.commit()

    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "employee_id": employee.id,
        "full_name": employee.full_name,
        "role": employee.role or "",
        "position": employee.position or "",
        "secondary_position": employee.secondary_position or "",
        "department": employee.department or ""
    }


@app.post("/api/auth/refresh")
async def refresh_token(
    refresh_token: str,
    db: Session = Depends(get_db)
):
    """Обновление access_token с помощью refresh_token"""
    # Проверяем refresh_token
    payload = verify_refresh_token(refresh_token)
    employee_id = payload.get("sub")

    if not employee_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Неверный refresh токен",
            headers={"WWW-Authenticate": "Bearer"},
        )

    # Проверяем, что сессия с этим refresh_token существует и активна
    session = db.query(UserSession).filter(
        UserSession.refresh_token == refresh_token,
        UserSession.is_active == True
    ).first()

    if not session:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Сессия не найдена или истекла",
            headers={"WWW-Authenticate": "Bearer"},
        )

    # Получаем сотрудника
    employee = db.query(Employee).filter(Employee.id == int(employee_id)).first()
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Пользователь не найден",
            headers={"WWW-Authenticate": "Bearer"},
        )

    # Создаем новый access_token
    new_access_token = create_access_token(data={"sub": str(employee.id)})

    # Обновляем сессию
    session.session_token = new_access_token
    employee.current_session_token = new_access_token
    employee.last_activity = datetime.utcnow()

    db.commit()

    return {
        "access_token": new_access_token,
        "token_type": "bearer",
        "employee_id": employee.id,
        "full_name": employee.full_name
    }


@app.post("/api/auth/logout")
async def logout(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Выход из системы"""
    current_user.is_online = False
    current_user.current_session_token = None

    # Закрываем сессию
    session = db.query(UserSession).filter(
        UserSession.employee_id == current_user.id,
        UserSession.is_active == True
    ).first()
    if session:
        session.is_active = False
        session.logout_time = datetime.utcnow()

    # Лог
    log = ActivityLog(
        employee_id=current_user.id,
        action_type="logout",
        entity_type="auth",
        entity_id=current_user.id
    )
    db.add(log)

    db.commit()

    return {"message": "Успешный выход"}


@app.get("/api/auth/me", response_model=EmployeeResponse)
async def get_me(current_user: Employee = Depends(get_current_user)):
    """Получить информацию о текущем пользователе"""
    return current_user


# =========================
# СОТРУДНИКИ
# =========================

@app.get("/api/employees", response_model=List[EmployeeResponse])
async def get_employees(
    skip: int = 0,
    limit: int = 100,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить список сотрудников"""
    employees = db.query(Employee).offset(skip).limit(limit).all()
    return employees


@app.get("/api/employees/{employee_id}", response_model=EmployeeResponse)
async def get_employee(
    employee_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить сотрудника по ID"""
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")
    return employee


@app.post("/api/employees", response_model=EmployeeResponse, status_code=201)
async def create_employee(
    employee_data: EmployeeCreate,
    current_user: Employee = Depends(require_permission("employees.create")),
    db: Session = Depends(get_db)
):
    """Создать нового сотрудника"""

    # Проверка уникальности логина
    existing = db.query(Employee).filter(Employee.login == employee_data.login).first()
    if existing:
        raise HTTPException(status_code=400, detail="Логин уже занят")

    # Создание
    employee = Employee(
        **employee_data.model_dump(exclude={'password'}),
        password_hash=get_password_hash(employee_data.password)
    )
    db.add(employee)
    db.commit()
    db.refresh(employee)

    # Лог
    log = ActivityLog(
        employee_id=current_user.id,
        action_type="create",
        entity_type="employee",
        entity_id=employee.id
    )
    db.add(log)
    db.commit()

    return employee


@app.put("/api/employees/{employee_id}", response_model=EmployeeResponse)
async def update_employee(
    employee_id: int,
    employee_data: EmployeeUpdate,
    current_user: Employee = Depends(require_permission("employees.update")),
    db: Session = Depends(get_db)
):
    """Обновить сотрудника"""

    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")

    # IDOR защита: Старший менеджер не может менять руководителей и других старших менеджеров
    protected_roles = ['Руководитель студии', 'admin', 'director']
    if current_user.role == 'Старший менеджер проектов':
        if employee.role in protected_roles:
            raise HTTPException(status_code=403, detail="Недостаточно прав для изменения этого сотрудника")
        # Запрет повышения роли до руководителя
        update_data_check = employee_data.model_dump(exclude_unset=True)
        if 'role' in update_data_check and update_data_check['role'] in protected_roles:
            raise HTTPException(status_code=403, detail="Недостаточно прав для назначения этой роли")

    # Обновление полей
    update_data = employee_data.model_dump(exclude_unset=True)

    # Если обновляется пароль
    if 'password' in update_data:
        update_data['password_hash'] = get_password_hash(update_data.pop('password'))

    for field, value in update_data.items():
        setattr(employee, field, value)

    employee.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(employee)

    # Лог
    log = ActivityLog(
        employee_id=current_user.id,
        action_type="update",
        entity_type="employee",
        entity_id=employee.id
    )
    db.add(log)
    db.commit()

    return employee


@app.delete("/api/employees/{employee_id}")
async def delete_employee(
    employee_id: int,
    current_user: Employee = Depends(require_permission("employees.delete")),
    db: Session = Depends(get_db)
):
    """Удалить сотрудника"""

    # Нельзя удалить самого себя
    if employee_id == current_user.id:
        raise HTTPException(status_code=400, detail="Нельзя удалить самого себя")

    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")

    # Каскадное удаление связанных записей
    session_ids = [s.id for s in db.query(UserSession).filter(UserSession.employee_id == employee_id).all()]
    if session_ids:
        db.query(ConcurrentEdit).filter(ConcurrentEdit.session_id.in_(session_ids)).delete(synchronize_session=False)
    db.query(ConcurrentEdit).filter(ConcurrentEdit.employee_id == employee_id).delete(synchronize_session=False)
    db.query(ActivityLog).filter(ActivityLog.employee_id == employee_id).delete(synchronize_session=False)
    if session_ids:
        db.query(ActivityLog).filter(ActivityLog.session_id.in_(session_ids)).delete(synchronize_session=False)

    # Удаляем stage_executors, payments, salaries, action_history
    db.query(StageExecutor).filter(StageExecutor.executor_id == employee_id).delete(synchronize_session=False)
    db.query(StageExecutor).filter(StageExecutor.assigned_by == employee_id).delete(synchronize_session=False)
    db.query(Payment).filter(Payment.employee_id == employee_id).delete(synchronize_session=False)
    db.query(Payment).filter(Payment.paid_by == employee_id).update({"paid_by": None}, synchronize_session=False)
    db.query(Salary).filter(Salary.employee_id == employee_id).delete(synchronize_session=False)
    db.query(ActionHistory).filter(ActionHistory.user_id == employee_id).delete(synchronize_session=False)

    # Обнуляем FK ссылки в crm_cards и supervision_cards
    for col in ['senior_manager_id', 'sdp_id', 'gap_id', 'manager_id', 'surveyor_id']:
        db.query(CRMCard).filter(getattr(CRMCard, col) == employee_id).update({col: None}, synchronize_session=False)
    for col in ['senior_manager_id', 'dan_id']:
        db.query(SupervisionCard).filter(getattr(SupervisionCard, col) == employee_id).update({col: None}, synchronize_session=False)

    # ORM cascade удалит: user_sessions, user_permissions, notifications
    db.delete(employee)
    db.commit()

    return {"status": "success", "message": "Сотрудник удален"}


# =========================
# ПРАВА ДОСТУПА (PERMISSIONS)
# =========================

@app.get("/api/permissions/definitions")
async def get_permission_definitions(
    current_user: Employee = Depends(get_current_user),
):
    """Получить список всех доступных прав с описаниями"""
    return [
        {"name": name, "description": desc}
        for name, desc in PERMISSION_NAMES.items()
    ]


@app.get("/api/permissions/{employee_id}", response_model=PermissionResponse)
async def get_permissions(
    employee_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить права сотрудника"""
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")

    perms = get_employee_permissions(employee_id, db)
    is_superuser = employee.role in SUPERUSER_ROLES
    return PermissionResponse(employee_id=employee_id, permissions=perms, is_superuser=is_superuser)


@app.put("/api/permissions/{employee_id}", response_model=PermissionResponse)
async def update_permissions(
    employee_id: int,
    request: PermissionSetRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Установить права сотрудника (полная замена)"""
    # Только Руководитель / admin / director могут менять права
    if current_user.role not in SUPERUSER_ROLES and current_user.role != 'Руководитель студии':
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")

    # Нельзя менять права superuser
    if employee.role in SUPERUSER_ROLES:
        raise HTTPException(status_code=400, detail="Нельзя менять права системного пользователя")

    # IDOR: Старший менеджер не может менять Руководителя (но мы уже ограничили до Руководителя)

    # Валидация имён прав
    invalid = [p for p in request.permissions if p not in PERMISSION_NAMES]
    if invalid:
        raise HTTPException(status_code=400, detail=f"Неизвестные права: {', '.join(invalid)}")

    set_employee_permissions(employee_id, request.permissions, current_user.id, db)
    perms = get_employee_permissions(employee_id, db)
    return PermissionResponse(employee_id=employee_id, permissions=perms)


@app.post("/api/permissions/{employee_id}/reset-to-defaults", response_model=PermissionResponse)
async def reset_permissions_to_defaults(
    employee_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Сбросить права до дефолтных по роли"""
    if current_user.role not in SUPERUSER_ROLES and current_user.role != 'Руководитель студии':
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")

    if employee.role in SUPERUSER_ROLES:
        raise HTTPException(status_code=400, detail="Нельзя сбросить права системного пользователя")

    reset_to_defaults(employee_id, db)
    perms = get_employee_permissions(employee_id, db)
    return PermissionResponse(employee_id=employee_id, permissions=perms)


# =========================
# КЛИЕНТЫ
# =========================

@app.get("/api/clients", response_model=List[ClientResponse])
async def get_clients(
    skip: int = 0,
    limit: int = 100,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить список клиентов"""
    clients = db.query(Client).offset(skip).limit(limit).all()
    return clients


@app.get("/api/clients/{client_id}", response_model=ClientResponse)
async def get_client(
    client_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить клиента по ID"""
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Клиент не найден")
    return client


@app.post("/api/clients", response_model=ClientResponse)
async def create_client(
    client_data: ClientCreate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Создать нового клиента"""
    client = Client(**client_data.model_dump())
    db.add(client)
    db.commit()
    db.refresh(client)

    # Лог
    log = ActivityLog(
        employee_id=current_user.id,
        action_type="create",
        entity_type="client",
        entity_id=client.id
    )
    db.add(log)
    db.commit()

    return client


@app.put("/api/clients/{client_id}", response_model=ClientResponse)
async def update_client(
    client_id: int,
    client_data: ClientUpdate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить клиента"""
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Клиент не найден")

    # Обновление полей
    for field, value in client_data.model_dump(exclude_unset=True).items():
        setattr(client, field, value)

    client.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(client)

    # Лог
    log = ActivityLog(
        employee_id=current_user.id,
        action_type="update",
        entity_type="client",
        entity_id=client.id
    )
    db.add(log)
    db.commit()

    return client


@app.delete("/api/clients/{client_id}")
async def delete_client(
    client_id: int,
    current_user: Employee = Depends(require_permission("clients.delete")),
    db: Session = Depends(get_db)
):
    """Удалить клиента и все связанные данные"""
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Клиент не найден")

    try:
        # Сначала удаляем все связанные договоры (каскадно удалит карточки, платежи и т.д.)
        contracts = db.query(Contract).filter(Contract.client_id == client_id).all()
        for contract in contracts:
            # Удаляем timeline записи
            db.query(ProjectTimelineEntry).filter(ProjectTimelineEntry.contract_id == contract.id).delete()
            # Удаляем CRM карточки и связанные данные
            crm_cards = db.query(CRMCard).filter(CRMCard.contract_id == contract.id).all()
            for card in crm_cards:
                db.query(StageExecutor).filter(StageExecutor.crm_card_id == card.id).delete()
                db.query(Payment).filter(Payment.crm_card_id == card.id).delete()
                db.delete(card)
            # Удаляем карточки надзора и связанные данные
            supervision_cards = db.query(SupervisionCard).filter(SupervisionCard.contract_id == contract.id).all()
            for card in supervision_cards:
                db.query(SupervisionTimelineEntry).filter(SupervisionTimelineEntry.supervision_card_id == card.id).delete()
                db.query(SupervisionProjectHistory).filter(SupervisionProjectHistory.supervision_card_id == card.id).delete()
                db.query(Payment).filter(Payment.supervision_card_id == card.id).delete()
                db.delete(card)
            # Удаляем оставшиеся платежи
            db.query(Payment).filter(Payment.contract_id == contract.id).delete()
            # Удаляем файлы
            db.query(ProjectFile).filter(ProjectFile.contract_id == contract.id).delete()
            # Удаляем сам договор
            db.delete(contract)

        # Лог перед удалением
        log = ActivityLog(
            employee_id=current_user.id,
            action_type="delete",
            entity_type="client",
            entity_id=client_id
        )
        db.add(log)

        db.delete(client)
        db.commit()

        return {"status": "success", "message": "Клиент и все связанные данные удалены"}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка удаления клиента: {str(e)}")


# =========================
# ДОГОВОРЫ
# =========================

@app.get("/api/contracts", response_model=List[ContractResponse])
async def get_contracts(
    skip: int = 0,
    limit: int = 100,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить список договоров"""
    contracts = db.query(Contract).offset(skip).limit(limit).all()
    return contracts


@app.get("/api/contracts/{contract_id}", response_model=ContractResponse)
async def get_contract(
    contract_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить договор по ID"""
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")
    return contract


@app.post("/api/contracts", response_model=ContractResponse)
async def create_contract(
    contract_data: ContractCreate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Создать новый договор"""
    # Проверяем существование клиента
    client = db.query(Client).filter(Client.id == contract_data.client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Клиент не найден")

    try:
        contract = Contract(**contract_data.model_dump())
        db.add(contract)
        db.commit()
        db.refresh(contract)

        # Автоматически создаем CRM карточку для нового договора
        # (кроме Авторского надзора - для него создается SupervisionCard)
        if contract.project_type != 'Авторский надзор':
            crm_card = CRMCard(
                contract_id=contract.id,
                column_name='Новый заказ',
                manager_id=current_user.id
            )
            db.add(crm_card)
            db.commit()
            print(f"[OK] Создана CRM карточка для договора {contract.id}")

        # Лог
        log = ActivityLog(
            employee_id=current_user.id,
            action_type="create",
            entity_type="contract",
            entity_id=contract.id
        )
        db.add(log)
        db.commit()

        return contract

    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Договор с таким номером уже существует")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка создания договора: {str(e)}")


@app.put("/api/contracts/{contract_id}", response_model=ContractResponse)
async def update_contract(
    contract_id: int,
    contract_data: ContractUpdate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить договор"""
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")

    # Проверяем, изменяется ли статус на "АВТОРСКИЙ НАДЗОР"
    old_status = contract.status
    update_data = contract_data.model_dump(exclude_unset=True)
    new_status = update_data.get('status')
    need_supervision_card = (
        new_status == 'АВТОРСКИЙ НАДЗОР' and
        old_status != 'АВТОРСКИЙ НАДЗОР'
    )

    # Обновление полей
    for field, value in update_data.items():
        setattr(contract, field, value)

    contract.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(contract)

    # BUG #2 FIX: Автоматическое создание карточки надзора
    if need_supervision_card:
        existing_supervision = db.query(SupervisionCard).filter(
            SupervisionCard.contract_id == contract_id
        ).first()
        if not existing_supervision:
            supervision_card = SupervisionCard(
                contract_id=contract_id,
                column_name='Новый заказ',
                created_at=datetime.utcnow()
            )
            db.add(supervision_card)
            db.commit()
            print(f"[OK] BUG #2 FIX: Автоматически создана карточка надзора для договора {contract_id}")

    # Лог
    log = ActivityLog(
        employee_id=current_user.id,
        action_type="update",
        entity_type="contract",
        entity_id=contract.id
    )
    db.add(log)
    db.commit()

    return contract


@app.patch("/api/contracts/{contract_id}/files")
async def update_contract_files(
    contract_id: int,
    files_data: ContractFilesUpdate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Обновить файлы договора (замер, референсы, фотофиксация)

    Этот endpoint используется для обновления:
    - measurement_image_link - ссылка на изображение замера
    - measurement_file_name - имя файла замера
    - measurement_yandex_path - путь к файлу на Яндекс.Диске
    - measurement_date - дата замера
    - contract_file_link - ссылка на договор
    - tech_task_link - ссылка на техническое задание
    """
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")

    # Обновляем только переданные поля
    for field, value in files_data.model_dump(exclude_unset=True).items():
        setattr(contract, field, value)

    contract.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(contract)

    # Лог
    log = ActivityLog(
        employee_id=current_user.id,
        action_type="update_files",
        entity_type="contract",
        entity_id=contract.id
    )
    db.add(log)
    db.commit()

    return {
        'id': contract.id,
        'measurement_image_link': contract.measurement_image_link,
        'measurement_file_name': contract.measurement_file_name,
        'measurement_yandex_path': contract.measurement_yandex_path,
        'measurement_date': contract.measurement_date,
        'contract_file_link': contract.contract_file_link,
        'tech_task_link': contract.tech_task_link,
        'updated_at': contract.updated_at.isoformat() if contract.updated_at else None
    }


@app.delete("/api/contracts/{contract_id}")
async def delete_contract(
    contract_id: int,
    current_user: Employee = Depends(require_permission("contracts.delete")),
    db: Session = Depends(get_db)
):
    """Удалить договор и все связанные данные"""
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")

    try:
        # Удаляем timeline записи проекта
        db.query(ProjectTimelineEntry).filter(ProjectTimelineEntry.contract_id == contract_id).delete()

        # Удаляем связанные CRM карточки
        crm_cards = db.query(CRMCard).filter(CRMCard.contract_id == contract_id).all()
        for card in crm_cards:
            # Удаляем связанные stage_executors
            db.query(StageExecutor).filter(StageExecutor.crm_card_id == card.id).delete()
            # Удаляем платежи привязанные к карточке
            db.query(Payment).filter(Payment.crm_card_id == card.id).delete()
            # Удаляем саму карточку
            db.delete(card)

        # Удаляем связанные SupervisionCard
        supervision_cards = db.query(SupervisionCard).filter(SupervisionCard.contract_id == contract_id).all()
        for card in supervision_cards:
            # Удаляем timeline записи надзора
            db.query(SupervisionTimelineEntry).filter(SupervisionTimelineEntry.supervision_card_id == card.id).delete()
            # Удаляем связанную историю
            db.query(SupervisionProjectHistory).filter(SupervisionProjectHistory.supervision_card_id == card.id).delete()
            # Удаляем платежи привязанные к карточке надзора
            db.query(Payment).filter(Payment.supervision_card_id == card.id).delete()
            db.delete(card)

        # Удаляем оставшиеся платежи по договору
        db.query(Payment).filter(Payment.contract_id == contract_id).delete()

        # Удаляем связанные файлы проекта
        db.query(ProjectFile).filter(ProjectFile.contract_id == contract_id).delete()

        # Лог перед удалением
        log = ActivityLog(
            employee_id=current_user.id,
            action_type="delete",
            entity_type="contract",
            entity_id=contract_id
        )
        db.add(log)

        db.delete(contract)
        db.commit()

        return {"status": "success", "message": "Договор и все связанные данные удалены"}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка удаления договора: {str(e)}")


# =========================
# СИНХРОНИЗАЦИЯ
# =========================

@app.post("/api/sync", response_model=SyncResponse)
async def sync_data(
    sync_request: SyncRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Синхронизация данных
    Возвращает все изменения после указанного timestamp
    """
    response = SyncResponse(timestamp=datetime.utcnow())

    # Клиенты
    if 'clients' in sync_request.entity_types:
        clients = db.query(Client).filter(
            Client.updated_at > sync_request.last_sync_timestamp
        ).all()
        response.clients = clients

    # Договоры
    if 'contracts' in sync_request.entity_types:
        contracts = db.query(Contract).filter(
            Contract.updated_at > sync_request.last_sync_timestamp
        ).all()
        response.contracts = contracts

    # Сотрудники
    if 'employees' in sync_request.entity_types:
        employees = db.query(Employee).filter(
            Employee.updated_at > sync_request.last_sync_timestamp
        ).all()
        response.employees = employees

    # Уведомления
    notifications = db.query(Notification).filter(
        Notification.employee_id == current_user.id,
        Notification.created_at > sync_request.last_sync_timestamp
    ).all()
    response.notifications = notifications

    return response


# =========================
# УВЕДОМЛЕНИЯ
# =========================

@app.get("/api/notifications", response_model=List[NotificationResponse])
async def get_notifications(
    unread_only: bool = False,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить уведомления текущего пользователя"""
    query = db.query(Notification).filter(Notification.employee_id == current_user.id)

    if unread_only:
        query = query.filter(Notification.is_read == False)

    notifications = query.order_by(Notification.created_at.desc()).all()
    return notifications


@app.put("/api/notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Отметить уведомление как прочитанное"""
    notification = db.query(Notification).filter(
        Notification.id == notification_id,
        Notification.employee_id == current_user.id
    ).first()

    if not notification:
        raise HTTPException(status_code=404, detail="Уведомление не найдено")

    notification.is_read = True
    notification.read_at = datetime.utcnow()
    db.commit()

    return {"message": "Уведомление прочитано"}


# =========================
# CRM КАРТОЧКИ
# =========================

@app.get("/api/crm/cards")
async def get_crm_cards(
    project_type: str,
    archived: bool = False,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить список CRM карточек по типу проекта

    Args:
        project_type: Тип проекта (Индивидуальный/Шаблонный)
        archived: Если True - возвращает архивные карточки (СДАН, РАСТОРГНУТ, АВТОРСКИЙ НАДЗОР)
    """
    try:
        query = db.query(CRMCard).join(
            Contract, CRMCard.contract_id == Contract.id
        ).filter(
            Contract.project_type == project_type
        )

        if archived:
            # Архивные карточки - статус СДАН, РАСТОРГНУТ или АВТОРСКИЙ НАДЗОР
            query = query.filter(
                Contract.status.in_(['СДАН', 'РАСТОРГНУТ', 'АВТОРСКИЙ НАДЗОР'])
            )
        else:
            # Активные карточки - статус НЕ в архивных
            query = query.filter(
                or_(
                    Contract.status == None,
                    Contract.status == '',
                    ~Contract.status.in_(['СДАН', 'РАСТОРГНУТ', 'АВТОРСКИЙ НАДЗОР'])
                )
            )

        cards = query.order_by(
            CRMCard.order_position,
            CRMCard.id
        ).all()

        # Batch-load all stage executors for all cards to avoid N+1 queries
        card_ids = [card.id for card in cards]
        all_executors = db.query(StageExecutor).filter(StageExecutor.crm_card_id.in_(card_ids)).all() if card_ids else []
        executors_by_card = {}
        for se in all_executors:
            if se.crm_card_id not in executors_by_card:
                executors_by_card[se.crm_card_id] = []
            executors_by_card[se.crm_card_id].append(se)

        # Batch-load executor Employee objects for all stage executors
        executor_employee_ids = list(set(se.executor_id for se in all_executors if se.executor_id))
        executor_employees_map = {e.id: e for e in db.query(Employee).filter(Employee.id.in_(executor_employee_ids)).all()} if executor_employee_ids else {}

        result = []
        for card in cards:
            contract = card.contract
            senior_manager_name = card.senior_manager.full_name if card.senior_manager else None
            sdp_name = card.sdp.full_name if card.sdp else None
            gap_name = card.gap.full_name if card.gap else None
            manager_name = card.manager.full_name if card.manager else None
            surveyor_name = card.surveyor.full_name if card.surveyor else None

            # ИСПРАВЛЕНИЕ 06.02.2026: Добавлен поиск по '3д визуализация' для шаблонных проектов (#10)
            # Use batch-loaded executors instead of per-card queries
            card_executors = executors_by_card.get(card.id, [])

            # Find designer executor: stage_name contains 'концепция' or 'визуализация', latest by id
            designer_candidates = [
                e for e in card_executors
                if 'концепция' in (e.stage_name or '').lower() or 'визуализация' in (e.stage_name or '').lower()
            ]
            designer_executor = max(designer_candidates, key=lambda e: e.id) if designer_candidates else None

            # Find draftsman executor: stage_name contains 'чертежи' or 'планировочные', latest by id
            draftsman_candidates = [
                e for e in card_executors
                if 'чертежи' in (e.stage_name or '').lower() or 'планировочные' in (e.stage_name or '').lower()
            ]
            draftsman_executor = max(draftsman_candidates, key=lambda e: e.id) if draftsman_candidates else None

            # Get executor names from batch-loaded employees map
            designer_employee = executor_employees_map.get(designer_executor.executor_id) if designer_executor else None
            draftsman_employee = executor_employees_map.get(draftsman_executor.executor_id) if draftsman_executor else None

            card_data = {
                'id': card.id,
                'contract_id': card.contract_id,
                'column_name': card.column_name,
                'deadline': str(card.deadline) if card.deadline else None,
                'tags': card.tags,
                'is_approved': card.is_approved,
                'approval_deadline': str(card.approval_deadline) if card.approval_deadline else None,
                'approval_stages': json.loads(card.approval_stages) if card.approval_stages else None,
                'project_data_link': card.project_data_link,
                'tech_task_file': card.tech_task_file,
                'tech_task_date': str(card.tech_task_date) if card.tech_task_date else None,
                'survey_date': str(card.survey_date) if card.survey_date else None,
                'senior_manager_id': card.senior_manager_id,
                'sdp_id': card.sdp_id,
                'gap_id': card.gap_id,
                'manager_id': card.manager_id,
                'surveyor_id': card.surveyor_id,
                'senior_manager_name': senior_manager_name,
                'sdp_name': sdp_name,
                'gap_name': gap_name,
                'manager_name': manager_name,
                'surveyor_name': surveyor_name,
                'contract_number': contract.contract_number,
                'address': contract.address,
                'area': contract.area,
                'city': contract.city,
                'agent_type': contract.agent_type,
                'project_type': contract.project_type,
                'project_subtype': contract.project_subtype if hasattr(contract, 'project_subtype') else None,
                'floors': contract.floors if hasattr(contract, 'floors') else 1,
                'contract_period': contract.contract_period,
                'contract_status': contract.status,
                # Поля ТЗ и замера из contracts
                'tech_task_link': contract.tech_task_link,
                'tech_task_file_name': contract.tech_task_file_name,
                'tech_task_yandex_path': contract.tech_task_yandex_path,
                'measurement_image_link': contract.measurement_image_link,
                'measurement_file_name': contract.measurement_file_name,
                'measurement_yandex_path': contract.measurement_yandex_path,
                'measurement_date': str(contract.measurement_date) if contract.measurement_date else None,
                'designer_name': designer_employee.full_name if designer_employee else None,
                'designer_completed': designer_executor.completed if designer_executor else False,
                'designer_deadline': str(designer_executor.deadline) if designer_executor and designer_executor.deadline else None,
                'draftsman_name': draftsman_employee.full_name if draftsman_employee else None,
                'draftsman_completed': draftsman_executor.completed if draftsman_executor else False,
                'draftsman_deadline': str(draftsman_executor.deadline) if draftsman_executor and draftsman_executor.deadline else None,
                'order_position': card.order_position,
                'created_at': card.created_at.isoformat() if card.created_at else None,
                'updated_at': card.updated_at.isoformat() if card.updated_at else None,
            }
            result.append(card_data)

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения CRM карточек: {str(e)}")


@app.get("/api/crm/cards/{card_id}")
async def get_crm_card(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить одну CRM карточку"""
    try:
        card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="CRM карточка не найдена")

        # Контракт — для полей project_subtype, agent_type, address и т.д.
        contract = db.query(Contract).filter(Contract.id == card.contract_id).first() if card.contract_id else None

        # Имена сотрудников
        def _emp_name(emp_id):
            if not emp_id:
                return None
            emp = db.query(Employee).filter(Employee.id == emp_id).first()
            return emp.full_name if emp else None

        stage_executors = db.query(StageExecutor).filter(
            StageExecutor.crm_card_id == card_id
        ).all()

        executor_data = []
        for se in stage_executors:
            executor_data.append({
                'id': se.id,
                'stage_name': se.stage_name,
                'executor_id': se.executor_id,
                'executor_name': se.executor.full_name,
                'assigned_by': se.assigned_by,
                'assigned_date': se.assigned_date.isoformat() if se.assigned_date else None,
                'deadline': str(se.deadline) if se.deadline else None,
                'submitted_date': se.submitted_date.isoformat() if se.submitted_date else None,
                'completed': se.completed,
                'completed_date': se.completed_date.isoformat() if se.completed_date else None,
            })

        result = {
            'id': card.id,
            'contract_id': card.contract_id,
            'column_name': card.column_name,
            'deadline': str(card.deadline) if card.deadline else None,
            'tags': card.tags,
            'is_approved': card.is_approved,
            'senior_manager_id': card.senior_manager_id,
            'sdp_id': card.sdp_id,
            'gap_id': card.gap_id,
            'manager_id': card.manager_id,
            'surveyor_id': card.surveyor_id,
            'senior_manager_name': _emp_name(card.senior_manager_id),
            'sdp_name': _emp_name(card.sdp_id),
            'gap_name': _emp_name(card.gap_id),
            'manager_name': _emp_name(card.manager_id),
            'surveyor_name': _emp_name(card.surveyor_id),
            'approval_deadline': str(card.approval_deadline) if card.approval_deadline else None,
            'approval_stages': json.loads(card.approval_stages) if card.approval_stages else None,
            'project_data_link': card.project_data_link,
            'tech_task_file': card.tech_task_file,
            'tech_task_date': str(card.tech_task_date) if card.tech_task_date else None,
            'survey_date': str(card.survey_date) if card.survey_date else None,
            'order_position': card.order_position,
            'stage_executors': executor_data,
        }

        # Поля из контракта
        if contract:
            result.update({
                'contract_number': contract.contract_number,
                'address': contract.address,
                'area': contract.area,
                'city': contract.city,
                'agent_type': contract.agent_type,
                'project_type': contract.project_type,
                'project_subtype': contract.project_subtype if hasattr(contract, 'project_subtype') else None,
                'floors': contract.floors if hasattr(contract, 'floors') else 1,
                'contract_period': contract.contract_period,
                'contract_status': contract.status,
            })

        return result

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения карточки: {str(e)}")


@app.post("/api/crm/cards", response_model=CRMCardResponse)
async def create_crm_card(
    card_data: CRMCardCreate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Создать новую CRM карточку"""
    try:
        card = CRMCard(**card_data.model_dump())
        db.add(card)
        db.commit()
        db.refresh(card)

        log = ActivityLog(
            employee_id=current_user.id,
            action_type="create",
            entity_type="crm_card",
            entity_id=card.id
        )
        db.add(log)
        db.commit()

        # Явная сериализация для корректного ответа
        return {
            "id": card.id,
            "contract_id": card.contract_id,
            "column_name": card.column_name,
            "deadline": card.deadline,
            "tags": card.tags,
            "is_approved": card.is_approved,
            "approval_deadline": card.approval_deadline,
            "approval_stages": json.loads(card.approval_stages) if card.approval_stages else None,
            "project_data_link": card.project_data_link,
            "tech_task_file": card.tech_task_file,
            "tech_task_date": card.tech_task_date,
            "survey_date": card.survey_date,
            "senior_manager_id": card.senior_manager_id,
            "sdp_id": card.sdp_id,
            "gap_id": card.gap_id,
            "manager_id": card.manager_id,
            "surveyor_id": card.surveyor_id,
            "order_position": card.order_position,
            "created_at": card.created_at,
            "updated_at": card.updated_at
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка создания карточки: {str(e)}")


@app.patch("/api/crm/cards/{card_id}")
async def update_crm_card(
    card_id: int,
    updates: CRMCardUpdate,
    current_user: Employee = Depends(require_permission("crm_cards.update")),
    db: Session = Depends(get_db)
):
    """Обновить CRM карточку"""
    try:
        card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="CRM карточка не найдена")

        update_data = updates.model_dump(exclude_unset=True)

        # Проверяем существование сотрудников перед обновлением FK полей
        employee_fields = ['senior_manager_id', 'sdp_id', 'gap_id', 'manager_id', 'surveyor_id']
        for field in employee_fields:
            if field in update_data and update_data[field] is not None:
                employee_exists = db.query(Employee).filter(Employee.id == update_data[field]).first()
                if not employee_exists:
                    # Удаляем несуществующий ID из обновлений (игнорируем)
                    print(f"[WARN] Employee ID {update_data[field]} not found for field {field}, skipping")
                    del update_data[field]

        for field, value in update_data.items():
            setattr(card, field, value)

        db.commit()
        db.refresh(card)

        return {
            'id': card.id,
            'contract_id': card.contract_id,
            'column_name': card.column_name,
            'deadline': str(card.deadline) if card.deadline else None,
            'tags': card.tags,
            'is_approved': card.is_approved,
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка обновления карточки: {str(e)}")


@app.patch("/api/crm/cards/{card_id}/column")
async def move_crm_card_to_column(
    card_id: int,
    move_request: ColumnMoveRequest,
    current_user: Employee = Depends(require_permission("crm_cards.move")),
    db: Session = Depends(get_db)
):
    """Переместить CRM карточку в другую колонку"""
    try:
        VALID_CRM_COLUMNS = [
            'Новый заказ', 'В ожидании',
            'Стадия 1: планировочные решения', 'Стадия 2: концепция дизайна',
            'Стадия 3: рабочие чертежи', 'Стадия 4: комплектация',
            'Выполненный проект'
        ]
        if move_request.column_name not in VALID_CRM_COLUMNS:
            raise HTTPException(status_code=422, detail=f"Недопустимая колонка: {move_request.column_name}")

        card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="CRM карточка не найдена")

        old_column = card.column_name
        new_column = move_request.column_name

        # Валидация последовательности переходов (руководство может перемещать свободно)
        free_move_roles = ['admin', 'director', 'Руководитель студии', 'Старший менеджер проектов']
        if current_user.role not in free_move_roles:
            CRM_COLUMN_ORDER = [
                'Новый заказ',
                'Стадия 1: планировочные решения', 'Стадия 2: концепция дизайна',
                'Стадия 3: рабочие чертежи', 'Стадия 4: комплектация',
                'Выполненный проект'
            ]
            # "В ожидании" — специальная колонка, можно перемещать туда и обратно
            if old_column != 'В ожидании' and new_column != 'В ожидании':
                old_idx = CRM_COLUMN_ORDER.index(old_column) if old_column in CRM_COLUMN_ORDER else -1
                new_idx = CRM_COLUMN_ORDER.index(new_column) if new_column in CRM_COLUMN_ORDER else -1
                if old_idx >= 0 and new_idx >= 0 and abs(new_idx - old_idx) > 1:
                    raise HTTPException(
                        status_code=422,
                        detail=f"Нельзя перескакивать стадии: {old_column} → {new_column}"
                    )

        card.column_name = new_column

        db.commit()
        db.refresh(card)

        # Хук: автоуведомление в чат при перемещении карточки
        if old_column != new_column:
            if new_column == 'Выполненный проект':
                asyncio.create_task(_trigger_messenger_notification(
                    db, card.id, 'project_end', stage_name=new_column
                ))
            elif 'Стадия' in new_column:
                asyncio.create_task(_trigger_messenger_notification(
                    db, card.id, 'stage_complete', stage_name=old_column
                ))

        return {
            'id': card.id,
            'contract_id': card.contract_id,
            'column_name': card.column_name,
            'old_column_name': old_column,
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка перемещения карточки: {str(e)}")


@app.post("/api/crm/cards/{card_id}/stage-executor")
async def assign_stage_executor(
    card_id: int,
    executor_data: StageExecutorCreate,
    current_user: Employee = Depends(require_permission("crm_cards.assign_executor")),
    db: Session = Depends(get_db)
):
    """Назначить исполнителя на стадию"""
    try:
        card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="CRM карточка не найдена")

        executor = db.query(Employee).filter(Employee.id == executor_data.executor_id).first()
        if not executor:
            raise HTTPException(status_code=404, detail="Исполнитель не найден")

        # Проверка дубликата: тот же исполнитель на ту же стадию
        existing = db.query(StageExecutor).filter(
            StageExecutor.crm_card_id == card_id,
            StageExecutor.stage_name == executor_data.stage_name,
            StageExecutor.executor_id == executor_data.executor_id
        ).first()
        if existing:
            raise HTTPException(
                status_code=409,
                detail=f"Исполнитель {executor.full_name} уже назначен на стадию '{executor_data.stage_name}'"
            )

        stage_executor = StageExecutor(
            crm_card_id=card_id,
            stage_name=executor_data.stage_name,
            executor_id=executor_data.executor_id,
            assigned_by=current_user.id,
            deadline=executor_data.deadline,
            assigned_date=datetime.utcnow()
        )

        db.add(stage_executor)
        db.commit()
        db.refresh(stage_executor)

        return {
            'id': stage_executor.id,
            'crm_card_id': stage_executor.crm_card_id,
            'stage_name': stage_executor.stage_name,
            'executor_id': stage_executor.executor_id,
            'executor_name': executor.full_name,
            'assigned_by': stage_executor.assigned_by,
            'assigned_date': stage_executor.assigned_date.isoformat(),
            'deadline': str(stage_executor.deadline) if stage_executor.deadline else None,
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка назначения исполнителя: {str(e)}")


@app.patch("/api/crm/cards/{card_id}/stage-executor/{stage_name}")
async def complete_stage(
    card_id: int,
    stage_name: str,
    update_data: StageExecutorUpdate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить статус выполнения стадии"""
    try:
        # ИСПРАВЛЕНИЕ 25.01.2026: Сначала ищем точное совпадение
        stage_executor = db.query(StageExecutor).filter(
            StageExecutor.crm_card_id == card_id,
            StageExecutor.stage_name == stage_name
        ).order_by(StageExecutor.id.desc()).first()

        # Fallback: поиск по LIKE если точное совпадение не найдено
        if not stage_executor:
            print(f"[DEBUG] Точное совпадение не найдено для stage_name='{stage_name}', пробуем LIKE")
            stage_executor = db.query(StageExecutor).filter(
                StageExecutor.crm_card_id == card_id,
                StageExecutor.stage_name.ilike(f'%{stage_name}%')
            ).order_by(StageExecutor.id.desc()).first()

        if not stage_executor:
            raise HTTPException(status_code=404, detail=f"Назначение стадии не найдено: card_id={card_id}, stage_name={stage_name}")

        update_dict = update_data.model_dump(exclude_unset=True)

        # Простое обновление всех полей
        for field, value in update_dict.items():
            setattr(stage_executor, field, value)

        if update_data.completed and not update_data.completed_date:
            stage_executor.completed_date = datetime.utcnow()

        db.commit()
        db.refresh(stage_executor)

        return {
            'id': stage_executor.id,
            'crm_card_id': stage_executor.crm_card_id,
            'stage_name': stage_executor.stage_name,
            'executor_id': stage_executor.executor_id,
            'completed': stage_executor.completed,
            'completed_date': stage_executor.completed_date.isoformat() if stage_executor.completed_date else None,
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка обновления стадии: {str(e)}")


@app.delete("/api/crm/cards/{card_id}")
async def delete_crm_card(
    card_id: int,
    current_user: Employee = Depends(require_permission("crm_cards.delete")),
    db: Session = Depends(get_db)
):
    """Удалить CRM карточку"""
    try:
        card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="CRM карточка не найдена")

        # Удаляем связанные stage_executors
        db.query(StageExecutor).filter(StageExecutor.crm_card_id == card_id).delete()

        # Удаляем связанные платежи
        db.query(Payment).filter(Payment.crm_card_id == card_id).delete()

        # Лог перед удалением
        log = ActivityLog(
            employee_id=current_user.id,
            action_type="delete",
            entity_type="crm_card",
            entity_id=card_id
        )
        db.add(log)

        db.delete(card)
        db.commit()

        return {"status": "success", "message": "CRM карточка удалена"}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка удаления карточки: {str(e)}")


@app.delete("/api/crm/stage-executors/{executor_id}")
async def delete_stage_executor(
    executor_id: int,
    current_user: Employee = Depends(require_permission("crm_cards.delete_executor")),
    db: Session = Depends(get_db)
):
    """Удалить назначение исполнителя на стадию"""
    try:
        executor = db.query(StageExecutor).filter(StageExecutor.id == executor_id).first()
        if not executor:
            raise HTTPException(status_code=404, detail="Назначение не найдено")

        # Лог перед удалением
        log = ActivityLog(
            employee_id=current_user.id,
            action_type="delete",
            entity_type="stage_executor",
            entity_id=executor_id
        )
        db.add(log)

        db.delete(executor)
        db.commit()

        return {"status": "success", "message": "Назначение исполнителя удалено"}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка удаления назначения: {str(e)}")


# =========================
# SUPERVISION (Авторский надзор)
# =========================

@app.get("/api/supervision/cards")
async def get_supervision_cards(
    status: str = "active",
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить список карточек авторского надзора"""
    try:
        if status == "active":
            cards = db.query(SupervisionCard).join(
                Contract, SupervisionCard.contract_id == Contract.id
            ).filter(
                Contract.status == 'АВТОРСКИЙ НАДЗОР'
            ).order_by(
                SupervisionCard.id.desc()
            ).all()
        else:
            cards = db.query(SupervisionCard).join(
                Contract, SupervisionCard.contract_id == Contract.id
            ).filter(
                Contract.status.in_(['СДАН', 'РАСТОРГНУТ'])
            ).order_by(
                SupervisionCard.id.desc()
            ).all()

        result = []
        for card in cards:
            contract = card.contract
            senior_manager_name = card.senior_manager.full_name if card.senior_manager else None
            dan_name = card.dan.full_name if card.dan else None

            card_data = {
                'id': card.id,
                'contract_id': card.contract_id,
                'column_name': card.column_name,
                'deadline': str(card.deadline) if card.deadline else None,
                'tags': card.tags,
                'senior_manager_id': card.senior_manager_id,
                'dan_id': card.dan_id,
                'dan_completed': card.dan_completed,
                'is_paused': card.is_paused,
                'pause_reason': card.pause_reason,
                'paused_at': card.paused_at.isoformat() if card.paused_at else None,
                'senior_manager_name': senior_manager_name,
                'dan_name': dan_name,
                'contract_number': contract.contract_number,
                'address': contract.address,
                'area': contract.area,
                'city': contract.city,
                'agent_type': contract.agent_type,
                'contract_status': contract.status,
                'termination_reason': contract.termination_reason if status == "archived" else None,
                'created_at': card.created_at.isoformat() if card.created_at else None,
                'updated_at': card.updated_at.isoformat() if card.updated_at else None,
            }
            result.append(card_data)

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения карточек надзора: {str(e)}")


@app.get("/api/supervision/cards/{card_id}")
async def get_supervision_card(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить одну карточку надзора"""
    try:
        card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

        return {
            'id': card.id,
            'contract_id': card.contract_id,
            'column_name': card.column_name,
            'deadline': str(card.deadline) if card.deadline else None,
            'tags': card.tags,
            'senior_manager_id': card.senior_manager_id,
            'dan_id': card.dan_id,
            'dan_completed': card.dan_completed,
            'is_paused': card.is_paused,
            'pause_reason': card.pause_reason,
            'paused_at': card.paused_at.isoformat() if card.paused_at else None,
            'created_at': card.created_at.isoformat() if card.created_at else None,
            'updated_at': card.updated_at.isoformat() if card.updated_at else None,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения карточки надзора: {str(e)}")


@app.post("/api/supervision/cards")
async def create_supervision_card(
    card_data: SupervisionCardCreate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Создать карточку надзора"""
    try:
        card = SupervisionCard(**card_data.model_dump())
        db.add(card)
        db.commit()
        db.refresh(card)

        log = ActivityLog(
            employee_id=current_user.id,
            action_type="create",
            entity_type="supervision_card",
            entity_id=card.id
        )
        db.add(log)
        db.commit()

        return {
            "id": card.id,
            "contract_id": card.contract_id,
            "column_name": card.column_name,
            "deadline": str(card.deadline) if card.deadline else None,
            "tags": card.tags,
            "senior_manager_id": card.senior_manager_id,
            "dan_id": card.dan_id,
            "dan_completed": card.dan_completed,
            "is_paused": card.is_paused,
            "pause_reason": card.pause_reason,
            "paused_at": card.paused_at.isoformat() if card.paused_at else None,
            "created_at": card.created_at.isoformat() if card.created_at else None,
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка создания карточки надзора: {str(e)}")


@app.patch("/api/supervision/cards/{card_id}")
async def update_supervision_card(
    card_id: int,
    updates: SupervisionCardUpdate,
    current_user: Employee = Depends(require_permission("supervision.update")),
    db: Session = Depends(get_db)
):
    """Обновить карточку надзора"""
    try:
        card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

        update_data = updates.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(card, field, value)

        db.commit()
        db.refresh(card)

        return {
            'id': card.id,
            'contract_id': card.contract_id,
            'column_name': card.column_name,
            'deadline': str(card.deadline) if card.deadline else None,
            'tags': card.tags,
            'senior_manager_id': card.senior_manager_id,
            'dan_id': card.dan_id,
            'dan_completed': card.dan_completed,
            'is_paused': card.is_paused,
            'pause_reason': card.pause_reason,
            'paused_at': card.paused_at.isoformat() if card.paused_at else None,
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка обновления карточки надзора: {str(e)}")


@app.patch("/api/supervision/cards/{card_id}/column")
async def move_supervision_card_to_column(
    card_id: int,
    move_request: SupervisionColumnMoveRequest,
    current_user: Employee = Depends(require_permission("supervision.move")),
    db: Session = Depends(get_db)
):
    """Переместить карточку надзора в другую колонку"""
    try:
        VALID_SUPERVISION_COLUMNS = [
            'Новый заказ', 'В ожидании',
            'Стадия 1: Закупка керамогранита', 'Стадия 2: Закупка сантехники',
            'Стадия 3: Закупка оборудования', 'Стадия 4: Закупка дверей и окон',
            'Стадия 5: Закупка настенных материалов', 'Стадия 6: Закупка напольных материалов',
            'Стадия 7: Лепного декора', 'Стадия 8: Освещения',
            'Стадия 9: бытовой техники', 'Стадия 10: Закупка заказной мебели',
            'Стадия 11: Закупка фабричной мебели', 'Стадия 12: Закупка декора',
            'Выполненный проект'
        ]
        if move_request.column_name not in VALID_SUPERVISION_COLUMNS:
            raise HTTPException(status_code=422, detail=f"Недопустимая колонка надзора: {move_request.column_name}")

        card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

        old_column = card.column_name
        card.column_name = move_request.column_name

        db.commit()
        db.refresh(card)

        return {
            'id': card.id,
            'contract_id': card.contract_id,
            'column_name': card.column_name,
            'old_column_name': old_column,
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка перемещения карточки надзора: {str(e)}")


@app.post("/api/supervision/cards/{card_id}/pause")
async def pause_supervision_card(
    card_id: int,
    pause_request: SupervisionPauseRequest,
    current_user: Employee = Depends(require_permission("supervision.pause_resume")),
    db: Session = Depends(get_db)
):
    """Приостановить карточку надзора"""
    try:
        card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

        card.is_paused = True
        card.pause_reason = pause_request.pause_reason
        card.paused_at = datetime.utcnow()

        # Добавляем запись в историю
        history = SupervisionProjectHistory(
            supervision_card_id=card_id,
            entry_type="pause",
            message=f"Приостановлено: {pause_request.pause_reason}",
            created_by=current_user.id
        )
        db.add(history)

        db.commit()
        db.refresh(card)

        return {
            'id': card.id,
            'is_paused': card.is_paused,
            'pause_reason': card.pause_reason,
            'paused_at': card.paused_at.isoformat(),
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка приостановки карточки: {str(e)}")


@app.post("/api/supervision/cards/{card_id}/resume")
async def resume_supervision_card(
    card_id: int,
    current_user: Employee = Depends(require_permission("supervision.pause_resume")),
    db: Session = Depends(get_db)
):
    """Возобновить карточку надзора"""
    try:
        card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

        card.is_paused = False
        card.pause_reason = None
        card.paused_at = None

        # Добавляем запись в историю
        history = SupervisionProjectHistory(
            supervision_card_id=card_id,
            entry_type="resume",
            message="Возобновлено",
            created_by=current_user.id
        )
        db.add(history)

        db.commit()
        db.refresh(card)

        return {
            'id': card.id,
            'is_paused': card.is_paused,
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка возобновления карточки: {str(e)}")


@app.get("/api/supervision/cards/{card_id}/history", response_model=List[SupervisionHistoryResponse])
async def get_supervision_card_history(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить историю карточки надзора"""
    try:
        history = db.query(SupervisionProjectHistory).filter(
            SupervisionProjectHistory.supervision_card_id == card_id
        ).order_by(SupervisionProjectHistory.created_at.desc()).all()

        return history

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения истории: {str(e)}")


# =========================
# ПЛАТЕЖИ
# =========================

@app.get("/api/payments/contract/{contract_id}")
async def get_payments_for_contract(
    contract_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить платежи по договору"""
    # Получаем статус договора для определения отчетного месяца
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    contract_status = contract.status if contract else ''

    payments = db.query(Payment).filter(Payment.contract_id == contract_id).all()

    # Добавляем имя сотрудника к каждому платежу
    result = []
    for payment in payments:
        payment_dict = {
            'id': payment.id,
            'contract_id': payment.contract_id,
            'crm_card_id': payment.crm_card_id,
            'supervision_card_id': payment.supervision_card_id,
            'employee_id': payment.employee_id,
            'role': payment.role,
            'stage_name': payment.stage_name,
            'calculated_amount': float(payment.calculated_amount) if payment.calculated_amount else 0,
            'manual_amount': float(payment.manual_amount) if payment.manual_amount else None,
            'final_amount': float(payment.final_amount) if payment.final_amount else 0,
            'is_manual': payment.is_manual,
            'payment_type': payment.payment_type,
            'report_month': payment.report_month,
            'payment_status': payment.payment_status if payment.payment_status else 'pending',  # ИСПРАВЛЕНИЕ: Статус по умолчанию
            'is_paid': payment.is_paid,
            'paid_date': payment.paid_date.isoformat() if payment.paid_date else None,
            'paid_by': payment.paid_by,
            'reassigned': payment.reassigned if payment.reassigned else False,
            'old_employee_id': payment.old_employee_id,
            'created_at': payment.created_at.isoformat() if payment.created_at else None,
            'updated_at': payment.updated_at.isoformat() if payment.updated_at else None,
        }

        # Получаем имя сотрудника
        employee = db.query(Employee).filter(Employee.id == payment.employee_id).first()
        payment_dict['employee_name'] = employee.full_name if employee else 'Неизвестный'
        payment_dict['position'] = employee.position if employee else ''  # ИСПРАВЛЕНИЕ: Добавлена должность

        # ИСПРАВЛЕНИЕ: Добавлены поля source и amount для совместимости
        payment_dict['source'] = 'CRM' if payment.crm_card_id else 'Оклад'
        payment_dict['amount'] = payment_dict['final_amount']  # Алиас

        # ИСПРАВЛЕНИЕ 25.01.2026: Добавлен contract_status для отображения "В работе" вместо "Не установлен"
        payment_dict['contract_status'] = contract_status

        result.append(payment_dict)

    return result


@app.post("/api/payments", response_model=PaymentResponse)
async def create_payment(
    payment_data: PaymentCreate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Создать платеж"""
    # Проверяем существование договора
    if payment_data.contract_id:
        contract = db.query(Contract).filter(Contract.id == payment_data.contract_id).first()
        if not contract:
            raise HTTPException(status_code=404, detail="Договор не найден")

    # Проверяем существование сотрудника
    if payment_data.employee_id:
        employee = db.query(Employee).filter(Employee.id == payment_data.employee_id).first()
        if not employee:
            raise HTTPException(status_code=404, detail="Сотрудник не найден")

    # Защита от дублей: проверяем нет ли уже платежа с теми же параметрами
    if payment_data.contract_id and payment_data.employee_id and payment_data.stage_name:
        duplicate_query = db.query(Payment).filter(
            Payment.contract_id == payment_data.contract_id,
            Payment.employee_id == payment_data.employee_id,
            Payment.stage_name == payment_data.stage_name,
            Payment.role == payment_data.role
        )
        if payment_data.crm_card_id:
            duplicate_query = duplicate_query.filter(Payment.crm_card_id == payment_data.crm_card_id)
        if payment_data.supervision_card_id:
            duplicate_query = duplicate_query.filter(Payment.supervision_card_id == payment_data.supervision_card_id)
        existing = duplicate_query.first()
        if existing:
            raise HTTPException(
                status_code=409,
                detail=f"Платёж уже существует (id={existing.id}). Используйте PUT для обновления."
            )

    try:
        # Получаем данные и устанавливаем значения по умолчанию для NOT NULL полей
        data = payment_data.model_dump(exclude_unset=True)

        # Устанавливаем значения по умолчанию для NOT NULL полей
        if data.get('calculated_amount') is None:
            data['calculated_amount'] = 0.0
        if data.get('final_amount') is None:
            data['final_amount'] = data.get('calculated_amount', 0.0)
        # payment_status остаётся NULL - платёж ещё не оплачен
        # Статус 'paid' устанавливается кнопкой "Оплачено" в UI

        payment = Payment(**data)
        db.add(payment)
        db.commit()
        db.refresh(payment)

        log = ActivityLog(
            employee_id=current_user.id,
            action_type="create",
            entity_type="payment",
            entity_id=payment.id
        )
        db.add(log)
        db.commit()

        return payment

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка создания платежа: {str(e)}")


# ИСПРАВЛЕНИЕ 30.01.2026: Endpoint перемещен ПЕРЕД /api/payments/{payment_id}
# чтобы FastAPI не перехватывал 'calculate' как payment_id
@app.get("/api/payments/calculate")
async def calculate_payment_amount(
    contract_id: int,
    employee_id: int,
    role: str,
    stage_name: Optional[str] = None,
    supervision_card_id: Optional[int] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Рассчитать сумму оплаты на основе тарифов - ПОЛНАЯ ЛОГИКА"""
    try:
        # Получаем договор
        contract = db.query(Contract).filter(Contract.id == contract_id).first()
        if not contract:
            return {'amount': 0, 'error': 'Contract not found'}

        area = float(contract.area) if contract.area else 0
        project_type = contract.project_type
        city = contract.city

        # ========== АВТОРСКИЙ НАДЗОР ==========
        if supervision_card_id:
            rate = db.query(Rate).filter(
                Rate.project_type == 'Авторский надзор',
                Rate.role == role
            )
            if stage_name:
                rate = rate.filter(or_(Rate.stage_name == stage_name, Rate.stage_name.is_(None)))
            rate = rate.order_by(
                # Приоритет: сначала с конкретной стадией, потом без стадии
                Rate.stage_name.desc().nullslast()
            ).first()

            if rate and rate.rate_per_m2:
                amount = area * float(rate.rate_per_m2)
                return {'amount': amount, 'rate_per_m2': float(rate.rate_per_m2)}
            return {'amount': 0}

        # ========== ЗАМЕРЩИК ==========
        if role == 'Замерщик':
            rate = db.query(Rate).filter(
                Rate.role == 'Замерщик',
                Rate.city == city
            ).first()
            if rate and rate.surveyor_price:
                return {'amount': float(rate.surveyor_price), 'surveyor_price': float(rate.surveyor_price)}
            return {'amount': 0}

        # ========== ИНДИВИДУАЛЬНЫЙ ==========
        if project_type == 'Индивидуальный':
            print(f"[CALC] Индивидуальный проект: role={role}, area={area}, stage_name={stage_name}")
            query = db.query(Rate).filter(
                Rate.project_type == 'Индивидуальный',
                Rate.role == role
            )
            if stage_name:
                # Сначала ищем с конкретной стадией
                rate = query.filter(Rate.stage_name == stage_name).first()
                print(f"[CALC] Поиск с stage_name='{stage_name}': rate={rate}")
                # Если не найден - ищем без стадии
                if not rate:
                    rate = query.filter(Rate.stage_name.is_(None)).first()
                    print(f"[CALC] Поиск с stage_name IS NULL: rate={rate}")
            else:
                rate = query.filter(Rate.stage_name.is_(None)).first()
                print(f"[CALC] Поиск без stage_name (IS NULL): rate={rate}")

            if rate and rate.rate_per_m2:
                amount = area * float(rate.rate_per_m2)
                print(f"[CALC] Найден тариф: rate_per_m2={rate.rate_per_m2}, amount={amount}")
                return {'amount': amount, 'rate_per_m2': float(rate.rate_per_m2)}
            print(f"[CALC] Тариф НЕ найден или rate_per_m2=0, возвращаем 0")
            return {'amount': 0}

        # ========== ШАБЛОННЫЙ ==========
        if project_type == 'Шаблонный':
            rate = db.query(Rate).filter(
                Rate.project_type == 'Шаблонный',
                Rate.role == role,
                Rate.area_from <= area,
                or_(Rate.area_to >= area, Rate.area_to.is_(None))
            ).order_by(Rate.area_from.asc()).first()

            if rate and rate.fixed_price:
                return {'amount': float(rate.fixed_price), 'fixed_price': float(rate.fixed_price)}
            return {'amount': 0}

        # ========== АВТОРСКИЙ НАДЗОР (по типу проекта) ==========
        if project_type == 'Авторский надзор':
            query = db.query(Rate).filter(
                Rate.project_type == 'Авторский надзор',
                Rate.role == role
            )
            if stage_name:
                query = query.filter(or_(Rate.stage_name == stage_name, Rate.stage_name.is_(None)))
            rate = query.first()

            if rate and rate.rate_per_m2:
                amount = area * float(rate.rate_per_m2)
                return {'amount': amount, 'rate_per_m2': float(rate.rate_per_m2)}
            return {'amount': 0}

        return {'amount': 0}

    except Exception as e:
        logger.error(f"Error calculating payment: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'amount': 0, 'error': str(e)}



# ============================================================================
# ИСПРАВЛЕНИЕ 30.01.2026: Статические endpoints /api/payments/* перемещены
# ПЕРЕД динамическим /api/payments/{payment_id} чтобы избежать HTTP 422
# ============================================================================

@app.get("/api/payments/summary")
async def get_payments_summary(
    year: int,
    month: Optional[int] = None,
    quarter: Optional[int] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить сводку по платежам"""
    try:
        from sqlalchemy import extract

        query = db.query(Payment).filter(extract('year', Payment.created_at) == year)

        if month:
            query = query.filter(extract('month', Payment.created_at) == month)
        if quarter:
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            query = query.filter(extract('month', Payment.created_at).between(start_month, end_month))

        payments = query.all()

        # Суммы по статусам
        paid_amount = sum(p.final_amount or 0 for p in payments if p.is_paid)
        pending_amount = sum(p.final_amount or 0 for p in payments if not p.is_paid)

        # По ролям
        by_role = {}
        for p in payments:
            role = p.role or 'Не указано'
            if role not in by_role:
                by_role[role] = {'paid': 0, 'pending': 0, 'count': 0}
            by_role[role]['count'] += 1
            if p.is_paid:
                by_role[role]['paid'] += p.final_amount or 0
            else:
                by_role[role]['pending'] += p.final_amount or 0

        return {
            'year': year,
            'month': month,
            'quarter': quarter,
            'total_paid': paid_amount,
            'total_pending': pending_amount,
            'total': paid_amount + pending_amount,
            'by_role': by_role,
            'payments_count': len(payments)
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения сводки: {str(e)}")


# =========================
# CRM EXTENDED ENDPOINTS
# =========================


@app.get("/api/payments/by-type")
async def get_payments_by_type(
    payment_type: str,
    project_type_filter: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Получить платежи по типу выплаты и фильтру типа проекта

    Сигнатура совпадает с database/db_manager.py:_get_payments_by_type_from_db()

    Args:
        payment_type: Тип выплаты ('Индивидуальные проекты', 'Шаблонные проекты', 'Оклады', 'Авторский надзор')
        project_type_filter: Фильтр типа проекта ('Индивидуальный', 'Шаблонный', 'Авторский надзор', None)

    Returns:
        Список платежей с полями:
            - id, contract_id, employee_id, employee_name, position
            - role, stage_name, final_amount, payment_type
            - report_month, payment_status
            - contract_number, address, area, city, agent_type
            - source ('CRM' или 'Оклад')
            - card_stage (для CRM)
    """
    try:
        result = []

        # Оклады - только из таблицы salaries
        if payment_type == 'Оклады':
            salaries = db.query(Salary).all()
            for s in salaries:
                employee = db.query(Employee).filter(Employee.id == s.employee_id).first()
                contract = db.query(Contract).filter(Contract.id == s.contract_id).first() if s.contract_id else None

                result.append({
                    'id': s.id,
                    'contract_id': s.contract_id,
                    'employee_id': s.employee_id,
                    'employee_name': employee.full_name if employee else 'Неизвестный',
                    'position': employee.position if employee else '',
                    'role': s.payment_type,
                    'stage_name': s.stage_name,
                    'final_amount': float(s.amount) if s.amount else 0,
                    'amount': float(s.amount) if s.amount else 0,
                    'payment_type': s.payment_type,
                    'report_month': s.report_month,
                    'payment_status': s.payment_status,
                    'project_type': s.project_type,
                    'contract_number': contract.contract_number if contract else None,
                    'address': contract.address if contract else None,
                    'area': float(contract.area) if contract and contract.area else None,
                    'city': contract.city if contract else None,
                    'agent_type': contract.agent_type if contract else None,
                    'source': 'Оклад',
                    'card_stage': None,
                    'comments': s.comments  # Добавлено поле комментария
                })

        # Авторский надзор - из payments с supervision_card + salaries
        elif project_type_filter == 'Авторский надзор':
            # Платежи из CRM надзора - ИСПРАВЛЕНО 06.02.2026
            payments = db.query(Payment).filter(
                Payment.supervision_card_id.isnot(None)
            ).all()

            for p in payments:
                employee = db.query(Employee).filter(Employee.id == p.employee_id).first()
                supervision_card = db.query(SupervisionCard).filter(SupervisionCard.id == p.supervision_card_id).first()
                contract = db.query(Contract).filter(Contract.id == supervision_card.contract_id).first() if supervision_card else None

                result.append({
                    'id': p.id,
                    'contract_id': p.contract_id,
                    'crm_card_id': p.crm_card_id,
                    'supervision_card_id': p.supervision_card_id,
                    'employee_id': p.employee_id,
                    'employee_name': employee.full_name if employee else 'Неизвестный',
                    'position': employee.position if employee else '',
                    'role': p.role,
                    'stage_name': p.stage_name,
                    'calculated_amount': float(p.calculated_amount) if p.calculated_amount else 0,
                    'final_amount': float(p.final_amount) if p.final_amount else 0,
                    'amount': float(p.final_amount) if p.final_amount else 0,
                    'payment_type': p.payment_type,
                    'payment_subtype': p.payment_type,
                    'report_month': p.report_month,
                    'payment_status': p.payment_status,
                    'contract_number': contract.contract_number if contract else None,
                    'address': contract.address if contract else None,
                    'area': float(contract.area) if contract and contract.area else None,
                    'city': contract.city if contract else None,
                    'agent_type': contract.agent_type if contract else None,
                    'project_type': contract.project_type if contract else 'Авторский надзор',
                    'source': 'CRM Надзор',
                    'card_stage': supervision_card.column_name if supervision_card else None,
                    'reassigned': p.reassigned if hasattr(p, 'reassigned') else False
                })

            # Добавляем оклады с типом "Авторский надзор"
            salaries = db.query(Salary).filter(Salary.project_type == project_type_filter).all()
            for s in salaries:
                employee = db.query(Employee).filter(Employee.id == s.employee_id).first()
                contract = db.query(Contract).filter(Contract.id == s.contract_id).first() if s.contract_id else None

                result.append({
                    'id': s.id,
                    'contract_id': s.contract_id,
                    'employee_id': s.employee_id,
                    'employee_name': employee.full_name if employee else 'Неизвестный',
                    'position': employee.position if employee else '',
                    'role': s.payment_type,
                    'stage_name': s.stage_name,
                    'final_amount': float(s.amount) if s.amount else 0,
                    'amount': float(s.amount) if s.amount else 0,
                    'payment_type': 'Оклад',
                    'report_month': s.report_month,
                    'payment_status': s.payment_status,
                    'contract_number': contract.contract_number if contract else None,
                    'address': contract.address if contract else None,
                    'area': float(contract.area) if contract and contract.area else None,
                    'city': contract.city if contract else None,
                    'agent_type': contract.agent_type if contract else None,
                    'source': 'Оклад',
                    'card_stage': None,
                    'comments': s.comments  # Добавлено поле комментария
                })

        # Индивидуальные и шаблонные проекты - из payments с crm_card + salaries
        elif project_type_filter:
            # Платежи из CRM
            payments_query = db.query(Payment).join(
                CRMCard, Payment.crm_card_id == CRMCard.id
            ).join(
                Contract, CRMCard.contract_id == Contract.id
            ).filter(
                Contract.project_type == project_type_filter
            )

            for p in payments_query.all():
                employee = db.query(Employee).filter(Employee.id == p.employee_id).first()
                crm_card = db.query(CRMCard).filter(CRMCard.id == p.crm_card_id).first()
                contract = db.query(Contract).filter(Contract.id == crm_card.contract_id).first() if crm_card else None

                result.append({
                    'id': p.id,
                    'contract_id': p.contract_id,
                    'employee_id': p.employee_id,
                    'employee_name': employee.full_name if employee else 'Неизвестный',
                    'position': employee.position if employee else '',
                    'role': p.role,
                    'stage_name': p.stage_name,
                    'final_amount': float(p.final_amount) if p.final_amount else 0,
                    'amount': float(p.final_amount) if p.final_amount else 0,
                    'payment_type': p.payment_type,
                    'report_month': p.report_month,
                    'payment_status': p.payment_status,
                    'contract_number': contract.contract_number if contract else None,
                    'address': contract.address if contract else None,
                    'area': float(contract.area) if contract and contract.area else None,
                    'city': contract.city if contract else None,
                    'agent_type': contract.agent_type if contract else None,
                    'source': 'CRM',
                    'card_stage': crm_card.column_name if crm_card else None,
                    'reassigned': p.reassigned if hasattr(p, 'reassigned') else False,
                    'old_employee_id': p.old_employee_id if hasattr(p, 'old_employee_id') else None
                })

            # Добавляем оклады с этим типом проекта
            salaries = db.query(Salary).filter(Salary.project_type == project_type_filter).all()
            for s in salaries:
                employee = db.query(Employee).filter(Employee.id == s.employee_id).first()
                contract = db.query(Contract).filter(Contract.id == s.contract_id).first() if s.contract_id else None

                result.append({
                    'id': s.id,
                    'contract_id': s.contract_id,
                    'employee_id': s.employee_id,
                    'employee_name': employee.full_name if employee else 'Неизвестный',
                    'position': employee.position if employee else '',
                    'role': s.payment_type,
                    'stage_name': s.stage_name,
                    'final_amount': float(s.amount) if s.amount else 0,
                    'amount': float(s.amount) if s.amount else 0,
                    'payment_type': 'Оклад',
                    'report_month': s.report_month,
                    'payment_status': s.payment_status,
                    'contract_number': contract.contract_number if contract else None,
                    'address': contract.address if contract else None,
                    'area': float(contract.area) if contract and contract.area else None,
                    'city': contract.city if contract else None,
                    'agent_type': contract.agent_type if contract else None,
                    'source': 'Оклад',
                    'card_stage': None,
                    'comments': s.comments  # Добавлено поле комментария
                })

        # Сортируем по id в обратном порядке
        result.sort(key=lambda x: x['id'], reverse=True)
        return result

    except Exception as e:
        logger.error(f"Error in get_payments_by_type: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Ошибка получения платежей по типу: {str(e)}")



# =========================

@app.get("/api/payments/all-optimized")
async def get_all_payments_optimized(
    year: Optional[int] = None,
    month: Optional[int] = None,
    quarter: Optional[int] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Оптимизированная загрузка всех выплат - один запрос вместо 132"""
    try:
        # Базовый запрос для payments
        query = db.query(Payment).join(Employee, Payment.employee_id == Employee.id)

        # Фильтрация по периоду
        if year and month:
            report_month_str = f"{year}-{month:02d}"
            query = query.filter(Payment.report_month == report_month_str)
        elif year and quarter:
            months = [(quarter - 1) * 3 + i for i in range(1, 4)]
            month_strs = [f"{year}-{m:02d}" for m in months]
            query = query.filter(Payment.report_month.in_(month_strs))
        elif year:
            query = query.filter(Payment.report_month.like(f"{year}-%"))

        payments = query.all()

        # Batch-load all contracts to avoid N+1 queries
        contract_ids = list(set(p.contract_id for p in payments if p.contract_id))
        contracts_list = db.query(Contract).filter(Contract.id.in_(contract_ids)).all() if contract_ids else []
        contracts_map = {c.id: c for c in contracts_list}

        result = []
        for p in payments:
            contract = contracts_map.get(p.contract_id)

            result.append({
                'id': p.id,
                'contract_id': p.contract_id,
                'employee_id': p.employee_id,
                'employee_name': p.employee.full_name if p.employee else '',
                'position': p.employee.position if p.employee else '',
                'role': p.role,
                'stage_name': p.stage_name,
                'calculated_amount': float(p.calculated_amount) if p.calculated_amount else 0,
                'final_amount': float(p.final_amount) if p.final_amount else 0,
                'amount': float(p.final_amount) if p.final_amount else 0,
                'payment_type': p.payment_type,
                'payment_status': p.payment_status,
                'report_month': p.report_month or '',
                'source': 'CRM',
                'project_type': contract.project_type if contract else '',
                'address': contract.address if contract else '',
                'agent_type': contract.agent_type if contract else ''
            })

        return result

    except Exception as e:
        logger.error(f"Error getting optimized payments: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



@app.post("/api/payments/recalculate")
async def recalculate_payments(
    contract_id: Optional[int] = None,
    role: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Пересчет выплат по текущим тарифам"""
    try:
        # Получаем выплаты для пересчета
        query = db.query(Payment)

        if contract_id:
            query = query.filter(Payment.contract_id == contract_id)
        if role:
            query = query.filter(Payment.role == role)

        payments = query.all()
        updated = 0
        errors = []

        for payment in payments:
            try:
                # Получаем договор
                contract = db.query(Contract).filter(Contract.id == payment.contract_id).first()
                if not contract:
                    continue

                area = float(contract.area) if contract.area else 0
                project_type = contract.project_type
                city = contract.city

                new_amount = 0

                # Замерщик
                if payment.role == 'Замерщик':
                    rate = db.query(Rate).filter(
                        Rate.role == 'Замерщик',
                        Rate.city == city
                    ).first()
                    if rate and rate.surveyor_price:
                        new_amount = float(rate.surveyor_price)

                # Индивидуальный
                elif project_type == 'Индивидуальный':
                    query_rate = db.query(Rate).filter(
                        Rate.project_type == 'Индивидуальный',
                        Rate.role == payment.role
                    )
                    if payment.stage_name:
                        rate = query_rate.filter(Rate.stage_name == payment.stage_name).first()
                        if not rate:
                            rate = query_rate.filter(Rate.stage_name.is_(None)).first()
                    else:
                        rate = query_rate.filter(Rate.stage_name.is_(None)).first()

                    if rate and rate.rate_per_m2:
                        new_amount = area * float(rate.rate_per_m2)

                # Шаблонный
                elif project_type == 'Шаблонный':
                    rate = db.query(Rate).filter(
                        Rate.project_type == 'Шаблонный',
                        Rate.role == payment.role,
                        Rate.area_from <= area,
                        or_(Rate.area_to >= area, Rate.area_to.is_(None))
                    ).order_by(Rate.area_from.asc()).first()

                    if rate and rate.fixed_price:
                        new_amount = float(rate.fixed_price)

                # Авторский надзор
                elif project_type == 'Авторский надзор' or payment.supervision_card_id:
                    rate = db.query(Rate).filter(
                        Rate.project_type == 'Авторский надзор',
                        Rate.role == payment.role
                    ).first()

                    if rate and rate.rate_per_m2:
                        new_amount = area * float(rate.rate_per_m2)

                # Обновляем если сумма изменилась
                if new_amount != payment.calculated_amount:
                    old_amount = payment.calculated_amount
                    payment.calculated_amount = new_amount
                    payment.final_amount = new_amount
                    updated += 1
                    print(f"[RECALC] Payment ID={payment.id}: {old_amount} -> {new_amount}")

            except Exception as e:
                errors.append({'payment_id': payment.id, 'error': str(e)})

        db.commit()

        return {
            'status': 'success',
            'updated': updated,
            'total': len(payments),
            'errors': errors
        }

    except Exception as e:
        db.rollback()
        logger.error(f"Error recalculating payments: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# =========================
# STATISTICS CRM ENDPOINTS
# =========================


@app.get("/api/payments")
async def get_all_payments(
    year: Optional[int] = None,
    payment_type: Optional[str] = None,
    month: Optional[int] = None,
    include_null_month: Optional[bool] = False,  # ДОБАВЛЕНО 06.02.2026: включить платежи без месяца
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить все платежи с фильтрами (включая оклады из таблицы salaries)"""
    try:
        result = []

        # 1. Платежи из таблицы payments (CRM и Надзор)
        payments_query = db.query(Payment)

        if year:
            if include_null_month:
                # ИСПРАВЛЕНИЕ 06.02.2026: Включаем платежи с NULL report_month (В работе)
                payments_query = payments_query.filter(
                    or_(
                        Payment.report_month.like(f'{year}%'),
                        Payment.report_month.is_(None)
                    )
                )
            else:
                payments_query = payments_query.filter(Payment.report_month.like(f'{year}%'))
        if month:
            payments_query = payments_query.filter(Payment.report_month.like(f'{year}-{month:02d}%' if year else f'%-{month:02d}%'))
        if payment_type and payment_type != 'Оклад':
            payments_query = payments_query.filter(Payment.payment_type == payment_type)

        payments = payments_query.all()

        # Batch-load all related entities to avoid N+1 queries
        employee_ids = list(set(p.employee_id for p in payments if p.employee_id))
        contract_ids = list(set(p.contract_id for p in payments if p.contract_id))
        crm_card_ids = list(set(p.crm_card_id for p in payments if p.crm_card_id))
        supervision_card_ids = list(set(p.supervision_card_id for p in payments if p.supervision_card_id))

        employees_map = {e.id: e for e in db.query(Employee).filter(Employee.id.in_(employee_ids)).all()} if employee_ids else {}
        contracts_map = {c.id: c for c in db.query(Contract).filter(Contract.id.in_(contract_ids)).all()} if contract_ids else {}
        crm_cards_map = {c.id: c for c in db.query(CRMCard).filter(CRMCard.id.in_(crm_card_ids)).all()} if crm_card_ids else {}
        supervision_cards_map = {c.id: c for c in db.query(SupervisionCard).filter(SupervisionCard.id.in_(supervision_card_ids)).all()} if supervision_card_ids else {}

        # Also load contracts referenced by CRM cards and supervision cards
        crm_contract_ids = list(set(c.contract_id for c in crm_cards_map.values() if c.contract_id))
        sup_contract_ids = list(set(c.contract_id for c in supervision_cards_map.values() if c.contract_id))
        extra_contract_ids = [cid for cid in crm_contract_ids + sup_contract_ids if cid not in contracts_map]
        if extra_contract_ids:
            extra_contracts = {c.id: c for c in db.query(Contract).filter(Contract.id.in_(extra_contract_ids)).all()}
            contracts_map.update(extra_contracts)

        for p in payments:
            employee = employees_map.get(p.employee_id)

            # Получаем данные контракта через CRM карточку или напрямую
            contract = None
            if p.crm_card_id:
                crm_card = crm_cards_map.get(p.crm_card_id)
                if crm_card:
                    contract = contracts_map.get(crm_card.contract_id)
            elif p.supervision_card_id:
                supervision_card = supervision_cards_map.get(p.supervision_card_id)
                if supervision_card:
                    contract = contracts_map.get(supervision_card.contract_id)
            elif p.contract_id:
                contract = contracts_map.get(p.contract_id)

            # Определяем source
            if p.crm_card_id:
                source = 'CRM'
            elif p.supervision_card_id:
                source = 'CRM Надзор'
            else:
                source = 'CRM'  # Если нет ни crm_card_id ни supervision_card_id, но есть contract_id

            # ИСПРАВЛЕНИЕ 06.02.2026: Для платежей надзора project_type = 'Авторский надзор'
            if p.supervision_card_id:
                payment_project_type = 'Авторский надзор'
            else:
                payment_project_type = contract.project_type if contract else None

            result.append({
                'id': p.id,
                'contract_id': p.contract_id,
                'crm_card_id': p.crm_card_id,
                'supervision_card_id': p.supervision_card_id,
                'employee_id': p.employee_id,
                'employee_name': employee.full_name if employee else 'Неизвестный',
                'position': employee.position if employee else '',
                'role': p.role,
                'stage_name': p.stage_name,
                'calculated_amount': float(p.calculated_amount) if p.calculated_amount else 0,
                'final_amount': float(p.final_amount) if p.final_amount else 0,
                'amount': float(p.final_amount) if p.final_amount else 0,
                'payment_type': p.payment_type,
                'payment_subtype': p.payment_type,  # Тип выплаты: Аванс, Доплата, Полная оплата
                'source': source,
                'report_month': p.report_month,
                'payment_status': p.payment_status if p.payment_status else 'pending',
                'is_paid': p.is_paid,
                'created_at': p.created_at.isoformat() if p.created_at else None,
                # Данные из контракта
                'project_type': payment_project_type,
                'agent_type': contract.agent_type if contract else None,
                'address': contract.address if contract else None,
                'contract_number': contract.contract_number if contract else None,
                'area': float(contract.area) if contract and contract.area else None,
                'city': contract.city if contract else None,
                'reassigned': p.reassigned if hasattr(p, 'reassigned') else False,
            })

        # 2. Оклады из таблицы salaries (если не фильтруем по payment_type отличному от Оклад)
        if not payment_type or payment_type == 'Оклад':
            salaries_query = db.query(Salary)

            if year:
                if include_null_month:
                    # ИСПРАВЛЕНИЕ 06.02.2026: Включаем оклады с NULL report_month
                    salaries_query = salaries_query.filter(
                        or_(
                            Salary.report_month.like(f'{year}%'),
                            Salary.report_month.is_(None)
                        )
                    )
                else:
                    salaries_query = salaries_query.filter(Salary.report_month.like(f'{year}%'))
            if month:
                salaries_query = salaries_query.filter(Salary.report_month.like(f'{year}-{month:02d}%' if year else f'%-{month:02d}%'))

            salaries = salaries_query.all()

            # Batch-load employees and contracts for salaries to avoid N+1 queries
            sal_employee_ids = list(set(s.employee_id for s in salaries if s.employee_id))
            sal_contract_ids = list(set(s.contract_id for s in salaries if s.contract_id))
            sal_employees_map = {e.id: e for e in db.query(Employee).filter(Employee.id.in_(sal_employee_ids)).all()} if sal_employee_ids else {}
            # Reuse already loaded contracts and load any missing ones
            missing_contract_ids = [cid for cid in sal_contract_ids if cid not in contracts_map]
            if missing_contract_ids:
                extra = {c.id: c for c in db.query(Contract).filter(Contract.id.in_(missing_contract_ids)).all()}
                contracts_map.update(extra)

            for s in salaries:
                employee = sal_employees_map.get(s.employee_id)
                contract = contracts_map.get(s.contract_id) if s.contract_id else None

                result.append({
                    'id': s.id,
                    'contract_id': s.contract_id,
                    'crm_card_id': None,
                    'supervision_card_id': None,
                    'employee_id': s.employee_id,
                    'employee_name': employee.full_name if employee else 'Неизвестный',
                    'position': employee.position if employee else '',
                    'role': s.payment_type,  # payment_type в salaries = роль
                    'stage_name': s.stage_name,
                    'calculated_amount': float(s.amount) if s.amount else 0,
                    'final_amount': float(s.amount) if s.amount else 0,
                    'amount': float(s.amount) if s.amount else 0,
                    'payment_type': s.payment_type,
                    'payment_subtype': 'Оклад',  # Всегда Оклад для salaries
                    'source': 'Оклад',
                    'report_month': s.report_month,
                    'payment_status': s.payment_status if s.payment_status else 'pending',
                    'is_paid': s.payment_status == 'paid' if s.payment_status else False,
                    'created_at': s.created_at.isoformat() if s.created_at else None,
                    # Данные из контракта
                    'project_type': s.project_type if s.project_type else (contract.project_type if contract else None),
                    'agent_type': contract.agent_type if contract else None,
                    'address': contract.address if contract else None,
                    'contract_number': contract.contract_number if contract else None,
                    'area': float(contract.area) if contract and contract.area else None,
                    'city': contract.city if contract else None,
                    'reassigned': False,  # Оклады не переназначаются
                    'comments': s.comments,  # Комментарий из оклада
                })

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения платежей: {str(e)}")


# =========================
# PROJECT FILES EXTENDED ENDPOINTS
# =========================

# ВАЖНО: Этот endpoint должен быть ПЕРЕД /api/files/{file_id}
@app.get("/api/payments/{payment_id}", response_model=PaymentResponse)
async def get_payment_by_id(
    payment_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить платеж по ID"""
    payment = db.query(Payment).filter(Payment.id == payment_id).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Платеж не найден")
    # Добавляем employee_name через JOIN
    employee = db.query(Employee).filter(Employee.id == payment.employee_id).first()
    payment.employee_name = employee.full_name if employee else 'Неизвестный'
    return payment


@app.put("/api/payments/{payment_id}", response_model=PaymentResponse)
async def update_payment(
    payment_id: int,
    payment_data: PaymentUpdate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить платеж"""
    payment = db.query(Payment).filter(Payment.id == payment_id).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Платеж не найден")

    for field, value in payment_data.model_dump(exclude_unset=True).items():
        setattr(payment, field, value)

    payment.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(payment)

    return payment


@app.delete("/api/payments/{payment_id}")
async def delete_payment(
    payment_id: int,
    current_user: Employee = Depends(require_permission("payments.delete")),
    db: Session = Depends(get_db)
):
    """Удалить платеж"""
    payment = db.query(Payment).filter(Payment.id == payment_id).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Платеж не найден")

    log = ActivityLog(
        employee_id=current_user.id,
        action_type="delete",
        entity_type="payment",
        entity_id=payment_id
    )
    db.add(log)

    db.delete(payment)
    db.commit()

    return {"status": "success", "message": "Платеж удален"}


@app.patch("/api/payments/contract/{contract_id}/report-month")
async def set_payments_report_month(
    contract_id: int,
    data: dict,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Установить отчетный месяц для всех платежей договора без месяца"""
    report_month = data.get('report_month')
    if not report_month:
        raise HTTPException(status_code=400, detail="report_month обязателен")

    # Обновляем все платежи без отчетного месяца
    result = db.query(Payment).filter(
        Payment.contract_id == contract_id,
        or_(Payment.report_month == None, Payment.report_month == '')
    ).update(
        {'report_month': report_month},
        synchronize_session='fetch'
    )

    db.commit()

    return {"status": "success", "updated_count": result}


# =========================
# ТАРИФЫ
# =========================

@app.get("/api/rates", response_model=List[RateResponse])
async def get_rates(
    project_type: Optional[str] = None,
    role: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить тарифы"""
    query = db.query(Rate)
    if project_type:
        query = query.filter(Rate.project_type == project_type)
    if role:
        query = query.filter(Rate.role == role)
    return query.all()


# ВАЖНО: Эти endpoints должны быть ПЕРЕД /api/rates/{rate_id}
# иначе FastAPI интерпретирует "template" как rate_id

@app.get("/api/rates/template")
async def get_template_rates_early(
    role: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить шаблонные тарифы"""
    query = db.query(Rate).filter(Rate.project_type == 'Шаблонный')
    if role:
        query = query.filter(Rate.role == role)
    return query.all()


@app.get("/api/rates/{rate_id}", response_model=RateResponse)
async def get_rate(
    rate_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить тариф по ID"""
    rate = db.query(Rate).filter(Rate.id == rate_id).first()
    if not rate:
        raise HTTPException(status_code=404, detail="Тариф не найден")
    return rate


@app.post("/api/rates", response_model=RateResponse)
async def create_rate(
    rate_data: RateCreate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Создать тариф"""
    rate = Rate(**rate_data.model_dump())
    db.add(rate)
    db.commit()
    db.refresh(rate)
    return rate


@app.put("/api/rates/{rate_id}", response_model=RateResponse)
async def update_rate(
    rate_id: int,
    rate_data: RateUpdate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить тариф"""
    rate = db.query(Rate).filter(Rate.id == rate_id).first()
    if not rate:
        raise HTTPException(status_code=404, detail="Тариф не найден")

    for field, value in rate_data.model_dump(exclude_unset=True).items():
        setattr(rate, field, value)

    rate.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(rate)

    return rate


@app.delete("/api/rates/{rate_id}")
async def delete_rate(
    rate_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Удалить тариф"""
    rate = db.query(Rate).filter(Rate.id == rate_id).first()
    if not rate:
        raise HTTPException(status_code=404, detail="Тариф не найден")

    db.delete(rate)
    db.commit()

    return {"status": "success", "message": "Тариф удален"}


# =========================
# ЗАРПЛАТЫ
# =========================

@app.get("/api/salaries", response_model=List[SalaryResponse])
async def get_salaries(
    report_month: Optional[str] = None,
    employee_id: Optional[int] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить зарплаты"""
    query = db.query(Salary)
    if report_month:
        query = query.filter(Salary.report_month == report_month)
    if employee_id:
        query = query.filter(Salary.employee_id == employee_id)
    return query.all()


@app.get("/api/salaries/report")
async def get_salary_report(
    report_month: Optional[str] = None,
    employee_id: Optional[int] = None,
    payment_type: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить отчет по зарплатам"""
    try:
        query = db.query(Salary)

        if report_month:
            query = query.filter(Salary.report_month == report_month)
        if employee_id:
            query = query.filter(Salary.employee_id == employee_id)
        if payment_type:
            query = query.filter(Salary.payment_type == payment_type)

        salaries = query.all()

        # Группировка по сотрудникам
        employee_totals = {}
        for s in salaries:
            emp_id = s.employee_id
            if emp_id not in employee_totals:
                emp = db.query(Employee).filter(Employee.id == emp_id).first()
                employee_totals[emp_id] = {
                    'employee_id': emp_id,
                    'employee_name': emp.full_name if emp else 'Неизвестный',
                    'position': emp.position if emp else '',
                    'total_amount': 0,
                    'advance_payment': 0,
                    'records': []
                }

            employee_totals[emp_id]['total_amount'] += s.amount or 0
            employee_totals[emp_id]['advance_payment'] += s.advance_payment or 0
            employee_totals[emp_id]['records'].append({
                'id': s.id,
                'payment_type': s.payment_type,
                'stage_name': s.stage_name,
                'amount': s.amount,
                'report_month': s.report_month,
                'payment_status': s.payment_status
            })

        return {
            'report_month': report_month,
            'total_amount': sum(e['total_amount'] for e in employee_totals.values()),
            'employees': list(employee_totals.values())
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения отчета: {str(e)}")


@app.get("/api/salaries/{salary_id}", response_model=SalaryResponse)
async def get_salary(
    salary_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить зарплату по ID"""
    salary = db.query(Salary).filter(Salary.id == salary_id).first()
    if not salary:
        raise HTTPException(status_code=404, detail="Запись о зарплате не найдена")
    return salary


@app.post("/api/salaries", response_model=SalaryResponse)
async def create_salary(
    salary_data: SalaryCreate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Создать запись о зарплате"""
    salary = Salary(**salary_data.model_dump())
    db.add(salary)
    db.commit()
    db.refresh(salary)
    return salary


@app.put("/api/salaries/{salary_id}", response_model=SalaryResponse)
async def update_salary(
    salary_id: int,
    salary_data: SalaryUpdate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить запись о зарплате"""
    salary = db.query(Salary).filter(Salary.id == salary_id).first()
    if not salary:
        raise HTTPException(status_code=404, detail="Запись о зарплате не найдена")

    for field, value in salary_data.model_dump(exclude_unset=True).items():
        setattr(salary, field, value)

    db.commit()
    db.refresh(salary)

    return salary


@app.delete("/api/salaries/{salary_id}")
async def delete_salary(
    salary_id: int,
    current_user: Employee = Depends(require_permission("salaries.delete")),
    db: Session = Depends(get_db)
):
    """Удалить запись о зарплате"""
    salary = db.query(Salary).filter(Salary.id == salary_id).first()
    if not salary:
        raise HTTPException(status_code=404, detail="Запись о зарплате не найдена")

    log = ActivityLog(
        employee_id=current_user.id,
        action_type="delete",
        entity_type="salary",
        entity_id=salary_id
    )
    db.add(log)

    db.delete(salary)
    db.commit()

    return {"status": "success", "message": "Запись о зарплате удалена"}


# =========================
# ИСТОРИЯ ДЕЙСТВИЙ
# =========================

@app.get("/api/action-history/{entity_type}/{entity_id}")
async def get_action_history(
    entity_type: str,
    entity_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить историю действий по сущности"""
    history = db.query(ActionHistory).filter(
        ActionHistory.entity_type == entity_type,
        ActionHistory.entity_id == entity_id
    ).order_by(ActionHistory.action_date.desc()).all()

    # Добавляем user_name к каждой записи
    result = []
    for item in history:
        employee = db.query(Employee).filter(Employee.id == item.user_id).first()
        result.append({
            'id': item.id,
            'user_id': item.user_id,
            'user_name': employee.full_name if employee else 'Неизвестно',
            'action_type': item.action_type,
            'entity_type': item.entity_type,
            'entity_id': item.entity_id,
            'description': item.description,
            'action_date': item.action_date.strftime('%Y-%m-%d %H:%M:%S') if item.action_date else None
        })
    return result


@app.post("/api/action-history", response_model=ActionHistoryResponse)
async def create_action_history(
    history_data: ActionHistoryCreate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Создать запись в истории действий"""
    history = ActionHistory(
        user_id=current_user.id,
        **history_data.model_dump()
    )
    db.add(history)
    db.commit()
    db.refresh(history)
    return history


@app.get("/api/action-history", response_model=List[ActionHistoryResponse])
async def get_all_action_history(
    entity_type: Optional[str] = None,
    user_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить всю историю действий"""
    query = db.query(ActionHistory)
    if entity_type:
        query = query.filter(ActionHistory.entity_type == entity_type)
    if user_id:
        query = query.filter(ActionHistory.user_id == user_id)
    return query.order_by(ActionHistory.action_date.desc()).offset(skip).limit(limit).all()


# =========================
# ФАЙЛЫ ПРОЕКТА
# =========================

@app.get("/api/files/contract/{contract_id}", response_model=List[ProjectFileResponse])
async def get_contract_files(
    contract_id: int,
    stage: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить файлы договора"""
    query = db.query(ProjectFile).filter(ProjectFile.contract_id == contract_id)
    if stage:
        query = query.filter(ProjectFile.stage == stage)
    return query.order_by(ProjectFile.file_order).all()


@app.post("/api/files", response_model=ProjectFileResponse)
async def create_file_record(
    file_data: ProjectFileCreate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Создать запись о файле"""
    # Проверяем дубликат по (contract_id, yandex_path) перед вставкой
    data = file_data.model_dump()
    yp = data.get('yandex_path', '')
    if yp:
        existing = db.query(ProjectFile).filter(
            ProjectFile.contract_id == data.get('contract_id'),
            ProjectFile.yandex_path == yp
        ).first()
        if existing:
            # Дубликат — возвращаем существующую запись
            return existing

    file_record = ProjectFile(**data)
    db.add(file_record)
    try:
        db.commit()
    except Exception:
        db.rollback()
        # После rollback пробуем найти существующую запись
        if yp:
            existing = db.query(ProjectFile).filter(
                ProjectFile.contract_id == data.get('contract_id'),
                ProjectFile.yandex_path == yp
            ).first()
            if existing:
                return existing
        raise HTTPException(status_code=409, detail="Дубликат файла")
    db.refresh(file_record)
    return file_record


@app.delete("/api/files/{file_id}")
async def delete_file_record(
    file_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Удалить запись о файле и файл с Яндекс.Диска"""
    file_record = db.query(ProjectFile).filter(ProjectFile.id == file_id).first()
    if not file_record:
        raise HTTPException(status_code=404, detail="Файл не найден")

    # Удаляем файл с Яндекс.Диска (до удаления из БД!)
    yandex_path = file_record.yandex_path
    if yandex_path and yandex_disk_available:
        try:
            yd = get_yandex_disk_service()
            yd.delete_file(yandex_path)
            logger.info(f"Файл удалён с Яндекс.Диска: {yandex_path}")
        except Exception as e:
            error_str = str(e)
            if "DiskNotFoundError" not in error_str and "not found" not in error_str.lower():
                logger.error(f"Не удалось удалить файл с Яндекс.Диска: {e}")
                raise HTTPException(status_code=500, detail=f"Не удалось удалить файл с Яндекс.Диска: {e}")
            logger.info(f"Файл уже удалён с Яндекс.Диска: {yandex_path}")

    db.delete(file_record)
    db.commit()

    return {"status": "success", "message": "Запись о файле удалена"}


# =========================
# YANDEX DISK API
# =========================

try:
    from yandex_disk_service import get_yandex_disk_service
    yandex_disk_available = True
except ImportError:
    yandex_disk_available = False
    logger.warning("YandexDiskService not available")


@app.post("/api/files/upload")
async def upload_file_to_yandex(
    file: UploadFile = File(...),
    yandex_path: str = None,
    current_user: Employee = Depends(get_current_user),
):
    """Загрузить файл на Яндекс.Диск"""
    if not yandex_disk_available:
        raise HTTPException(status_code=503, detail="Yandex Disk service not available")

    try:
        yd_service = get_yandex_disk_service()
        if not yd_service.token:
            raise HTTPException(status_code=503, detail="Yandex Disk token not configured")
        file_bytes = await file.read()

        # Проверка размера файла
        max_size = int(os.environ.get("MAX_FILE_SIZE_MB", 50)) * 1024 * 1024
        if len(file_bytes) > max_size:
            raise HTTPException(
                status_code=413,
                detail=f"Размер файла превышает максимально допустимый ({os.environ.get('MAX_FILE_SIZE_MB', 50)} МБ)"
            )

        if not yandex_path:
            # Защита от path traversal в имени файла
            safe_filename = os.path.basename(file.filename or "unnamed")
            yandex_path = f"/uploads/{safe_filename}"
        else:
            # Защита от path traversal: запрещаем ".." в пути
            if ".." in yandex_path:
                raise HTTPException(status_code=400, detail="Недопустимый путь файла")

        result = yd_service.upload_file_from_bytes(file_bytes, yandex_path)

        if result:
            public_link = yd_service.get_public_link(yandex_path)
            return {
                "status": "success",
                "yandex_path": yandex_path,
                "public_link": public_link,
                "file_name": file.filename
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to upload file")

    except HTTPException:
        raise
    except Exception as e:
        error_msg = str(e).lower()
        if "unauthorized" in error_msg or "token" in error_msg or "401" in error_msg:
            raise HTTPException(status_code=503, detail="Yandex Disk not configured or token expired")
        raise HTTPException(status_code=500, detail=f"Upload error: {str(e)}")


@app.post("/api/files/folder")
async def create_yandex_folder(
    folder_path: str,
    current_user: Employee = Depends(get_current_user),
):
    """Создать папку на Яндекс.Диске"""
    if not yandex_disk_available:
        raise HTTPException(status_code=503, detail="Yandex Disk service not available")

    # Защита от path traversal
    if ".." in folder_path:
        raise HTTPException(status_code=400, detail="Недопустимый путь папки")

    try:
        yd_service = get_yandex_disk_service()
        result = yd_service.create_folder(folder_path)

        return {
            "status": "success" if result else "exists",
            "folder_path": folder_path
        }

    except HTTPException:
        raise
    except Exception as e:
        error_str = str(e)
        if "DiskNotFoundError" in error_str or "not found" in error_str.lower():
            raise HTTPException(status_code=404, detail=f"Путь не найден: {folder_path}")
        raise HTTPException(status_code=500, detail=f"Folder creation error: {error_str}")


@app.get("/api/files/public-link")
async def get_public_link(
    yandex_path: str,
    current_user: Employee = Depends(get_current_user),
):
    """Получить публичную ссылку на файл"""
    if not yandex_disk_available:
        raise HTTPException(status_code=503, detail="Yandex Disk service not available")

    try:
        yd_service = get_yandex_disk_service()
        public_link = yd_service.get_public_link(yandex_path)

        if public_link:
            return {
                "status": "success",
                "public_link": public_link,
                "yandex_path": yandex_path
            }
        else:
            raise HTTPException(status_code=404, detail="File not found or cannot create public link")

    except HTTPException:
        raise
    except Exception as e:
        error_str = str(e)
        if "DiskNotFoundError" in error_str or "not found" in error_str.lower():
            raise HTTPException(status_code=404, detail=f"Файл не найден: {yandex_path}")
        raise HTTPException(status_code=500, detail=f"Error getting public link: {error_str}")


@app.get("/api/files/list")
async def list_yandex_files(
    folder_path: Optional[str] = None,
    path: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
):
    """Получить список файлов в папке Яндекс.Диска"""
    # Принимаем и folder_path и path как алиасы
    resolved_path = folder_path or path
    if not resolved_path:
        raise HTTPException(status_code=422, detail="Необходимо указать folder_path или path")

    if not yandex_disk_available:
        raise HTTPException(status_code=503, detail="Yandex Disk service not available")

    try:
        yd_service = get_yandex_disk_service()
        files = yd_service.list_files(resolved_path)

        return {
            "status": "success",
            "folder_path": resolved_path,
            "files": files
        }

    except HTTPException:
        raise
    except Exception as e:
        error_str = str(e)
        if "DiskNotFoundError" in error_str or "not found" in error_str.lower():
            raise HTTPException(status_code=404, detail=f"Папка не найдена: {resolved_path}")
        raise HTTPException(status_code=500, detail=f"Error listing files: {error_str}")


@app.delete("/api/files/yandex")
async def delete_yandex_file(
    yandex_path: str,
    current_user: Employee = Depends(get_current_user),
):
    """Удалить файл с Яндекс.Диска"""
    if not yandex_disk_available:
        raise HTTPException(status_code=503, detail="Yandex Disk service not available")

    try:
        yd_service = get_yandex_disk_service()
        result = yd_service.delete_file(yandex_path)

        return {
            "status": "success" if result else "not_found",
            "yandex_path": yandex_path
        }

    except HTTPException:
        raise
    except Exception as e:
        error_str = str(e)
        if "DiskNotFoundError" in error_str or "not found" in error_str.lower():
            raise HTTPException(status_code=404, detail=f"Файл не найден: {yandex_path}")
        raise HTTPException(status_code=500, detail=f"Error deleting file: {error_str}")


@app.post("/api/files/scan/{contract_id}")
async def scan_contract_files_on_yandex(
    contract_id: int,
    scope: str = "all",
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Сканирование файлов на Яндекс.Диске для договора.

    Находит файлы, которые есть на ЯД но отсутствуют в БД, и создаёт записи.
    scope: 'all' — вся папка проекта, 'supervision' — только Авторский надзор.
    """
    if not yandex_disk_available:
        raise HTTPException(status_code=503, detail="Yandex Disk service not available")

    # Получаем договор для определения папки
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")

    folder_path = contract.yandex_folder_path
    if not folder_path:
        raise HTTPException(status_code=404, detail="Папка договора на ЯД не задана")

    # Защита от параллельных сканирований одного договора
    with _scanning_contracts_lock:
        if contract_id in _scanning_contracts:
            return {
                "status": "already_scanning",
                "total_on_disk": 0,
                "already_in_db": 0,
                "new_files_added": 0,
                "new_files": []
            }
        _scanning_contracts.add(contract_id)

    try:
        yd_service = get_yandex_disk_service()

        # Точный маппинг папок → стадий
        folder_to_stage_exact = {
            'Замер': 'measurement',
            'Замеры': 'measurement',
            '1 стадия - Планировочное решение': 'stage1',
            'Планировочное решение': 'stage1',
            'Концепция-коллажи': 'stage2_concept',
            'Коллажи': 'stage2_concept',
            '3D визуализация': 'stage2_3d',
            '3D': 'stage2_3d',
            '3 стадия - Чертежный проект': 'stage3',
            'Чертежный проект': 'stage3',
            'Чертежи': 'stage3',
            'Референсы': 'references',
            'Фотофиксация': 'photo_documentation',
            'Фото': 'photo_documentation',
            'Анкета': 'questionnaire',
            'Анкеты': 'questionnaire',
            'Документы': 'documents',
            'Техническое задание': 'tech_task',
            'ТЗ': 'tech_task',
            'Авторский надзор': 'supervision',
        }

        # Нечёткий маппинг: ключевые слова → стадия (для папок с нестандартными именами)
        folder_keywords_to_stage = [
            ('замер', 'measurement'),
            ('1 стадия', 'stage1'),
            ('1стадия', 'stage1'),
            ('планировочн', 'stage1'),
            ('концепция', 'stage2_concept'),
            ('коллаж', 'stage2_concept'),
            ('3d', 'stage2_3d'),
            ('визуализ', 'stage2_3d'),
            ('2 стадия', 'stage2_concept'),
            ('2стадия', 'stage2_concept'),
            ('3 стадия', 'stage3'),
            ('3стадия', 'stage3'),
            ('чертеж', 'stage3'),
            ('рабочи', 'stage3'),
            ('референ', 'references'),
            ('фотофикс', 'photo_documentation'),
            ('фото', 'photo_documentation'),
            ('анкет', 'questionnaire'),
            ('документ', 'documents'),
            ('техническ', 'tech_task'),
            ('надзор', 'supervision'),
        ]

        def match_folder_to_stage(folder_name):
            """Определить стадию по имени папки: сначала точное, потом нечёткое"""
            # Точное совпадение
            if folder_name in folder_to_stage_exact:
                return folder_to_stage_exact[folder_name]
            # Нечёткое: ищем ключевое слово в нижнем регистре
            name_lower = folder_name.lower()
            for keyword, stage_id in folder_keywords_to_stage:
                if keyword in name_lower:
                    return stage_id
            return None

        def detect_file_type(name):
            ext = name.rsplit('.', 1)[-1].lower() if '.' in name else ''
            if ext in ('png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp', 'tiff', 'svg'):
                return 'image'
            elif ext == 'pdf':
                return 'pdf'
            elif ext in ('xls', 'xlsx', 'csv'):
                return 'excel'
            elif ext in ('doc', 'docx'):
                return 'word'
            elif ext in ('dwg', 'dxf'):
                return 'cad'
            return 'other'

        def normalize_path(p):
            """Нормализация пути: убираем 'disk:' префикс для сравнения"""
            if p and p.startswith('disk:'):
                return p[5:]
            return p or ''

        found_files = []

        def scan_folder(path, stage=None):
            try:
                items = yd_service.list_files(path)
                for item in items:
                    item_name = item.get('name', '')
                    item_path = item.get('path', '')
                    item_type = item.get('type', '')

                    if item_type == 'dir':
                        # Пропускаем папку "правки" — файлы правок отображаются отдельно
                        if item_name.lower() == 'правки':
                            continue
                        child_stage = match_folder_to_stage(item_name)
                        if child_stage is None:
                            child_stage = stage  # наследуем стадию от родителя
                        # Внутри Авторского надзора подпапки "Стадия ..." остаются supervision
                        if stage == 'supervision' and item_name.startswith('Стадия'):
                            child_stage = 'supervision'
                        # Подпапки "Вариация N" наследуют стадию от родителя
                        if item_name.startswith('Вариация') or item_name.startswith('вариация'):
                            child_stage = stage
                        scan_folder(item_path, child_stage)
                    elif item_type == 'file':
                        # Файлы с определённой стадией добавляем
                        # Файлы в корне (stage=None) — пропускаем
                        if stage:
                            found_files.append({
                                'yandex_path': item_path,
                                'file_name': item_name,
                                'stage': stage,
                                'file_type': detect_file_type(item_name),
                            })
            except Exception as e:
                logger.warning(f"Ошибка сканирования {path}: {e}")

        if scope == 'supervision':
            # Для надзора сканируем только подпапку "Авторский надзор"
            supervision_path = folder_path.rstrip('/') + '/Авторский надзор'
            logger.info(f"Scan scope=supervision: сканируем только {supervision_path}")
            scan_folder(supervision_path, stage='supervision')
        else:
            scan_folder(folder_path)

        # Получаем существующие записи — нормализуем пути для сравнения
        existing_paths_normalized = set()
        existing_records = db.query(ProjectFile).filter(ProjectFile.contract_id == contract_id).all()
        for rec in existing_records:
            if rec.yandex_path:
                # Добавляем оба варианта пути (с disk: и без) для надёжного сравнения
                norm = normalize_path(rec.yandex_path)
                existing_paths_normalized.add(norm)
                if not norm.startswith('/'):
                    existing_paths_normalized.add('/' + norm)
                else:
                    existing_paths_normalized.add(norm.lstrip('/'))

        # Создаём записи для новых файлов (сравнение по нормализованному пути)
        new_files = []
        for f in found_files:
            yp = f['yandex_path']
            yp_normalized = normalize_path(yp)

            if yp_normalized in existing_paths_normalized:
                continue

            # Добавляем в множество чтобы не дублировать внутри одного скана
            existing_paths_normalized.add(yp_normalized)

            # Получаем публичную ссылку
            try:
                public_link = yd_service.get_public_link(yp)
            except Exception:
                public_link = ''

            # Для файлов надзора file_type хранит название стадии
            file_type_val = f['file_type']
            if f['stage'] == 'supervision':
                # Определяем стадию надзора из пути
                parts = yp.split('/')
                for part in parts:
                    if part.startswith('Стадия'):
                        file_type_val = part
                        break

            # Дополнительная проверка: прямой запрос в БД (защита от дубликатов)
            existing_exact = db.query(ProjectFile).filter(
                ProjectFile.contract_id == contract_id,
                ProjectFile.yandex_path == yp
            ).first()
            if existing_exact:
                logger.info(f"Scan: файл уже есть в БД (exact match), пропускаем: {f['file_name']}")
                continue

            new_record = ProjectFile(
                contract_id=contract_id,
                stage=f['stage'],
                file_type=file_type_val,
                yandex_path=yp,
                public_link=public_link,
                file_name=f['file_name']
            )
            try:
                # Используем savepoint чтобы rollback не затронул предыдущие записи
                savepoint = db.begin_nested()
                db.add(new_record)
                db.flush()
            except Exception as insert_err:
                savepoint.rollback()
                logger.warning(f"Scan: не удалось добавить файл (дубликат?): {f['file_name']}: {insert_err}")
                continue
            new_files.append({
                'yandex_path': yp,
                'file_name': f['file_name'],
                'stage': f['stage'],
                'file_type': file_type_val,
                'public_link': public_link,
            })

        # Для файлов из "Анкета" (questionnaire/tech_task): обновляем contract.tech_task_link
        tech_task_files = [f for f in new_files if f['stage'] in ('questionnaire', 'tech_task')]
        if tech_task_files and not contract.tech_task_link:
            first_tt = tech_task_files[0]
            contract.tech_task_link = first_tt.get('public_link', '')
            contract.tech_task_yandex_path = first_tt.get('yandex_path', '')
            contract.tech_task_file_name = first_tt.get('file_name', '')
            logger.info(f"Scan: обновлён tech_task_link для contract {contract_id}")

        # Обновляем references_yandex_path и photo_documentation_yandex_path
        # Логика: файлы есть → создать ссылку; файлов нет → очистить ссылку
        contract_updated = False

        if scope == 'all':
            ref_files = [f for f in found_files if f['stage'] == 'references']
            if ref_files:
                # Файлы есть — создаём ссылку если нет
                if not contract.references_yandex_path:
                    try:
                        first_ref_path = ref_files[0]['yandex_path']
                        ref_folder = '/'.join(first_ref_path.split('/')[:-1])
                        logger.info(f"Scan: публикуем папку референсов: {ref_folder}")
                        ref_link = yd_service.get_public_link(ref_folder)
                        if ref_link:
                            contract.references_yandex_path = ref_link
                            contract_updated = True
                            logger.info(f"Scan: обновлён references_yandex_path: {ref_link}")
                    except Exception as e:
                        logger.warning(f"Scan: не удалось получить ссылку на Референсы: {e}")
            else:
                # Файлов нет — очищаем ссылку если была
                if contract.references_yandex_path:
                    logger.info(f"Scan: папка референсов пуста, очищаем references_yandex_path")
                    contract.references_yandex_path = ''
                    contract_updated = True

            photo_files = [f for f in found_files if f['stage'] == 'photo_documentation']
            if photo_files:
                # Файлы есть — создаём ссылку если нет
                if not contract.photo_documentation_yandex_path:
                    try:
                        first_photo_path = photo_files[0]['yandex_path']
                        photo_folder = '/'.join(first_photo_path.split('/')[:-1])
                        logger.info(f"Scan: публикуем папку фотофиксации: {photo_folder}")
                        photo_link = yd_service.get_public_link(photo_folder)
                        if photo_link:
                            contract.photo_documentation_yandex_path = photo_link
                            contract_updated = True
                            logger.info(f"Scan: обновлён photo_documentation_yandex_path: {photo_link}")
                    except Exception as e:
                        logger.warning(f"Scan: не удалось получить ссылку на Фотофиксацию: {e}")
            else:
                # Файлов нет — очищаем ссылку если была
                if contract.photo_documentation_yandex_path:
                    logger.info(f"Scan: папка фотофиксации пуста, очищаем photo_documentation_yandex_path")
                    contract.photo_documentation_yandex_path = ''
                    contract_updated = True

        if new_files or contract_updated:
            db.commit()
            logger.info(f"Scan contract {contract_id}: новых файлов={len(new_files)}, contract_updated={contract_updated}")

        return {
            "status": "success",
            "total_on_disk": len(found_files),
            "already_in_db": len(existing_records),
            "new_files_added": len(new_files),
            "new_files": new_files,
            "contract_updated": contract_updated
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка сканирования: {str(e)}")
    finally:
        with _scanning_contracts_lock:
            _scanning_contracts.discard(contract_id)


@app.get("/api/files/updated")
async def get_updated_files(
    since: str = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить файлы, загруженные после указанного timestamp"""
    if not since:
        raise HTTPException(status_code=400, detail="Parameter 'since' is required")

    try:
        since_dt = datetime.fromisoformat(since.replace('Z', '+00:00'))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid timestamp format")

    files = db.query(ProjectFile).filter(
        ProjectFile.upload_date > since_dt
    ).all()

    return [
        {
            "id": f.id,
            "contract_id": f.contract_id,
            "stage": f.stage,
            "file_type": f.file_type,
            "public_link": f.public_link,
            "yandex_path": f.yandex_path,
            "file_name": f.file_name,
            "upload_date": f.upload_date.isoformat() if f.upload_date else None,
            "variation": f.variation
        }
        for f in files
    ]


@app.post("/api/files/validate")
async def validate_files(
    request: dict,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Пакетная проверка существования файлов на Яндекс.Диске"""
    file_ids = request.get("file_ids", [])
    auto_clean = request.get("auto_clean", False)

    if not file_ids:
        return []

    if len(file_ids) > 50:
        raise HTTPException(status_code=400, detail="Максимум 50 файлов за запрос")

    if not yandex_disk_available:
        raise HTTPException(status_code=503, detail="Yandex Disk service not available")

    yd = get_yandex_disk_service()
    results = []

    for file_id in file_ids:
        file_record = db.query(ProjectFile).filter(ProjectFile.id == file_id).first()
        if not file_record:
            results.append({"file_id": file_id, "exists": False, "reason": "not_in_db"})
            continue

        yandex_path = file_record.yandex_path
        if not yandex_path:
            results.append({"file_id": file_id, "exists": False, "reason": "no_path"})
            if auto_clean:
                db.delete(file_record)
            continue

        try:
            # Нормализация пути: Яндекс API принимает и disk: и без
            check_path = yandex_path
            exists = yd.file_exists(check_path)
            # Если не найден с disk: префиксом — попробуем без
            if not exists and check_path.startswith('disk:'):
                exists = yd.file_exists(check_path[5:])
            # И наоборот
            if not exists and not check_path.startswith('disk:'):
                exists = yd.file_exists('disk:' + check_path)
        except Exception as e:
            logger.warning(f"Ошибка проверки файла {file_id} на YD: {e}")
            results.append({"file_id": file_id, "exists": True, "reason": "check_error"})
            continue

        results.append({"file_id": file_id, "exists": exists})

        if not exists and auto_clean:
            db.delete(file_record)
            logger.info(f"Автоочистка: удалена запись файла {file_id} path={yandex_path} (нет на YD)")

    if auto_clean:
        db.commit()

    return results


# =========================
# СТАТИСТИКА И ОТЧЕТЫ
# =========================

@app.get("/api/statistics/dashboard")
async def get_dashboard_statistics(
    year: Optional[int] = None,
    month: Optional[int] = None,
    quarter: Optional[int] = None,
    agent_type: Optional[str] = None,
    city: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить статистику для дашборда"""
    try:
        from sqlalchemy import func, extract, and_

        # Базовый запрос договоров
        query = db.query(Contract)

        # Фильтры
        if year:
            query = query.filter(extract('year', Contract.created_at) == year)
        if month:
            query = query.filter(extract('month', Contract.created_at) == month)
        if quarter:
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            query = query.filter(extract('month', Contract.created_at).between(start_month, end_month))
        if agent_type and agent_type != 'Все':
            query = query.filter(Contract.agent_type == agent_type)
        if city and city != 'Все':
            query = query.filter(Contract.city == city)

        contracts = query.all()

        # Подсчет статистики
        total_contracts = len(contracts)
        total_amount = sum(c.total_amount or 0 for c in contracts)
        total_area = sum(c.area or 0 for c in contracts)

        # По типам проектов
        individual_count = len([c for c in contracts if c.project_type == 'Индивидуальный'])
        template_count = len([c for c in contracts if c.project_type == 'Шаблонный'])

        # По статусам
        status_counts = {}
        for c in contracts:
            status = c.status or 'Новый заказ'
            status_counts[status] = status_counts.get(status, 0) + 1

        # По городам
        city_counts = {}
        for c in contracts:
            c_city = c.city or 'Не указан'
            city_counts[c_city] = city_counts.get(c_city, 0) + 1

        # По месяцам (для графиков)
        monthly_data = {}
        for c in contracts:
            if c.created_at:
                month_key = c.created_at.strftime('%Y-%m')
                if month_key not in monthly_data:
                    monthly_data[month_key] = {'count': 0, 'amount': 0}
                monthly_data[month_key]['count'] += 1
                monthly_data[month_key]['amount'] += c.total_amount or 0

        # Активные CRM карточки
        active_cards = db.query(CRMCard).join(Contract).filter(
            ~Contract.status.in_(['СДАН', 'РАСТОРГНУТ', 'АВТОРСКИЙ НАДЗОР'])
        ).count()

        # Карточки надзора
        supervision_cards = db.query(SupervisionCard).join(Contract).filter(
            Contract.status == 'АВТОРСКИЙ НАДЗОР'
        ).count()

        return {
            'total_contracts': total_contracts,
            'total_amount': total_amount,
            'total_area': total_area,
            'individual_count': individual_count,
            'template_count': template_count,
            'status_counts': status_counts,
            'city_counts': city_counts,
            'monthly_data': monthly_data,
            'active_crm_cards': active_cards,
            'supervision_cards': supervision_cards
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения статистики: {str(e)}")


@app.get("/api/statistics/employees")
async def get_employee_statistics(
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить статистику по сотрудникам"""
    try:
        from sqlalchemy import func, extract, case

        employees = db.query(Employee).filter(Employee.status == 'активный').all()
        emp_ids = [emp.id for emp in employees]

        if not emp_ids:
            return []

        # Batch-load stage executor counts to avoid N+1 queries
        stage_query = db.query(
            StageExecutor.executor_id,
            func.count(StageExecutor.id).label('total'),
            func.count(case((StageExecutor.completed == True, 1))).label('completed')
        ).filter(StageExecutor.executor_id.in_(emp_ids))

        if year:
            stage_query = stage_query.filter(extract('year', StageExecutor.assigned_date) == year)
        if month:
            stage_query = stage_query.filter(extract('month', StageExecutor.assigned_date) == month)

        stage_counts = stage_query.group_by(StageExecutor.executor_id).all()
        stage_map = {sc[0]: {'total': sc[1], 'completed': sc[2]} for sc in stage_counts}

        # Batch-load salary totals to avoid N+1 queries
        salary_query = db.query(
            Salary.employee_id,
            func.sum(Salary.amount).label('total')
        ).filter(Salary.employee_id.in_(emp_ids))

        if year and month:
            report_month_str = f"{year}-{month:02d}"
            salary_query = salary_query.filter(Salary.report_month == report_month_str)

        salary_totals = salary_query.group_by(Salary.employee_id).all()
        salary_map = {st[0]: float(st[1]) if st[1] else 0 for st in salary_totals}

        result = []
        for emp in employees:
            stage_data = stage_map.get(emp.id, {'total': 0, 'completed': 0})
            total_stages = stage_data['total']
            completed_stages = stage_data['completed']
            total_salary = salary_map.get(emp.id, 0)

            result.append({
                'id': emp.id,
                'full_name': emp.full_name,
                'position': emp.position,
                'total_stages': total_stages,
                'completed_stages': completed_stages,
                'completion_rate': (completed_stages / total_stages * 100) if total_stages > 0 else 0,
                'total_salary': total_salary
            })

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения статистики: {str(e)}")


@app.get("/api/statistics/contracts-by-period")
async def get_contracts_by_period(
    year: int,
    group_by: str = "month",  # month, quarter, status
    project_type: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить договоры сгруппированные по периоду"""
    try:
        from sqlalchemy import func, extract

        query = db.query(Contract).filter(extract('year', Contract.created_at) == year)

        if project_type and project_type != 'Все':
            query = query.filter(Contract.project_type == project_type)

        contracts = query.all()

        if group_by == "month":
            result = {i: {'count': 0, 'amount': 0} for i in range(1, 13)}
            for c in contracts:
                if c.created_at:
                    m = c.created_at.month
                    result[m]['count'] += 1
                    result[m]['amount'] += c.total_amount or 0

        elif group_by == "quarter":
            result = {i: {'count': 0, 'amount': 0} for i in range(1, 5)}
            for c in contracts:
                if c.created_at:
                    q = (c.created_at.month - 1) // 3 + 1
                    result[q]['count'] += 1
                    result[q]['amount'] += c.total_amount or 0

        elif group_by == "status":
            result = {}
            for c in contracts:
                status = c.status or 'Новый заказ'
                if status not in result:
                    result[status] = {'count': 0, 'amount': 0}
                result[status]['count'] += 1
                result[status]['amount'] += c.total_amount or 0

        else:
            result = {}

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения данных: {str(e)}")


@app.get("/api/statistics/agent-types")
async def get_agent_types(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить список типов агентов"""
    try:
        result = db.query(Contract.agent_type).distinct().filter(
            Contract.agent_type != None,
            Contract.agent_type != ''
        ).all()
        return ['Все'] + [r[0] for r in result if r[0]]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.get("/api/statistics/cities")
async def get_cities(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить список городов"""
    try:
        result = db.query(Contract.city).distinct().filter(
            Contract.city != None,
            Contract.city != ''
        ).all()
        return ['Все'] + [r[0] for r in result if r[0]]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.post("/api/crm/cards/{card_id}/reset-stages")
async def reset_crm_card_stages(
    card_id: int,
    current_user: Employee = Depends(require_permission("crm_cards.reset_stages")),
    db: Session = Depends(get_db)
):
    """Сбросить выполнение стадий карточки"""
    try:
        card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="CRM карточка не найдена")

        # Сбрасываем все stage_executors
        stage_executors = db.query(StageExecutor).filter(StageExecutor.crm_card_id == card_id).all()
        for se in stage_executors:
            se.completed = False
            se.completed_date = None
            se.submitted_date = None

        db.commit()

        return {"status": "success", "message": "Стадии сброшены", "card_id": card_id}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка сброса стадий: {str(e)}")


@app.post("/api/crm/cards/{card_id}/reset-approval")
async def reset_crm_card_approval(
    card_id: int,
    current_user: Employee = Depends(require_permission("crm_cards.reset_approval")),
    db: Session = Depends(get_db)
):
    """Сбросить стадии согласования карточки"""
    try:
        card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="CRM карточка не найдена")

        # Сбрасываем согласования
        card.is_approved = False
        card.approval_stages = None
        card.approval_deadline = None

        db.commit()

        return {"status": "success", "message": "Согласования сброшены", "card_id": card_id}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка сброса согласований: {str(e)}")


# =========================
# SUPERVISION EXTENDED ENDPOINTS
# =========================

@app.post("/api/supervision/cards/{card_id}/reset-stages")
async def reset_supervision_card_stages(
    card_id: int,
    current_user: Employee = Depends(require_permission("supervision.reset_stages")),
    db: Session = Depends(get_db)
):
    """Сбросить выполнение стадий надзора"""
    try:
        card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

        # Сбрасываем dan_completed
        card.dan_completed = False

        db.commit()

        return {"status": "success", "message": "Стадии надзора сброшены", "card_id": card_id}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка сброса стадий надзора: {str(e)}")


@app.post("/api/supervision/cards/{card_id}/complete-stage")
async def complete_supervision_stage(
    card_id: int,
    current_user: Employee = Depends(require_permission("supervision.complete_stage")),
    db: Session = Depends(get_db)
):
    """Завершить стадию надзора"""
    try:
        card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

        card.dan_completed = True

        # Добавляем запись в историю
        history = SupervisionProjectHistory(
            supervision_card_id=card_id,
            entry_type="stage_completed",
            message="Стадия завершена",
            created_by=current_user.id
        )
        db.add(history)

        db.commit()

        return {"status": "success", "message": "Стадия завершена", "card_id": card_id}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка завершения стадии: {str(e)}")


@app.delete("/api/supervision/orders/{supervision_card_id}")
async def delete_supervision_order(
    supervision_card_id: int,
    contract_id: int,
    current_user: Employee = Depends(require_permission("supervision.delete_order")),
    db: Session = Depends(get_db)
):
    """Удалить заказ надзора"""
    try:
        card = db.query(SupervisionCard).filter(SupervisionCard.id == supervision_card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

        # Удаляем историю
        db.query(SupervisionProjectHistory).filter(
            SupervisionProjectHistory.supervision_card_id == supervision_card_id
        ).delete()

        # Удаляем связанные платежи
        db.query(Payment).filter(Payment.supervision_card_id == supervision_card_id).delete()

        # Удаляем карточку
        db.delete(card)
        db.commit()

        return {"status": "success", "message": "Заказ надзора удален"}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка удаления заказа надзора: {str(e)}")


@app.get("/api/payments/supervision/{contract_id}")
async def get_payments_for_supervision(
    contract_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить платежи для надзора - ИСПРАВЛЕНО 06.02.2026"""
    # ИСПРАВЛЕНИЕ: Фильтруем по supervision_card_id через JOIN с SupervisionCard
    # Платежи надзора имеют payment_type='Полная оплата' или другой тип
    payments = db.query(Payment).join(
        SupervisionCard, Payment.supervision_card_id == SupervisionCard.id
    ).filter(
        SupervisionCard.contract_id == contract_id
    ).all()

    result = []
    for payment in payments:
        employee = db.query(Employee).filter(Employee.id == payment.employee_id).first()
        result.append({
            'id': payment.id,
            'contract_id': payment.contract_id,
            'supervision_card_id': payment.supervision_card_id,
            'employee_id': payment.employee_id,
            'employee_name': employee.full_name if employee else 'Неизвестный',
            'role': payment.role,
            'stage_name': payment.stage_name,
            'calculated_amount': float(payment.calculated_amount) if payment.calculated_amount else 0,
            'manual_amount': float(payment.manual_amount) if payment.manual_amount else None,
            'final_amount': float(payment.final_amount) if payment.final_amount else 0,
            'is_manual': payment.is_manual,
            'payment_type': payment.payment_type,
            'report_month': payment.report_month,
            'payment_status': payment.payment_status,
            'is_paid': payment.is_paid,
            'reassigned': payment.reassigned if hasattr(payment, 'reassigned') else False,
            'old_employee_id': payment.old_employee_id if hasattr(payment, 'old_employee_id') else None,
        })

    return result


@app.get("/api/payments/by-supervision-card/{supervision_card_id}")
async def get_payments_by_supervision_card(
    supervision_card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """ДОБАВЛЕНО 30.01.2026: Получить платежи по ID карточки надзора"""
    payments = db.query(Payment).filter(
        Payment.supervision_card_id == supervision_card_id
    ).all()

    result = []
    for payment in payments:
        employee = db.query(Employee).filter(Employee.id == payment.employee_id).first()
        result.append({
            'id': payment.id,
            'contract_id': payment.contract_id,
            'supervision_card_id': payment.supervision_card_id,
            'employee_id': payment.employee_id,
            'employee_name': employee.full_name if employee else 'Неизвестный',
            'role': payment.role,
            'stage_name': payment.stage_name,
            'calculated_amount': float(payment.calculated_amount) if payment.calculated_amount else 0,
            'manual_amount': float(payment.manual_amount) if payment.manual_amount else None,
            'final_amount': float(payment.final_amount) if payment.final_amount else 0,
            'is_manual': payment.is_manual,
            'payment_type': payment.payment_type,
            'report_month': payment.report_month,
            'payment_status': payment.payment_status,
            'is_paid': payment.is_paid,
            'reassigned': payment.reassigned if hasattr(payment, 'reassigned') else False,
            'old_employee_id': payment.old_employee_id if hasattr(payment, 'old_employee_id') else None,
        })

    return result


@app.patch("/api/payments/{payment_id}/manual")
async def update_payment_manual(
    payment_id: int,
    data: PaymentManualUpdateRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить платеж вручную"""
    try:
        payment = db.query(Payment).filter(Payment.id == payment_id).first()
        if not payment:
            raise HTTPException(status_code=404, detail="Платеж не найден")

        payment.manual_amount = data.amount
        payment.final_amount = data.amount
        payment.is_manual = True
        payment.report_month = data.report_month
        payment.updated_at = datetime.utcnow()

        db.commit()
        db.refresh(payment)

        return {
            'id': payment.id,
            'manual_amount': float(payment.manual_amount),
            'final_amount': float(payment.final_amount),
            'report_month': payment.report_month
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка обновления платежа: {str(e)}")


@app.patch("/api/payments/{payment_id}/mark-paid")
async def mark_payment_as_paid(
    payment_id: int,
    employee_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Отметить платеж как выплаченный"""
    try:
        payment = db.query(Payment).filter(Payment.id == payment_id).first()
        if not payment:
            raise HTTPException(status_code=404, detail="Платеж не найден")

        payment.is_paid = True
        payment.paid_date = datetime.utcnow()
        payment.paid_by = current_user.id
        payment.payment_status = 'paid'
        payment.updated_at = datetime.utcnow()

        db.commit()

        return {
            'id': payment.id,
            'is_paid': payment.is_paid,
            'paid_date': payment.paid_date.isoformat(),
            'paid_by': payment.paid_by
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка отметки платежа: {str(e)}")


@app.get("/api/crm/cards/{card_id}/submitted-stages")
async def get_submitted_stages(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить отправленные стадии карточки"""
    stages = db.query(StageExecutor).filter(
        StageExecutor.crm_card_id == card_id,
        StageExecutor.submitted_date != None
    ).all()

    return [{
        'id': s.id,
        'stage_name': s.stage_name,
        'executor_id': s.executor_id,
        'executor_name': s.executor.full_name if s.executor else None,
        'submitted_date': s.submitted_date.isoformat() if s.submitted_date else None,
        'completed': s.completed,
        'completed_date': s.completed_date.isoformat() if s.completed_date else None,
    } for s in stages]


@app.get("/api/crm/cards/{card_id}/stage-history")
async def get_stage_history(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить историю стадий карточки"""
    # История из ActionHistory для данной карточки
    history = db.query(ActionHistory).filter(
        ActionHistory.entity_type == 'stage_executor',
        ActionHistory.entity_id == card_id
    ).order_by(ActionHistory.action_date.desc()).all()

    return [{
        'id': h.id,
        'action_type': h.action_type,
        'old_value': h.old_value,
        'new_value': h.new_value,
        'action_date': h.action_date.isoformat() if h.action_date else None,
        'user_id': h.user_id
    } for h in history]


@app.get("/api/supervision/addresses")
async def get_supervision_addresses(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить адреса надзора"""
    try:
        result = db.query(Contract.address).join(
            SupervisionCard, SupervisionCard.contract_id == Contract.id
        ).distinct().filter(
            Contract.address != None,
            Contract.address != ''
        ).all()
        return [r[0] for r in result if r[0]]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.get("/api/supervision/cards/{card_id}/contract")
async def get_contract_id_by_supervision_card(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить ID договора по ID карточки надзора"""
    card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка надзора не найдена")
    return {"contract_id": card.contract_id}


@app.get("/api/statistics/projects")
async def get_project_statistics(
    project_type: str,
    year: Optional[int] = None,
    quarter: Optional[int] = None,
    month: Optional[int] = None,
    agent_type: Optional[str] = None,
    city: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить статистику проектов (формат совместим с db_manager)"""
    try:
        from sqlalchemy import func, cast, Date, Integer
        from sqlalchemy.sql import text

        # Базовый запрос
        query = db.query(Contract).filter(Contract.project_type == project_type)

        # Фильтр по дате договора (contract_date хранится как строка YYYY-MM-DD)
        # Используем CAST к DATE для PostgreSQL или работу со строками
        if year or month or quarter:
            # Фильтруем только договоры с валидными датами
            query = query.filter(
                Contract.contract_date.isnot(None),
                Contract.contract_date != ''
            )

        if year:
            # Приводим contract_date к DATE и извлекаем год
            query = query.filter(
                func.extract('year', cast(Contract.contract_date, Date)) == year
            )
        if month:
            query = query.filter(
                func.extract('month', cast(Contract.contract_date, Date)) == month
            )
        if quarter:
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            query = query.filter(
                func.extract('year', cast(Contract.contract_date, Date)) == year,
                func.extract('month', cast(Contract.contract_date, Date)).between(start_month, end_month)
            )
        if agent_type and agent_type != 'Все':
            query = query.filter(Contract.agent_type == agent_type)
        if city and city != 'Все':
            query = query.filter(Contract.city == city)

        contracts = query.all()

        # Подсчёт статистики
        total_orders = len(contracts)
        total_area = sum(c.area or 0 for c in contracts)

        # Активные (не сданы и не расторгнуты)
        active = len([c for c in contracts if c.status not in ['СДАН', 'АВТОРСКИЙ НАДЗОР', 'РАСТОРГНУТ'] or c.status is None or c.status == ''])

        # Выполненные (СДАН или АВТОРСКИЙ НАДЗОР)
        completed = len([c for c in contracts if c.status in ['СДАН', 'АВТОРСКИЙ НАДЗОР']])

        # Расторгнуты
        cancelled = len([c for c in contracts if c.status == 'РАСТОРГНУТ'])

        # Просрочки - считаем договоры с незавершёнными просроченными этапами
        overdue = 0
        try:
            from datetime import date as date_today
            overdue_query = db.query(func.count(func.distinct(Contract.id))).join(
                CRMCard, CRMCard.contract_id == Contract.id
            ).join(
                StageExecutor, StageExecutor.crm_card_id == CRMCard.id
            ).filter(
                Contract.project_type == project_type,
                StageExecutor.completed == False,
                StageExecutor.deadline < date_today.today()
            )
            # Применяем те же фильтры
            if year or month or quarter:
                # Фильтруем только договоры с валидными датами
                overdue_query = overdue_query.filter(
                    Contract.contract_date.isnot(None),
                    Contract.contract_date != ''
                )

            if year:
                overdue_query = overdue_query.filter(
                    func.extract('year', cast(Contract.contract_date, Date)) == year
                )
            if month:
                overdue_query = overdue_query.filter(
                    func.extract('month', cast(Contract.contract_date, Date)) == month
                )
            if quarter:
                start_month = (quarter - 1) * 3 + 1
                end_month = quarter * 3
                overdue_query = overdue_query.filter(
                    func.extract('year', cast(Contract.contract_date, Date)) == year,
                    func.extract('month', cast(Contract.contract_date, Date)).between(start_month, end_month)
                )
            if agent_type and agent_type != 'Все':
                overdue_query = overdue_query.filter(Contract.agent_type == agent_type)
            if city and city != 'Все':
                overdue_query = overdue_query.filter(Contract.city == city)

            overdue = overdue_query.scalar() or 0
        except Exception:
            overdue = 0

        # По городам
        by_cities = {}
        for c in contracts:
            if c.city:
                by_cities[c.city] = by_cities.get(c.city, 0) + 1

        # По агентам
        by_agents = {}
        for c in contracts:
            if c.agent_type:
                by_agents[c.agent_type] = by_agents.get(c.agent_type, 0) + 1

        return {
            'total_orders': total_orders,
            'total_area': float(total_area),
            'active': active,
            'completed': completed,
            'cancelled': cancelled,
            'overdue': overdue,
            'by_cities': by_cities,
            'by_agents': by_agents,
            'by_stages': {}  # TODO: реализовать если нужно
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения статистики: {str(e)}")


@app.get("/api/statistics/supervision")
async def get_supervision_statistics(
    year: Optional[int] = None,
    quarter: Optional[int] = None,
    month: Optional[int] = None,
    agent_type: Optional[str] = None,
    city: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить статистику авторского надзора (формат совместим с db_manager)"""
    try:
        from sqlalchemy import func, cast, Date
        from datetime import date as date_type

        # Базовый запрос
        query = db.query(SupervisionCard).join(Contract, SupervisionCard.contract_id == Contract.id)

        # Фильтры по дате договора (contract_date хранится как строка YYYY-MM-DD)
        if year or month or quarter:
            # Фильтруем только договоры с валидными датами
            query = query.filter(
                Contract.contract_date.isnot(None),
                Contract.contract_date != ''
            )

        if year:
            query = query.filter(
                func.extract('year', cast(Contract.contract_date, Date)) == year
            )
        if month:
            query = query.filter(
                func.extract('month', cast(Contract.contract_date, Date)) == month
            )
        if quarter:
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            query = query.filter(
                func.extract('year', cast(Contract.contract_date, Date)) == year,
                func.extract('month', cast(Contract.contract_date, Date)).between(start_month, end_month)
            )
        if agent_type and agent_type != 'Все':
            query = query.filter(Contract.agent_type == agent_type)
        if city and city != 'Все':
            query = query.filter(Contract.city == city)

        cards = query.all()

        # Подсчёт статистики
        total_orders = len(cards)
        total_area = sum(c.contract.area or 0 for c in cards)

        # Активные (статус АВТОРСКИЙ НАДЗОР)
        active = len([c for c in cards if c.contract.status == 'АВТОРСКИЙ НАДЗОР'])

        # Выполненные (статус СДАН)
        completed = len([c for c in cards if c.contract.status == 'СДАН'])

        # Расторгнуты
        cancelled = len([c for c in cards if c.contract.status == 'РАСТОРГНУТ'])

        # Просрочки (дедлайн прошел, статус АВТОРСКИЙ НАДЗОР)
        today = date_type.today()
        overdue = len([
            c for c in cards
            if c.deadline and c.deadline < today and c.contract.status == 'АВТОРСКИЙ НАДЗОР'
        ])

        # По городам
        by_cities = {}
        for c in cards:
            if c.contract.city:
                by_cities[c.contract.city] = by_cities.get(c.contract.city, 0) + 1

        # По агентам
        by_agents = {}
        for c in cards:
            if c.contract.agent_type:
                by_agents[c.contract.agent_type] = by_agents.get(c.contract.agent_type, 0) + 1

        return {
            'total_orders': total_orders,
            'total_area': float(total_area),
            'active': active,
            'completed': completed,
            'cancelled': cancelled,
            'overdue': overdue,
            'by_cities': by_cities,
            'by_agents': by_agents,
            'by_stages': {}  # TODO: реализовать если нужно
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения статистики надзора: {str(e)}")


@app.get("/api/statistics/supervision/filtered")
async def get_supervision_statistics_filtered(
    year: Optional[int] = None,
    quarter: Optional[int] = None,
    month: Optional[int] = None,
    agent_type: Optional[str] = None,
    city: Optional[str] = None,
    address: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить отфильтрованную статистику надзора"""
    try:
        from sqlalchemy import extract

        query = db.query(SupervisionCard).join(Contract)

        if year:
            query = query.filter(extract('year', SupervisionCard.created_at) == year)
        if month:
            query = query.filter(extract('month', SupervisionCard.created_at) == month)
        if quarter:
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            query = query.filter(extract('month', SupervisionCard.created_at).between(start_month, end_month))
        if agent_type and agent_type != 'Все':
            query = query.filter(Contract.agent_type == agent_type)
        if city and city != 'Все':
            query = query.filter(Contract.city == city)
        if address:
            query = query.filter(Contract.address.ilike(f'%{address}%'))

        cards = query.all()

        total_count = len(cards)
        total_area = sum(c.contract.area or 0 for c in cards)

        return {
            'total_count': total_count,
            'total_area': total_area,
            'cards': [{
                'id': c.id,
                'contract_number': c.contract.contract_number,
                'address': c.contract.address,
                'area': c.contract.area,
                'status': c.contract.status
            } for c in cards]
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения статистики: {str(e)}")


# =========================
# RATES EXTENDED ENDPOINTS
# =========================

# GET /api/rates/template перемещен выше /api/rates/{rate_id} для корректной маршрутизации

@app.post("/api/rates/template")
async def save_template_rate(
    data: TemplateRateRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Сохранить шаблонный тариф"""
    try:
        # Проверяем существующий тариф
        existing = db.query(Rate).filter(
            Rate.project_type == 'Шаблонный',
            Rate.role == data.role,
            Rate.area_from == data.area_from,
            Rate.area_to == data.area_to
        ).first()

        if existing:
            existing.fixed_price = data.price  # ИСПРАВЛЕНО: price -> fixed_price
            existing.updated_at = datetime.utcnow()
            db.commit()
            db.refresh(existing)
            return existing
        else:
            rate = Rate(
                project_type='Шаблонный',
                role=data.role,
                area_from=data.area_from,
                area_to=data.area_to,
                fixed_price=data.price  # ИСПРАВЛЕНО: price -> fixed_price
            )
            db.add(rate)
            db.commit()
            db.refresh(rate)
            return rate

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка сохранения тарифа: {str(e)}")


@app.post("/api/rates/individual")
async def save_individual_rate(
    data: IndividualRateRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Сохранить индивидуальный тариф"""
    try:
        # Проверяем существующий тариф
        query = db.query(Rate).filter(
            Rate.project_type == 'Индивидуальный',
            Rate.role == data.role
        )
        if data.stage_name:
            query = query.filter(Rate.stage_name == data.stage_name)
        else:
            query = query.filter(Rate.stage_name.is_(None))  # ИСПРАВЛЕНО: добавлен фильтр IS NULL

        existing = query.first()

        if existing:
            existing.rate_per_m2 = data.rate_per_m2
            existing.updated_at = datetime.utcnow()
            db.commit()
            db.refresh(existing)
            return existing
        else:
            rate = Rate(
                project_type='Индивидуальный',
                role=data.role,
                rate_per_m2=data.rate_per_m2,
                stage_name=data.stage_name
            )
            db.add(rate)
            db.commit()
            db.refresh(rate)
            return rate

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка сохранения тарифа: {str(e)}")


@app.delete("/api/rates/individual")
async def delete_individual_rate(
    role: str,
    stage_name: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Удалить индивидуальный тариф"""
    try:
        query = db.query(Rate).filter(
            Rate.project_type == 'Индивидуальный',
            Rate.role == role
        )
        if stage_name:
            query = query.filter(Rate.stage_name == stage_name)

        deleted = query.delete()
        db.commit()

        return {"status": "success", "deleted_count": deleted}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка удаления тарифа: {str(e)}")


@app.post("/api/rates/supervision")
async def save_supervision_rate(
    data: SupervisionRateRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Сохранить тариф надзора - ИСПРАВЛЕНО: создаем отдельные записи для каждой роли"""
    try:
        results = []

        # Сохраняем тариф для ДАН (исполнитель)
        if data.executor_rate is not None:
            existing_dan = db.query(Rate).filter(
                Rate.project_type == 'Авторский надзор',
                Rate.role == 'ДАН',
                Rate.stage_name == data.stage_name
            ).first()

            if existing_dan:
                existing_dan.rate_per_m2 = data.executor_rate
                existing_dan.updated_at = datetime.utcnow()
            else:
                rate_dan = Rate(
                    project_type='Авторский надзор',
                    role='ДАН',
                    stage_name=data.stage_name,
                    rate_per_m2=data.executor_rate
                )
                db.add(rate_dan)
            results.append({'role': 'ДАН', 'rate': data.executor_rate})

        # Сохраняем тариф для Старшего менеджера
        if data.manager_rate is not None:
            existing_manager = db.query(Rate).filter(
                Rate.project_type == 'Авторский надзор',
                Rate.role == 'Старший менеджер проектов',
                Rate.stage_name == data.stage_name
            ).first()

            if existing_manager:
                existing_manager.rate_per_m2 = data.manager_rate
                existing_manager.updated_at = datetime.utcnow()
            else:
                rate_manager = Rate(
                    project_type='Авторский надзор',
                    role='Старший менеджер проектов',
                    stage_name=data.stage_name,
                    rate_per_m2=data.manager_rate
                )
                db.add(rate_manager)
            results.append({'role': 'Старший менеджер проектов', 'rate': data.manager_rate})

        db.commit()
        return {'status': 'success', 'stage_name': data.stage_name, 'rates': results}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка сохранения тарифа: {str(e)}")


@app.post("/api/rates/surveyor")
async def save_surveyor_rate(
    data: SurveyorRateRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Сохранить тариф замерщика"""
    try:
        existing = db.query(Rate).filter(
            Rate.role == 'Замерщик',
            Rate.city == data.city
        ).first()

        if existing:
            existing.surveyor_price = data.price  # ИСПРАВЛЕНО: price -> surveyor_price
            existing.updated_at = datetime.utcnow()
            db.commit()
            db.refresh(existing)
            return existing
        else:
            rate = Rate(
                role='Замерщик',
                city=data.city,
                surveyor_price=data.price  # ИСПРАВЛЕНО: price -> surveyor_price
            )
            db.add(rate)
            db.commit()
            db.refresh(rate)
            return rate

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка сохранения тарифа: {str(e)}")


@app.get("/api/agents")
async def get_all_agents(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить список всех агентов"""
    agents = db.query(Employee).filter(
        or_(
            Employee.position == 'Агент',
            Employee.secondary_position == 'Агент'
        )
    ).all()
    return [{
        "id": a.id,
        "name": a.full_name,
        "full_name": a.full_name,
        "color": a.agent_color or "#FFFFFF",
        "status": a.status
    } for a in agents]


@app.get("/api/reports/employee")
async def get_employee_report_data(
    employee_id: int,
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить данные для отчета сотрудника"""
    try:
        from sqlalchemy import extract

        employee = db.query(Employee).filter(Employee.id == employee_id).first()
        if not employee:
            raise HTTPException(status_code=404, detail="Сотрудник не найден")

        # Получаем назначения на стадии
        stage_query = db.query(StageExecutor).filter(StageExecutor.executor_id == employee_id)
        if year:
            stage_query = stage_query.filter(extract('year', StageExecutor.assigned_date) == year)
        if month:
            stage_query = stage_query.filter(extract('month', StageExecutor.assigned_date) == month)

        stages = stage_query.all()

        # Получаем платежи
        payment_query = db.query(Payment).filter(Payment.employee_id == employee_id)
        if year and month:
            report_month = f"{year}-{month:02d}"
            payment_query = payment_query.filter(Payment.report_month == report_month)

        payments = payment_query.all()

        total_stages = len(stages)
        completed_stages = len([s for s in stages if s.completed])
        total_payments = sum(p.final_amount or 0 for p in payments)

        return {
            'employee_id': employee_id,
            'employee_name': employee.full_name,
            'position': employee.position,
            'total_stages': total_stages,
            'completed_stages': completed_stages,
            'completion_rate': (completed_stages / total_stages * 100) if total_stages > 0 else 0,
            'total_payments': total_payments,
            'stages': [{
                'id': s.id,
                'stage_name': s.stage_name,
                'completed': s.completed,
                'assigned_date': s.assigned_date.isoformat() if s.assigned_date else None,
                'completed_date': s.completed_date.isoformat() if s.completed_date else None,
            } for s in stages],
            'payments': [{
                'id': p.id,
                'payment_type': p.payment_type,
                'amount': float(p.final_amount) if p.final_amount else 0,
                'report_month': p.report_month,
                'is_paid': p.is_paid
            } for p in payments]
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения отчета: {str(e)}")


@app.get("/api/reports/employee-report")
async def get_employee_report_by_type(
    project_type: str,
    period: str,
    year: int,
    quarter: Optional[int] = None,
    month: Optional[int] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Получить данные для отчета по сотрудникам

    Сигнатура совпадает с database/db_manager.py:get_employee_report_data()

    Args:
        project_type: Тип проекта ('Индивидуальный' или 'Шаблонный')
        period: Период ('За год', 'За квартал', 'За месяц')
        year: Год
        quarter: Квартал (1-4)
        month: Месяц (1-12)

    Returns:
        completed: Список выполненных заказов по сотрудникам
        area: Список площадей по сотрудникам
        deadlines: Список просрочек
        salaries: Список зарплат
    """
    try:
        from sqlalchemy import extract, func, and_

        # Строим условие по периоду
        def build_period_condition(date_column):
            """Строит условие фильтрации по периоду"""
            conditions = [extract('year', date_column) == year]

            if period == 'За квартал' and quarter:
                start_month = (quarter - 1) * 3 + 1
                end_month = quarter * 3
                conditions.append(extract('month', date_column).between(start_month, end_month))
            elif period == 'За месяц' and month:
                conditions.append(extract('month', date_column) == month)

            return and_(*conditions) if len(conditions) > 1 else conditions[0]

        # 1. Выполненные заказы - группировка по сотрудникам
        completed_query = db.query(
            Employee.full_name.label('employee_name'),
            Employee.position,
            func.count(StageExecutor.id).label('count')
        ).join(
            StageExecutor, StageExecutor.executor_id == Employee.id
        ).join(
            CRMCard, StageExecutor.crm_card_id == CRMCard.id
        ).join(
            Contract, CRMCard.contract_id == Contract.id
        ).filter(
            Contract.project_type == project_type,
            StageExecutor.completed == True
        )

        # Добавляем фильтр по периоду для completed_date
        if period == 'За квартал' and quarter:
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            completed_query = completed_query.filter(
                extract('year', StageExecutor.completed_date) == year,
                extract('month', StageExecutor.completed_date).between(start_month, end_month)
            )
        elif period == 'За месяц' and month:
            completed_query = completed_query.filter(
                extract('year', StageExecutor.completed_date) == year,
                extract('month', StageExecutor.completed_date) == month
            )
        else:  # За год
            completed_query = completed_query.filter(
                extract('year', StageExecutor.completed_date) == year
            )

        completed_data = completed_query.group_by(
            StageExecutor.executor_id,
            Employee.full_name,
            Employee.position
        ).order_by(func.count(StageExecutor.id).desc()).all()

        completed = [
            {'employee_name': row.employee_name, 'position': row.position, 'count': row.count}
            for row in completed_data
        ]

        # 2. Площадь по сотрудникам
        area_query = db.query(
            Employee.full_name.label('employee_name'),
            Employee.position,
            func.sum(Contract.area).label('total_area')
        ).join(
            StageExecutor, StageExecutor.executor_id == Employee.id
        ).join(
            CRMCard, StageExecutor.crm_card_id == CRMCard.id
        ).join(
            Contract, CRMCard.contract_id == Contract.id
        ).filter(
            Contract.project_type == project_type,
            StageExecutor.completed == True
        )

        # Добавляем фильтр по периоду
        if period == 'За квартал' and quarter:
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            area_query = area_query.filter(
                extract('year', StageExecutor.completed_date) == year,
                extract('month', StageExecutor.completed_date).between(start_month, end_month)
            )
        elif period == 'За месяц' and month:
            area_query = area_query.filter(
                extract('year', StageExecutor.completed_date) == year,
                extract('month', StageExecutor.completed_date) == month
            )
        else:  # За год
            area_query = area_query.filter(
                extract('year', StageExecutor.completed_date) == year
            )

        area_data = area_query.group_by(
            StageExecutor.executor_id,
            Employee.full_name,
            Employee.position
        ).order_by(func.sum(Contract.area).desc()).all()

        area = [
            {'employee_name': row.employee_name, 'position': row.position, 'total_area': float(row.total_area or 0)}
            for row in area_data
        ]

        # 3. Просрочки дедлайнов (PostgreSQL-совместимый вариант)
        # Используем EXTRACT(EPOCH FROM ...) / 86400 для получения дней
        deadlines_query = db.query(
            Employee.full_name.label('employee_name'),
            func.count().label('overdue_count'),
            func.avg(
                func.extract('epoch', StageExecutor.completed_date - StageExecutor.deadline) / 86400.0
            ).label('avg_overdue_days')
        ).join(
            StageExecutor, StageExecutor.executor_id == Employee.id
        ).join(
            CRMCard, StageExecutor.crm_card_id == CRMCard.id
        ).join(
            Contract, CRMCard.contract_id == Contract.id
        ).filter(
            Contract.project_type == project_type,
            StageExecutor.completed == True,
            StageExecutor.completed_date > StageExecutor.deadline
        )

        # Добавляем фильтр по периоду
        if period == 'За квартал' and quarter:
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            deadlines_query = deadlines_query.filter(
                extract('year', StageExecutor.completed_date) == year,
                extract('month', StageExecutor.completed_date).between(start_month, end_month)
            )
        elif period == 'За месяц' and month:
            deadlines_query = deadlines_query.filter(
                extract('year', StageExecutor.completed_date) == year,
                extract('month', StageExecutor.completed_date) == month
            )
        else:  # За год
            deadlines_query = deadlines_query.filter(
                extract('year', StageExecutor.completed_date) == year
            )

        deadlines_data = deadlines_query.group_by(
            StageExecutor.executor_id,
            Employee.full_name
        ).order_by(func.count().desc()).all()

        deadlines = [
            {
                'employee_name': row.employee_name,
                'overdue_count': row.overdue_count,
                'avg_overdue_days': float(row.avg_overdue_days or 0)
            }
            for row in deadlines_data
        ]

        # 4. Зарплаты по сотрудникам
        salaries_query = db.query(
            Employee.full_name.label('employee_name'),
            Employee.position,
            func.sum(Salary.amount).label('total_salary')
        ).join(
            Salary, Salary.employee_id == Employee.id
        ).filter(
            Salary.payment_type.ilike(f'%{project_type}%')
        )

        # Фильтр по report_month
        if period == 'За квартал' and quarter:
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            months_pattern = '|'.join([f'{year}-{m:02d}' for m in range(start_month, end_month + 1)])
            salaries_query = salaries_query.filter(
                Salary.report_month.op('~')(months_pattern)
            )
        elif period == 'За месяц' and month:
            salaries_query = salaries_query.filter(
                Salary.report_month == f'{year}-{month:02d}'
            )
        else:  # За год
            salaries_query = salaries_query.filter(
                Salary.report_month.like(f'{year}-%')
            )

        salaries_data = salaries_query.group_by(
            Salary.employee_id,
            Employee.full_name,
            Employee.position
        ).order_by(func.sum(Salary.amount).desc()).all()

        salaries = [
            {'employee_name': row.employee_name, 'position': row.position, 'total_salary': float(row.total_salary or 0)}
            for row in salaries_data
        ]

        return {
            'completed': completed,
            'area': area,
            'deadlines': deadlines,
            'salaries': salaries
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in get_employee_report_by_type: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Ошибка получения отчета по сотрудникам: {str(e)}")


@app.post("/api/supervision/cards/{card_id}/history")
async def add_supervision_history(
    card_id: int,
    data: SupervisionHistoryCreate,  # ИСПРАВЛЕНИЕ 06.02.2026: Принимаем body вместо query (#22)
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Добавить запись в историю надзора"""
    try:
        card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

        history = SupervisionProjectHistory(
            supervision_card_id=card_id,
            entry_type=data.entry_type,
            message=data.message,
            created_by=data.created_by or current_user.id
        )
        db.add(history)
        db.commit()
        db.refresh(history)

        return {
            'id': history.id,
            'supervision_card_id': history.supervision_card_id,
            'entry_type': history.entry_type,
            'message': history.message,
            'created_at': history.created_at.isoformat() if history.created_at else None,
            'created_by': history.created_by
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка добавления истории: {str(e)}")



@app.get("/api/files/all")
async def get_all_project_files(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить все файлы проектов для синхронизации"""
    try:
        files = db.query(ProjectFile).all()

        return [{
            'id': f.id,
            'contract_id': f.contract_id,
            'stage': f.stage,
            'file_type': f.file_type,
            'public_link': f.public_link,
            'yandex_path': f.yandex_path,
            'file_name': f.file_name,
            'preview_cache_path': f.preview_cache_path,
            'file_order': f.file_order,
            'variation': f.variation,
            'upload_date': f.upload_date.isoformat() if f.upload_date else None
        } for f in files]

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения файлов: {str(e)}")


@app.get("/api/files/{file_id}")
async def get_file_record(
    file_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить информацию о файле"""
    file_record = db.query(ProjectFile).filter(ProjectFile.id == file_id).first()
    if not file_record:
        raise HTTPException(status_code=404, detail="Файл не найден")
    return {
        'id': file_record.id,
        'contract_id': file_record.contract_id,
        'stage': file_record.stage,
        'file_type': file_record.file_type,
        'public_link': file_record.public_link,
        'yandex_path': file_record.yandex_path,
        'file_name': file_record.file_name,
        'file_order': file_record.file_order,
        'variation': file_record.variation
    }


@app.patch("/api/files/{file_id}/order")
async def update_file_order(
    file_id: int,
    file_order: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить порядок файла в галерее"""
    file_record = db.query(ProjectFile).filter(ProjectFile.id == file_id).first()
    if not file_record:
        raise HTTPException(status_code=404, detail="Файл не найден")

    file_record.file_order = file_order
    db.commit()

    return {"status": "success", "file_id": file_id, "file_order": file_order}


# =========================
# TIMELINE ENDPOINTS (CRM)
# =========================

def calc_contract_term(project_type_code: int, area: float):
    """Расчёт срока договора. 1=Полный, 2=Эскизный, 3=Планировочный"""
    if project_type_code == 1:
        thresholds = [(70,50),(100,60),(130,70),(160,80),(190,90),(220,100),
                      (250,110),(300,120),(350,130),(400,140),(450,150),(500,160)]
    elif project_type_code == 3:
        thresholds = [(70,10),(100,15),(130,20),(160,25),(190,30),(220,35),
                      (250,40),(300,45),(350,50),(400,55),(450,60),(500,65)]
    else:
        thresholds = [(70,30),(100,35),(130,40),(160,45),(190,50),(220,55),
                      (250,60),(300,65),(350,70),(400,75),(450,80),(500,85)]
    for max_area, days in thresholds:
        if area <= max_area:
            return days
    return 0  # индивидуальный расчёт


def calc_area_coefficient(area: float) -> int:
    return max(0, int((area - 1) // 100))


def build_project_timeline_template(project_type: str, area: float, project_subtype: str = None):
    """Генерация полного шаблона подэтапов с формулами.
    project_subtype: 'Полный (с 3д визуализацией)', 'Эскизный (с коллажами)', 'Планировочный'
    """
    K = calc_area_coefficient(area)
    # Определяем pt_code из project_subtype (если задан)
    if project_subtype:
        if 'Полный' in project_subtype:
            pt_code = 1
        elif 'Планировочный' in project_subtype:
            pt_code = 3
        else:
            pt_code = 2
    else:
        pt_code = 1 if project_type == 'Индивидуальный' else 2
    contract_term = calc_contract_term(pt_code, area)

    # Все подэтапы: (stage_code, stage_name, stage_group, substage_group, raw_G, executor, in_scope)
    entries = []
    order = 0

    def add(code, name, group, subgroup, g, executor, in_scope=True):
        nonlocal order
        order += 1
        entries.append({
            'stage_code': code, 'stage_name': name, 'stage_group': group,
            'substage_group': subgroup, 'raw_norm_days': g, 'executor_role': executor,
            'is_in_contract_scope': in_scope, 'sort_order': order
        })

    def add_header(code, name, group, subgroup=''):
        nonlocal order
        order += 1
        entries.append({
            'stage_code': code, 'stage_name': name, 'stage_group': group,
            'substage_group': subgroup, 'raw_norm_days': 0, 'executor_role': 'header',
            'is_in_contract_scope': False, 'sort_order': order
        })

    # --- ДАТА НАЧАЛА ---
    add('START', 'ДАТА НАЧАЛА РАЗРАБОТКИ', 'START', '', 0, 'Менеджер', True)

    # --- ЭТАП 1: ПЛАНИРОВОЧНОЕ РЕШЕНИЕ ---
    add_header('S1_HDR', 'ЭТАП 1: ПЛАНИРОВОЧНОЕ РЕШЕНИЕ', 'STAGE1')

    # Подэтап 1.1 — входит
    add_header('S1_1_HDR', 'Подэтап 1.1', 'STAGE1', 'Подэтап 1.1')
    add('S1_1_01', 'Разработка 3 вар. планировок', 'STAGE1', 'Подэтап 1.1', 4 + K*2, 'Чертежник', True)
    add('S1_1_02', 'Проверка СДП', 'STAGE1', 'Подэтап 1.1', 1 + K*0.5, 'СДП', True)
    add('S1_1_03', 'Правка чертежником', 'STAGE1', 'Подэтап 1.1', 1.5 + K*1, 'Чертежник', True)
    add('S1_1_04', 'Проверка повторная СДП', 'STAGE1', 'Подэтап 1.1', 0.5 + K*0.5, 'СДП', True)
    add('S1_1_05', 'Отправка клиенту', 'STAGE1', 'Подэтап 1.1', 0, 'Клиент', False)
    add('S1_1_06', 'Сбор правок от клиента СДП', 'STAGE1', 'Подэтап 1.1', 1 + K*0.5, 'СДП', False)

    # Подэтап 1.2 — не входит
    add_header('S1_2_HDR', 'Подэтап 1.2 — Фин. план 1 круг', 'STAGE1', 'Подэтап 1.2')
    add('S1_2_01', 'Фин. план. решение (1 круг)', 'STAGE1', 'Подэтап 1.2', 1 + K*1, 'Чертежник', True)
    add('S1_2_02', 'Проверка СДП', 'STAGE1', 'Подэтап 1.2', 1 + K*0.5, 'СДП', False)
    add('S1_2_03', 'Правка чертежником', 'STAGE1', 'Подэтап 1.2', 1 + K*0.5, 'Чертежник', False)
    add('S1_2_04', 'Проверка повторная СДП', 'STAGE1', 'Подэтап 1.2', 1 + K*0.5, 'СДП', False)
    add('S1_2_05', 'Отправка клиенту', 'STAGE1', 'Подэтап 1.2', 0, 'Клиент', False)
    add('S1_2_06', 'Сбор правок от клиента СДП', 'STAGE1', 'Подэтап 1.2', 1 + K*0.5, 'СДП', False)

    # Подэтап 1.3 — не входит
    add_header('S1_3_HDR', 'Подэтап 1.3 — Фин. план 2 круг', 'STAGE1', 'Подэтап 1.3')
    add('S1_3_01', 'Фин. план. решение (2 круг)', 'STAGE1', 'Подэтап 1.3', 1 + K*1, 'Чертежник', False)
    add('S1_3_02', 'Проверка СДП', 'STAGE1', 'Подэтап 1.3', 1 + K*0.5, 'СДП', False)
    add('S1_3_03', 'Правка чертежником', 'STAGE1', 'Подэтап 1.3', 1 + K*0.5, 'Чертежник', False)
    add('S1_3_04', 'Проверка СДП', 'STAGE1', 'Подэтап 1.3', 1 + K*0.5, 'СДП', False)
    add('S1_3_05', 'Согласование планировки. Акт', 'STAGE1', 'Подэтап 1.3', 0, 'Клиент', False)

    # --- ЭТАП 2: КОНЦЕПЦИЯ ДИЗАЙНА ---
    add_header('S2_HDR', 'ЭТАП 2: КОНЦЕПЦИЯ ДИЗАЙНА', 'STAGE2')

    # 2.1 Мудборды
    add_header('S2_1_HDR', 'Подэтап 2.1 — Мудборды', 'STAGE2', 'Подэтап 2.1')
    add('S2_1_01', 'Разработка мудбордов', 'STAGE2', 'Подэтап 2.1', 3 + K*2, 'Дизайнер', True)
    add('S2_1_02', 'Проверка СДП', 'STAGE2', 'Подэтап 2.1', 1 + K*1, 'СДП', True)
    add('S2_1_03', 'Правка дизайнером', 'STAGE2', 'Подэтап 2.1', 2 + K*1, 'Дизайнер', True)
    add('S2_1_04', 'Проверка повторная СДП', 'STAGE2', 'Подэтап 2.1', 1 + K*0.5, 'СДП', True)
    add('S2_1_05', 'Отправка клиенту', 'STAGE2', 'Подэтап 2.1', 0, 'Клиент', False)
    add('S2_1_06', 'Сбор правок СДП', 'STAGE2', 'Подэтап 2.1', 1 + K*0.5, 'СДП', False)
    add('S2_1_07', 'Правка дизайнером', 'STAGE2', 'Подэтап 2.1', 1 + K*1, 'Дизайнер', False)
    add('S2_1_08', 'Проверка СДП', 'STAGE2', 'Подэтап 2.1', 1, 'СДП', False)
    add('S2_1_09', 'Согласование мудборда', 'STAGE2', 'Подэтап 2.1', 0, 'Клиент', False)

    # 2.2 Виз 1 пом.
    add_header('S2_2_HDR', 'Подэтап 2.2 — Виз 1 пом.', 'STAGE2', 'Подэтап 2.2')
    add('S2_2_01', 'Разработка визуализации 1 пом.', 'STAGE2', 'Подэтап 2.2', 3 + K*0.5, 'Дизайнер', True)
    add('S2_2_02', 'Проверка СДП', 'STAGE2', 'Подэтап 2.2', 1, 'СДП', True)
    add('S2_2_03', 'Правка дизайнером', 'STAGE2', 'Подэтап 2.2', 2, 'Дизайнер', True)
    add('S2_2_04', 'Проверка повторная СДП', 'STAGE2', 'Подэтап 2.2', 1, 'СДП', True)
    add('S2_2_05', 'Отправка клиенту', 'STAGE2', 'Подэтап 2.2', 0, 'Клиент', False)
    add('S2_2_06', 'Сбор правок СДП', 'STAGE2', 'Подэтап 2.2', 1, 'СДП', False)

    # 2.3 Виз 1 пом. 1 круг — не входит
    add_header('S2_3_HDR', 'Подэтап 2.3 — Виз 1 пом. 1 круг', 'STAGE2', 'Подэтап 2.3')
    add('S2_3_01', 'Правка визуализации (1 круг)', 'STAGE2', 'Подэтап 2.3', 2 + K*0.5, 'Дизайнер', False)
    add('S2_3_02', 'Проверка СДП', 'STAGE2', 'Подэтап 2.3', 1, 'СДП', False)
    add('S2_3_03', 'Правка дизайнером', 'STAGE2', 'Подэтап 2.3', 1, 'Дизайнер', False)
    add('S2_3_04', 'Проверка повторная СДП', 'STAGE2', 'Подэтап 2.3', 1, 'СДП', False)
    add('S2_3_05', 'Отправка клиенту', 'STAGE2', 'Подэтап 2.3', 0, 'Клиент', False)
    add('S2_3_06', 'Сбор правок СДП', 'STAGE2', 'Подэтап 2.3', 1, 'СДП', False)

    # 2.4 Виз 1 пом. 2 круг — не входит
    add_header('S2_4_HDR', 'Подэтап 2.4 — Виз 1 пом. 2 круг', 'STAGE2', 'Подэтап 2.4')
    add('S2_4_01', 'Правка визуализации (2 круг)', 'STAGE2', 'Подэтап 2.4', 1 + K*1, 'Дизайнер', False)
    add('S2_4_02', 'Проверка СДП', 'STAGE2', 'Подэтап 2.4', 1, 'СДП', False)
    add('S2_4_03', 'Правка дизайнером', 'STAGE2', 'Подэтап 2.4', 1, 'Дизайнер', False)
    add('S2_4_04', 'Проверка СДП', 'STAGE2', 'Подэтап 2.4', 1, 'СДП', False)
    add('S2_4_05', 'Согласование 1 пом.', 'STAGE2', 'Подэтап 2.4', 0, 'Клиент', False)

    # 2.5 Виз остальных — входит
    add_header('S2_5_HDR', 'Подэтап 2.5 — Виз остальных', 'STAGE2', 'Подэтап 2.5')
    add('S2_5_01', 'Разработка визуализаций всех', 'STAGE2', 'Подэтап 2.5', 10 + K*10, 'Дизайнер', True)
    add('S2_5_02', 'Проверка СДП', 'STAGE2', 'Подэтап 2.5', 3 + K*2.5, 'СДП', True)
    add('S2_5_03', 'Правка дизайнером', 'STAGE2', 'Подэтап 2.5', 5 + K*5, 'Дизайнер', True)
    add('S2_5_04', 'Проверка повторная СДП', 'STAGE2', 'Подэтап 2.5', 2 + K*1.5, 'СДП', True)
    add('S2_5_05', 'Отправка клиенту', 'STAGE2', 'Подэтап 2.5', 0, 'Клиент', False)
    add('S2_5_06', 'Сбор правок СДП', 'STAGE2', 'Подэтап 2.5', 2 + K*1.5, 'СДП', False)

    # 2.6 Виз все 1 круг — не входит
    add_header('S2_6_HDR', 'Подэтап 2.6 — Виз все 1 круг', 'STAGE2', 'Подэтап 2.6')
    add('S2_6_01', 'Правка визуализаций (1 круг)', 'STAGE2', 'Подэтап 2.6', 5 + K*5, 'Дизайнер', False)
    add('S2_6_02', 'Проверка СДП', 'STAGE2', 'Подэтап 2.6', 2 + K*1.5, 'СДП', False)
    add('S2_6_03', 'Правка дизайнером', 'STAGE2', 'Подэтап 2.6', 2 + K*1.5, 'Дизайнер', False)
    add('S2_6_04', 'Проверка повторная СДП', 'STAGE2', 'Подэтап 2.6', 2 + K*1.5, 'СДП', False)
    add('S2_6_05', 'Согласование визуализаций', 'STAGE2', 'Подэтап 2.6', 0, 'Клиент', False)

    # 2.7 Виз все 2 круг — не входит
    add_header('S2_7_HDR', 'Подэтап 2.7 — Виз все 2 круг', 'STAGE2', 'Подэтап 2.7')
    add('S2_7_01', 'Правка визуализаций (2 круг)', 'STAGE2', 'Подэтап 2.7', 3 + K*3, 'Дизайнер', False)
    add('S2_7_02', 'Проверка СДП', 'STAGE2', 'Подэтап 2.7', 1 + K*1, 'СДП', False)
    add('S2_7_03', 'Правка дизайнером', 'STAGE2', 'Подэтап 2.7', 1 + K*1, 'Дизайнер', False)
    add('S2_7_04', 'Проверка СДП', 'STAGE2', 'Подэтап 2.7', 1 + K*1, 'СДП', False)
    add('S2_7_05', 'Согласование дизайна. Акт', 'STAGE2', 'Подэтап 2.7', 0, 'Клиент', False)

    # --- ЭТАП 3: РАБОЧАЯ ДОКУМЕНТАЦИЯ ---
    add_header('S3_HDR', 'ЭТАП 3: РАБОЧАЯ ДОКУМЕНТАЦИЯ', 'STAGE3')

    add('S3_01', 'Подготовка файлов, выдача', 'STAGE3', '', 1, 'СДП', True)
    add('S3_02', 'Разработка комплекта РД', 'STAGE3', '', 10 + K*2, 'Чертежник', True)
    add('S3_03', 'Проверка ГАП (1 круг)', 'STAGE3', '', 3 + K*0.5, 'ГАП', True)
    add('S3_04', 'Правка чертежником', 'STAGE3', '', 2 + K*1, 'Чертежник', True)
    add('S3_05', 'Проверка ГАП (2 круг)', 'STAGE3', '', 1 + K*0.5, 'ГАП', True)
    add('S3_06', 'Правка чертежником (при необх.)', 'STAGE3', '', 1, 'Чертежник', True)
    add('S3_07', 'Проверка ГАП (3 круг)', 'STAGE3', '', 1, 'ГАП', True)
    add('S3_08', 'Отправка клиенту', 'STAGE3', '', 0, 'Клиент', False)
    add('S3_09', 'Сбор правок от клиента', 'STAGE3', '', 1 + K*0.5, 'Менеджер', False)
    add('S3_10', 'Внесение правок чертежником', 'STAGE3', '', 1 + K*1, 'Чертежник', False)
    add('S3_11', 'Проверка ГАП (4 круг)', 'STAGE3', '', 1 + K*0.5, 'ГАП', False)
    add('S3_12', 'Принятие проекта. Акт финальный', 'STAGE3', '', 0, 'Клиент', False)

    # --- Фильтрация по подтипу проекта ---
    if project_subtype and 'Планировочный' in project_subtype:
        # Только START + STAGE1
        entries = [e for e in entries if e['stage_group'] in ('START', 'STAGE1')]
    elif project_subtype and 'Эскизный' in project_subtype:
        # START + STAGE1 + мудборды (Подэтап 2.1) + STAGE3
        entries = [e for e in entries if e['stage_group'] in ('START', 'STAGE1', 'STAGE3')
                   or (e['stage_group'] == 'STAGE2' and e['substage_group'] == 'Подэтап 2.1')
                   or e['stage_code'] == 'S2_HDR']
    # Полный / None — все этапы

    # Перенумерация sort_order после фильтрации
    for i, e in enumerate(entries, 1):
        e['sort_order'] = i

    # Пропорциональный расчёт норм
    in_scope = [e for e in entries if e['is_in_contract_scope'] and e['executor_role'] != 'header' and e['raw_norm_days'] > 0]
    total_raw = sum(e['raw_norm_days'] for e in in_scope)

    if total_raw > 0 and contract_term > 0:
        cumulative = 0
        prev_rounded = 0
        for e in in_scope:
            cumulative += e['raw_norm_days']
            e['cumulative_days'] = cumulative
            current_rounded = round(cumulative / total_raw * contract_term)
            e['norm_days'] = max(1, current_rounded - prev_rounded) if prev_rounded > 0 else max(1, current_rounded)
            prev_rounded = current_rounded

    # Не в сроке: norm = max(1, round(raw))
    for e in entries:
        if e['executor_role'] == 'header':
            e['norm_days'] = 0
            continue
        if not e['is_in_contract_scope'] and e['raw_norm_days'] > 0:
            e['norm_days'] = max(1, round(e['raw_norm_days']))

    return entries, contract_term, K


def calc_template_contract_term(template_subtype: str, area: float, floors: int = 1) -> int:
    """Расчёт срока для шаблонных проектов (рабочие дни)"""
    sub = template_subtype.lower()
    if 'ванной' in sub:
        if 'визуализац' in sub:
            return 20
        return 10

    # Стандарт / Стандарт с визуализацией
    if area <= 90:
        base_days = 20
    else:
        extra = int((area - 90 - 1) // 50) + 1
        base_days = 20 + extra * 10

    # Доп. этажи
    if floors > 1:
        for _ in range(1, floors):
            if area <= 90:
                base_days += 10
            else:
                extra = int((area - 90 - 1) // 50) + 1
                base_days += 10 + extra * 10

    # Визуализация
    if 'визуализац' in sub:
        if area <= 90:
            base_days += 25
        else:
            extra = int((area - 90 - 1) // 50) + 1
            base_days += 25 + extra * 15

    return int(base_days)


def build_template_project_timeline(template_subtype: str, area: float, floors: int = 1):
    """Генерация шаблона таблицы сроков для шаблонных проектов.
    template_subtype: 'Стандарт', 'Стандарт с визуализацией',
                      'Проект ванной комнаты', 'Проект ванной комнаты с визуализацией'
    """
    contract_term = calc_template_contract_term(template_subtype, area, floors)
    entries = []
    order = 0

    def add(code, name, group, subgroup, g, executor, in_scope=True):
        nonlocal order
        order += 1
        entries.append({
            'stage_code': code, 'stage_name': name, 'stage_group': group,
            'substage_group': subgroup, 'raw_norm_days': g, 'executor_role': executor,
            'is_in_contract_scope': in_scope, 'sort_order': order
        })

    def add_header(code, name, group, subgroup=''):
        nonlocal order
        order += 1
        entries.append({
            'stage_code': code, 'stage_name': name, 'stage_group': group,
            'substage_group': subgroup, 'raw_norm_days': 0, 'executor_role': 'header',
            'is_in_contract_scope': False, 'sort_order': order
        })

    # --- ДАТА НАЧАЛА ---
    add('START', 'ДАТА НАЧАЛА РАЗРАБОТКИ', 'START', '', 0, 'Менеджер', True)

    # --- СТАДИЯ 1: ПЛАНИРОВОЧНЫЕ РЕШЕНИЯ ---
    add_header('T1_HDR', 'СТАДИЯ 1: ПЛАНИРОВОЧНЫЕ РЕШЕНИЯ', 'STAGE1')

    # Подэтап 1.1
    add_header('T1_1_HDR', 'Подэтап 1.1', 'STAGE1', 'Подэтап 1.1')
    add('T1_1_01', 'Разработка 3 вар. план. решений', 'STAGE1', 'Подэтап 1.1', 3, 'Чертежник', True)
    add('T1_1_02', 'Проверка менеджером', 'STAGE1', 'Подэтап 1.1', 1, 'Менеджер', True)
    add('T1_1_03', 'Правка чертежником', 'STAGE1', 'Подэтап 1.1', 1, 'Чертежник', True)
    add('T1_1_04', 'Проверка повторная менеджером', 'STAGE1', 'Подэтап 1.1', 0.5, 'Менеджер', True)
    add('T1_1_05', 'Отправка клиенту / Согласование', 'STAGE1', 'Подэтап 1.1', 0, 'Клиент', False)
    add('T1_1_06', 'Сбор правок от клиента', 'STAGE1', 'Подэтап 1.1', 1, 'Менеджер', False)

    # Подэтап 1.2
    add_header('T1_2_HDR', 'Подэтап 1.2 — Финальное план. решение', 'STAGE1', 'Подэтап 1.2')
    add('T1_2_01', 'Финальное план. решение (1 круг)', 'STAGE1', 'Подэтап 1.2', 1, 'Чертежник', True)
    add('T1_2_02', 'Проверка менеджером', 'STAGE1', 'Подэтап 1.2', 1, 'Менеджер', False)
    add('T1_2_03', 'Правка чертежником', 'STAGE1', 'Подэтап 1.2', 1, 'Чертежник', False)
    add('T1_2_04', 'Проверка повторная менеджером', 'STAGE1', 'Подэтап 1.2', 0.5, 'Менеджер', False)
    add('T1_2_05', 'Отправка клиенту / Согласование', 'STAGE1', 'Подэтап 1.2', 0, 'Клиент', False)

    # --- СТАДИЯ 2: РАБОЧИЕ ЧЕРТЕЖИ ---
    add_header('T2_HDR', 'СТАДИЯ 2: РАБОЧИЕ ЧЕРТЕЖИ', 'STAGE2')

    add('T2_01', 'Подготовка файлов, выдача чертежнику', 'STAGE2', '', 1, 'Менеджер', True)
    add('T2_02', 'Разработка комплекта РД', 'STAGE2', '', 5, 'Чертежник', True)
    add('T2_03', 'Проверка ГАП (1 круг)', 'STAGE2', '', 2, 'ГАП', True)
    add('T2_04', 'Правка чертежником', 'STAGE2', '', 1, 'Чертежник', True)
    add('T2_05', 'Проверка ГАП (2 круг)', 'STAGE2', '', 1, 'ГАП', True)
    add('T2_06', 'Правка чертежником (при необх.)', 'STAGE2', '', 1, 'Чертежник', False)
    add('T2_07', 'Проверка ГАП (3 круг)', 'STAGE2', '', 1, 'ГАП', False)
    add('T2_08', 'Отправка клиенту / Согласование', 'STAGE2', '', 0, 'Клиент', False)
    add('T2_09', 'Сбор правок от клиента', 'STAGE2', '', 1, 'Менеджер', False)
    add('T2_10', 'Внесение правок чертежником', 'STAGE2', '', 1, 'Чертежник', False)
    add('T2_11', 'Проверка ГАП (4 круг)', 'STAGE2', '', 1, 'ГАП', False)
    add('T2_12', 'Принятие проекта. Закрытие.', 'STAGE2', '', 0, 'Клиент', False)

    # --- СТАДИЯ 3: 3Д ВИЗУАЛИЗАЦИЯ (только для подтипов с визуализацией) ---
    has_viz = 'визуализац' in template_subtype.lower()
    if has_viz:
        add_header('T3_HDR', 'СТАДИЯ 3: 3Д ВИЗУАЛИЗАЦИЯ', 'STAGE3')

        add('T3_01', 'Разработка визуализаций всех пом.', 'STAGE3', '', 10, 'Дизайнер', True)
        add('T3_02', 'Проверка менеджером', 'STAGE3', '', 2, 'Менеджер', True)
        add('T3_03', 'Правка дизайнером', 'STAGE3', '', 3, 'Дизайнер', True)
        add('T3_04', 'Проверка повторная менеджером', 'STAGE3', '', 1, 'Менеджер', True)
        add('T3_05', 'Отправка клиенту / Согласование', 'STAGE3', '', 0, 'Клиент', False)
        add('T3_06', 'Принятие проекта. Закрытие.', 'STAGE3', '', 0, 'Клиент', False)

    # Перенумерация sort_order
    for i, e in enumerate(entries, 1):
        e['sort_order'] = i

    # Пропорциональный расчёт норм
    in_scope = [e for e in entries if e['is_in_contract_scope'] and e['executor_role'] != 'header' and e['raw_norm_days'] > 0]
    total_raw = sum(e['raw_norm_days'] for e in in_scope)

    if total_raw > 0 and contract_term > 0:
        cumulative = 0
        prev_rounded = 0
        for e in in_scope:
            cumulative += e['raw_norm_days']
            e['cumulative_days'] = cumulative
            current_rounded = round(cumulative / total_raw * contract_term)
            e['norm_days'] = max(1, current_rounded - prev_rounded) if prev_rounded > 0 else max(1, current_rounded)
            prev_rounded = current_rounded

    # Не в сроке: norm = max(1, round(raw))
    for e in entries:
        if e['executor_role'] == 'header':
            e['norm_days'] = 0
            continue
        if not e['is_in_contract_scope'] and e['raw_norm_days'] > 0:
            e['norm_days'] = max(1, round(e['raw_norm_days']))

    return entries, contract_term, 0


@app.get("/api/timeline")
async def get_all_timelines(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить все записи таймлайна проектов"""
    entries = db.query(ProjectTimelineEntry).order_by(
        ProjectTimelineEntry.contract_id,
        ProjectTimelineEntry.sort_order
    ).all()
    return [
        {c.name: getattr(e, c.name) for c in e.__table__.columns}
        for e in entries
    ]


@app.get("/api/timeline/{contract_id}")
async def get_project_timeline(
    contract_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить таблицу сроков проекта"""
    entries = db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id
    ).order_by(ProjectTimelineEntry.sort_order).all()
    return [
        {c.name: getattr(e, c.name) for c in e.__table__.columns}
        for e in entries
    ]


@app.post("/api/timeline/{contract_id}/init")
async def init_project_timeline(
    contract_id: int,
    request: TimelineInitRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Инициализация таблицы сроков проекта из шаблона"""
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")

    # Проверяем, есть ли уже записи
    existing = db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id
    ).count()
    if existing > 0:
        return {"status": "already_initialized", "count": existing}

    if request.project_type == 'Шаблонный':
        template_subtype = request.project_subtype or 'Стандарт'
        floors = getattr(request, 'floors', 1) or 1
        entries, contract_term, K = build_template_project_timeline(
            template_subtype, request.area, floors
        )
    else:
        entries, contract_term, K = build_project_timeline_template(
            request.project_type, request.area, request.project_subtype
        )

    for e in entries:
        record = ProjectTimelineEntry(
            contract_id=contract_id,
            stage_code=e['stage_code'],
            stage_name=e['stage_name'],
            stage_group=e['stage_group'],
            substage_group=e.get('substage_group', ''),
            executor_role=e['executor_role'],
            is_in_contract_scope=e['is_in_contract_scope'],
            sort_order=e['sort_order'],
            raw_norm_days=e.get('raw_norm_days', 0),
            cumulative_days=e.get('cumulative_days', 0),
            norm_days=e.get('norm_days', 0)
        )
        db.add(record)

    db.commit()
    return {
        "status": "initialized",
        "count": len(entries),
        "contract_term": contract_term,
        "area_coefficient": K
    }


@app.post("/api/timeline/{contract_id}/reinit")
async def reinit_project_timeline(
    contract_id: int,
    request: TimelineInitRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Пересоздать таблицу сроков проекта (удалить старые записи и создать заново)"""
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")

    # Удаляем старые записи
    db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id
    ).delete()
    db.flush()

    if request.project_type == 'Шаблонный':
        template_subtype = request.project_subtype or 'Стандарт'
        floors = getattr(request, 'floors', 1) or 1
        entries, contract_term, K = build_template_project_timeline(
            template_subtype, request.area, floors
        )
    else:
        entries, contract_term, K = build_project_timeline_template(
            request.project_type, request.area, request.project_subtype
        )

    for e in entries:
        record = ProjectTimelineEntry(
            contract_id=contract_id,
            stage_code=e['stage_code'],
            stage_name=e['stage_name'],
            stage_group=e['stage_group'],
            substage_group=e.get('substage_group', ''),
            executor_role=e['executor_role'],
            is_in_contract_scope=e['is_in_contract_scope'],
            sort_order=e['sort_order'],
            raw_norm_days=e.get('raw_norm_days', 0),
            cumulative_days=e.get('cumulative_days', 0),
            norm_days=e.get('norm_days', 0)
        )
        db.add(record)

    db.commit()
    return {
        "status": "reinitialized",
        "count": len(entries),
        "contract_term": contract_term,
        "area_coefficient": K
    }


@app.put("/api/timeline/{contract_id}/entry/{stage_code}")
async def update_timeline_entry(
    contract_id: int,
    stage_code: str,
    update: TimelineEntryUpdate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить запись таблицы сроков"""
    entry = db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id,
        ProjectTimelineEntry.stage_code == stage_code
    ).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    update_data = update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(entry, key, value)
    entry.updated_at = datetime.utcnow()

    db.commit()
    return {c.name: getattr(entry, c.name) for c in entry.__table__.columns}


@app.get("/api/timeline/{contract_id}/summary")
async def get_timeline_summary(
    contract_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Сводка по таблице сроков"""
    entries = db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id,
        ProjectTimelineEntry.executor_role != 'header'
    ).all()

    total_norm = sum(e.norm_days for e in entries if e.is_in_contract_scope)
    total_actual = sum(e.actual_days for e in entries if e.is_in_contract_scope and e.actual_days > 0)
    overdue_count = sum(1 for e in entries if e.actual_days > e.norm_days > 0)

    return {
        "total_norm_days": total_norm,
        "total_actual_days": total_actual,
        "overdue_count": overdue_count,
        "entries_count": len(entries)
    }


# =========================
# WORKFLOW ENDPOINTS (CRM)
# =========================

@app.get("/api/crm/cards/{card_id}/workflow/state")
async def get_workflow_state(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить текущее состояние рабочего процесса карточки"""
    states = db.query(StageWorkflowState).filter(
        StageWorkflowState.crm_card_id == card_id
    ).all()
    return [
        {c.name: getattr(s, c.name) for c in s.__table__.columns}
        for s in states
    ]


@app.post("/api/crm/cards/{card_id}/workflow/submit")
async def workflow_submit_work(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Сдача работы исполнителем — записывает дату в timeline"""
    card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка не найдена")

    contract_id = card.contract_id
    stage_name = card.column_name

    # Находим текущий подэтап (первый без actual_date, is_in_contract_scope, не header)
    stage_group = _resolve_stage_group(stage_name)
    if not stage_group:
        return {"status": "no_stage_group"}

    entries = db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id,
        ProjectTimelineEntry.stage_group == stage_group,
        ProjectTimelineEntry.executor_role != 'header',
        ProjectTimelineEntry.actual_date.is_(None)
    ).order_by(ProjectTimelineEntry.sort_order).all()

    # Также check entries with empty string
    if not entries:
        entries = db.query(ProjectTimelineEntry).filter(
            ProjectTimelineEntry.contract_id == contract_id,
            ProjectTimelineEntry.stage_group == stage_group,
            ProjectTimelineEntry.executor_role != 'header',
            ProjectTimelineEntry.actual_date == ''
        ).order_by(ProjectTimelineEntry.sort_order).all()

    if entries:
        entry = entries[0]
        entry.actual_date = datetime.utcnow().strftime('%Y-%m-%d')
        entry.updated_at = datetime.utcnow()

    # Обновляем workflow state
    wf = db.query(StageWorkflowState).filter(
        StageWorkflowState.crm_card_id == card_id,
        StageWorkflowState.stage_name == stage_name
    ).first()
    if not wf:
        wf = StageWorkflowState(
            crm_card_id=card_id,
            stage_name=stage_name,
            current_substep_code=entries[0].stage_code if entries else None,
            status='in_progress'
        )
        db.add(wf)
    wf.current_substep_code = entries[0].stage_code if entries else wf.current_substep_code
    wf.updated_at = datetime.utcnow()

    db.commit()

    # Хук: уведомление в чат о сдаче работы
    asyncio.create_task(_trigger_messenger_notification(
        db, card_id, 'stage_complete', stage_name=stage_name
    ))

    return {"status": "submitted", "substep": entries[0].stage_code if entries else None}


@app.post("/api/crm/cards/{card_id}/workflow/accept")
async def workflow_accept_work(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Приемка работы — записывает дату проверки в timeline"""
    card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка не найдена")

    contract_id = card.contract_id
    stage_name = card.column_name
    stage_group = _resolve_stage_group(stage_name)

    if stage_group:
        entries = db.query(ProjectTimelineEntry).filter(
            ProjectTimelineEntry.contract_id == contract_id,
            ProjectTimelineEntry.stage_group == stage_group,
            ProjectTimelineEntry.executor_role != 'header',
            ProjectTimelineEntry.actual_date.is_(None) | (ProjectTimelineEntry.actual_date == '')
        ).order_by(ProjectTimelineEntry.sort_order).all()

        if entries:
            entry = entries[0]
            entry.actual_date = datetime.utcnow().strftime('%Y-%m-%d')
            entry.updated_at = datetime.utcnow()

    # Обновляем workflow state
    wf = db.query(StageWorkflowState).filter(
        StageWorkflowState.crm_card_id == card_id,
        StageWorkflowState.stage_name == stage_name
    ).first()
    if wf:
        wf.status = 'in_progress'
        wf.updated_at = datetime.utcnow()

    db.commit()
    return {"status": "accepted"}


@app.post("/api/crm/cards/{card_id}/workflow/reject")
async def workflow_reject_work(
    card_id: int,
    request: Request,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Отправить на исправление — обновляет workflow state и сбрасывает completed.
    Опционально принимает revision_file_path (путь к папке правок на ЯД)."""
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка не найдена")

    stage_name = card.column_name
    revision_file_path = body.get('revision_file_path', '')

    # Сбрасываем completed у исполнителя текущей стадии
    # чтобы он увидел кнопку "Сдать работу" снова
    executors = db.query(StageExecutor).filter(
        StageExecutor.crm_card_id == card_id,
        StageExecutor.completed == True
    ).all()
    for ex in executors:
        # Сбрасываем только если stage_name совпадает с текущей стадией
        if ex.stage_name and (
            ex.stage_name == stage_name or
            stage_name.lower() in ex.stage_name.lower() or
            ex.stage_name.lower() in stage_name.lower()
        ):
            ex.completed = False
            ex.completed_date = None

    wf = db.query(StageWorkflowState).filter(
        StageWorkflowState.crm_card_id == card_id,
        StageWorkflowState.stage_name == stage_name
    ).first()
    if not wf:
        wf = StageWorkflowState(
            crm_card_id=card_id,
            stage_name=stage_name,
            status='revision',
            revision_count=1,
            revision_file_path=revision_file_path or None
        )
        db.add(wf)
    else:
        wf.status = 'revision'
        wf.revision_count = (wf.revision_count or 0) + 1
        if revision_file_path:
            wf.revision_file_path = revision_file_path
        wf.updated_at = datetime.utcnow()

    db.commit()
    return {"status": "rejected", "revision_count": wf.revision_count, "revision_file_path": wf.revision_file_path}


@app.post("/api/crm/cards/{card_id}/workflow/client-send")
async def workflow_client_send(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Отправить на согласование клиенту — приостанавливает дедлайн"""
    card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка не найдена")

    stage_name = card.column_name
    contract_id = card.contract_id
    stage_group = _resolve_stage_group(stage_name)

    # Записываем дату в timeline (Отправка клиенту)
    if stage_group:
        entries = db.query(ProjectTimelineEntry).filter(
            ProjectTimelineEntry.contract_id == contract_id,
            ProjectTimelineEntry.stage_group == stage_group,
            ProjectTimelineEntry.executor_role == 'Клиент',
            ProjectTimelineEntry.actual_date.is_(None) | (ProjectTimelineEntry.actual_date == '')
        ).order_by(ProjectTimelineEntry.sort_order).first()
        if entries:
            entries.actual_date = datetime.utcnow().strftime('%Y-%m-%d')

    wf = db.query(StageWorkflowState).filter(
        StageWorkflowState.crm_card_id == card_id,
        StageWorkflowState.stage_name == stage_name
    ).first()
    if not wf:
        wf = StageWorkflowState(
            crm_card_id=card_id,
            stage_name=stage_name,
            status='client_approval',
            client_approval_started_at=datetime.utcnow(),
            client_approval_deadline_paused=True
        )
        db.add(wf)
    else:
        wf.status = 'client_approval'
        wf.client_approval_started_at = datetime.utcnow()
        wf.client_approval_deadline_paused = True
        wf.updated_at = datetime.utcnow()

    db.commit()

    # Хук: уведомление в чат об отправке клиенту
    asyncio.create_task(_trigger_messenger_notification(
        db, card_id, 'stage_complete', stage_name=f"{stage_name} (отправлено клиенту)"
    ))

    return {"status": "sent_to_client"}


@app.post("/api/crm/cards/{card_id}/workflow/client-ok")
async def workflow_client_approved(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Клиент согласовал — возобновляет дедлайн"""
    card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка не найдена")

    stage_name = card.column_name

    wf = db.query(StageWorkflowState).filter(
        StageWorkflowState.crm_card_id == card_id,
        StageWorkflowState.stage_name == stage_name
    ).first()
    if wf:
        wf.status = 'in_progress'
        wf.client_approval_deadline_paused = False
        wf.updated_at = datetime.utcnow()

    db.commit()

    # Хук: уведомление в чат о согласовании клиентом
    asyncio.create_task(_trigger_messenger_notification(
        db, card_id, 'stage_complete', stage_name=f"{stage_name} (клиент согласовал)"
    ))

    return {"status": "client_approved"}


def _resolve_stage_group(column_name: str) -> str:
    """Определить stage_group по имени колонки канбана"""
    col = column_name.lower()
    if 'стадия 1' in col:
        return 'STAGE1'
    elif 'стадия 2' in col and ('концепция' in col or 'дизайн' in col):
        return 'STAGE2'
    elif 'стадия 2' in col:
        return 'STAGE2'
    elif 'стадия 3' in col:
        return 'STAGE3'
    return ''


# =========================
# SUPERVISION TIMELINE ENDPOINTS
# =========================

SUPERVISION_STAGES = [
    ('STAGE_1_CERAMIC', 'Стадия 1: Закупка керамогранита'),
    ('STAGE_2_PLUMBING', 'Стадия 2: Закупка сантехники'),
    ('STAGE_3_EQUIPMENT', 'Стадия 3: Закупка оборудования'),
    ('STAGE_4_DOORS', 'Стадия 4: Закупка дверей и окон'),
    ('STAGE_5_WALL', 'Стадия 5: Закупка настенных материалов'),
    ('STAGE_6_FLOOR', 'Стадия 6: Закупка напольных материалов'),
    ('STAGE_7_STUCCO', 'Стадия 7: Лепной декор'),
    ('STAGE_8_LIGHTING', 'Стадия 8: Освещение'),
    ('STAGE_9_APPLIANCES', 'Стадия 9: Бытовая техника'),
    ('STAGE_10_CUSTOM_FURNITURE', 'Стадия 10: Закупка заказной мебели'),
    ('STAGE_11_FACTORY_FURNITURE', 'Стадия 11: Закупка фабричной мебели'),
    ('STAGE_12_DECOR', 'Стадия 12: Закупка декора'),
]


@app.get("/api/supervision-timeline")
async def get_all_supervision_timelines(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить все записи таймлайна надзора"""
    entries = db.query(SupervisionTimelineEntry).order_by(
        SupervisionTimelineEntry.supervision_card_id,
        SupervisionTimelineEntry.sort_order
    ).all()
    return {
        "entries": [
            {c.name: getattr(e, c.name) for c in e.__table__.columns}
            for e in entries
        ],
        "total": len(entries)
    }


@app.get("/api/supervision-timeline/{card_id}")
async def get_supervision_timeline(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить таблицу сроков надзора"""
    entries = db.query(SupervisionTimelineEntry).filter(
        SupervisionTimelineEntry.supervision_card_id == card_id
    ).order_by(SupervisionTimelineEntry.sort_order).all()
    return [
        {c.name: getattr(e, c.name) for c in e.__table__.columns}
        for e in entries
    ]


@app.post("/api/supervision-timeline/{card_id}/init")
async def init_supervision_timeline(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Инициализация таблицы сроков надзора (12 стадий)"""
    sv_card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
    if not sv_card:
        raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

    existing = db.query(SupervisionTimelineEntry).filter(
        SupervisionTimelineEntry.supervision_card_id == card_id
    ).count()
    if existing > 0:
        return {"status": "already_initialized", "count": existing}

    for i, (code, name) in enumerate(SUPERVISION_STAGES):
        record = SupervisionTimelineEntry(
            supervision_card_id=card_id,
            stage_code=code,
            stage_name=name,
            sort_order=i + 1,
            status='Не начато'
        )
        db.add(record)

    db.commit()
    return {"status": "initialized", "count": len(SUPERVISION_STAGES)}


@app.put("/api/supervision-timeline/{card_id}/entry/{stage_code}")
async def update_supervision_timeline_entry(
    card_id: int,
    stage_code: str,
    update: SupervisionTimelineUpdate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить запись таблицы сроков надзора"""
    entry = db.query(SupervisionTimelineEntry).filter(
        SupervisionTimelineEntry.supervision_card_id == card_id,
        SupervisionTimelineEntry.stage_code == stage_code
    ).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    update_data = update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(entry, key, value)

    # Автоматический расчёт экономии
    if entry.budget_planned is not None and entry.budget_actual is not None:
        entry.budget_savings = entry.budget_planned - entry.budget_actual

    entry.updated_at = datetime.utcnow()
    db.commit()
    return {c.name: getattr(entry, c.name) for c in entry.__table__.columns}


@app.get("/api/supervision-timeline/{card_id}/summary")
async def get_supervision_timeline_summary(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Сводка по таблице сроков надзора"""
    entries = db.query(SupervisionTimelineEntry).filter(
        SupervisionTimelineEntry.supervision_card_id == card_id
    ).all()

    total_budget_planned = sum(e.budget_planned or 0 for e in entries)
    total_budget_actual = sum(e.budget_actual or 0 for e in entries)
    total_savings = sum(e.budget_savings or 0 for e in entries)
    total_defects = sum(e.defects_found or 0 for e in entries)
    total_resolved = sum(e.defects_resolved or 0 for e in entries)
    total_visits = sum(e.site_visits or 0 for e in entries)

    status_counts = {}
    for e in entries:
        status_counts[e.status] = status_counts.get(e.status, 0) + 1

    return {
        "total_budget_planned": total_budget_planned,
        "total_budget_actual": total_budget_actual,
        "total_savings": total_savings,
        "total_defects_found": total_defects,
        "total_defects_resolved": total_resolved,
        "total_site_visits": total_visits,
        "status_counts": status_counts,
        "entries_count": len(entries)
    }


@app.get("/api/timeline/{contract_id}/export/excel")
async def export_timeline_excel(
    contract_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Экспорт таблицы сроков CRM в Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")

    entries = db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id
    ).order_by(ProjectTimelineEntry.sort_order).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица сроков"

    # Шапка
    ws.append(["Адрес:", contract.address or ""])
    ws.append(["Тип проекта:", contract.project_type or ""])
    ws.append(["Площадь:", f"{contract.area or 0} м²"])
    ws.append([])

    # Заголовки таблицы
    headers = ["Действия по этапам", "Дата", "Кол-во дней", "Норма дней", "Статус", "Исполнитель"]
    ws.append(headers)

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Данные
    stage_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    stage_font = Font(bold=True, color="FFFFFF")
    substage_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    substage_font = Font(bold=True)
    overdue_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

    for entry in entries:
        row_data = [
            entry.stage_name,
            entry.actual_date or "",
            entry.actual_days or 0,
            entry.norm_days or 0,
            entry.status or "",
            entry.executor_role or ""
        ]
        ws.append(row_data)
        row_idx = ws.max_row

        # Стили для заголовков этапов
        if entry.executor_role == 'header':
            for col_idx in range(1, 7):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = stage_fill
                cell.font = stage_font
                cell.border = thin_border
        elif entry.executor_role == 'subheader':
            for col_idx in range(1, 7):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = substage_fill
                cell.font = substage_font
                cell.border = thin_border
        else:
            for col_idx in range(1, 7):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
            # Подсветка просроченных
            if entry.actual_days and entry.norm_days and entry.actual_days > entry.norm_days:
                for col_idx in range(1, 7):
                    ws.cell(row=row_idx, column=col_idx).fill = overdue_fill

    # Ширина колонок
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"timeline_{contract_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/timeline/{contract_id}/export/pdf")
async def export_timeline_pdf(
    contract_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Экспорт таблицы сроков CRM в PDF (для клиента)"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")

    entries = db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id
    ).order_by(ProjectTimelineEntry.sort_order).all()

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)

    # Попытка зарегистрировать шрифт с кириллицей
    font_name = "Helvetica"
    for font_path in ["/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                      "/usr/share/fonts/TTF/DejaVuSans.ttf"]:
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
            font_name = "DejaVuSans"
            break

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title_RU', parent=styles['Title'],
                                  fontName=font_name, fontSize=14)
    cell_style = ParagraphStyle('Cell_RU', fontName=font_name, fontSize=8,
                                 leading=10)
    header_style = ParagraphStyle('Header_RU', fontName=font_name, fontSize=9,
                                   leading=11, textColor=colors.white)

    elements = []
    elements.append(Paragraph(f"Таблица сроков проекта", title_style))
    elements.append(Spacer(1, 5*mm))

    # Информация о проекте
    info_text = f"Адрес: {contract.address or '-'}  |  Тип: {contract.project_type or '-'}  |  Площадь: {contract.area or 0} м²"
    elements.append(Paragraph(info_text, ParagraphStyle('Info', fontName=font_name, fontSize=10)))
    elements.append(Spacer(1, 5*mm))

    # Таблица
    table_data = [[
        Paragraph("Этап", header_style),
        Paragraph("Дата", header_style),
        Paragraph("Дней", header_style),
        Paragraph("Норма", header_style),
        Paragraph("Статус", header_style),
        Paragraph("Исполнитель", header_style)
    ]]

    for entry in entries:
        table_data.append([
            Paragraph(entry.stage_name or "", cell_style),
            Paragraph(entry.actual_date or "", cell_style),
            Paragraph(str(entry.actual_days or ""), cell_style),
            Paragraph(str(entry.norm_days or ""), cell_style),
            Paragraph(entry.status or "", cell_style),
            Paragraph(entry.executor_role if entry.executor_role not in ('header', 'subheader') else "", cell_style)
        ])

    col_widths = [180, 60, 50, 50, 50, 80]
    table = Table(table_data, colWidths=col_widths)

    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8F8')]),
    ]

    # Подсветка строк-заголовков
    for i, entry in enumerate(entries, 1):
        if entry.executor_role == 'header':
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#2F5496')))
            style_cmds.append(('TEXTCOLOR', (0, i), (-1, i), colors.white))
        elif entry.executor_role == 'subheader':
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#D6E4F0')))
        elif entry.actual_days and entry.norm_days and entry.actual_days > entry.norm_days:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F2DCDB')))

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    doc.build(elements)
    output.seek(0)

    filename = f"timeline_{contract_id}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/supervision-timeline/{card_id}/export/excel")
async def export_supervision_timeline_excel(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Экспорт таблицы сроков надзора в Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

    contract = db.query(Contract).filter(Contract.id == card.contract_id).first()
    entries = db.query(SupervisionTimelineEntry).filter(
        SupervisionTimelineEntry.supervision_card_id == card_id
    ).order_by(SupervisionTimelineEntry.sort_order).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица сроков надзора"

    # Шапка
    ws.append(["Адрес:", contract.address if contract else ""])
    ws.append([])

    # Заголовки
    headers = ["Стадия", "План. дата", "Факт. дата", "Дней", "Бюджет план",
               "Бюджет факт", "Экономия", "Поставщик", "Статус", "Примечания"]
    ws.append(headers)

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Цвета статусов
    status_colors = {
        'В работе': 'FFF8E1',
        'Закуплено': 'E3F2FD',
        'Доставлено': 'E8F5E9',
        'Просрочено': 'FFEBEE',
    }

    for entry in entries:
        row_data = [
            entry.stage_name,
            entry.plan_date or "",
            entry.actual_date or "",
            entry.actual_days or 0,
            entry.budget_planned or 0,
            entry.budget_actual or 0,
            entry.budget_savings or 0,
            entry.supplier or "",
            entry.status or "",
            entry.notes or ""
        ]
        ws.append(row_data)
        row_idx = ws.max_row

        for col_idx in range(1, 11):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

        # Цвет по статусу
        color_hex = status_colors.get(entry.status)
        if color_hex:
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            for col_idx in range(1, 11):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Сводка внизу
    ws.append([])
    total_planned = sum(e.budget_planned or 0 for e in entries)
    total_actual = sum(e.budget_actual or 0 for e in entries)
    total_savings = sum(e.budget_savings or 0 for e in entries)
    ws.append(["ИТОГО:", "", "", "", total_planned, total_actual, total_savings])
    total_row = ws.max_row
    for col_idx in range(1, 11):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.border = thin_border

    # Ширина колонок
    widths = [30, 14, 14, 10, 14, 14, 14, 20, 14, 25]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"supervision_timeline_{card_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/supervision-timeline/{card_id}/export/pdf")
async def export_supervision_timeline_pdf(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Экспорт таблицы сроков надзора в PDF (без бюджетов)"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

    contract = db.query(Contract).filter(Contract.id == card.contract_id).first()
    entries = db.query(SupervisionTimelineEntry).filter(
        SupervisionTimelineEntry.supervision_card_id == card_id
    ).order_by(SupervisionTimelineEntry.sort_order).all()

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)

    font_name = "Helvetica"
    for font_path in ["/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                      "/usr/share/fonts/TTF/DejaVuSans.ttf"]:
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
            font_name = "DejaVuSans"
            break

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title_RU', parent=styles['Title'],
                                  fontName=font_name, fontSize=14)
    cell_style = ParagraphStyle('Cell_RU', fontName=font_name, fontSize=8,
                                 leading=10)
    header_style = ParagraphStyle('Header_RU', fontName=font_name, fontSize=9,
                                   leading=11, textColor=colors.white)

    elements = []
    elements.append(Paragraph("Таблица сроков надзора", title_style))
    elements.append(Spacer(1, 5*mm))

    addr = contract.address if contract else "-"
    elements.append(Paragraph(f"Адрес: {addr}", ParagraphStyle('Info', fontName=font_name, fontSize=10)))
    elements.append(Spacer(1, 5*mm))

    # PDF без бюджетов — только: Стадия, План, Факт, Дней, Поставщик, Статус, Примечания
    table_data = [[
        Paragraph("Стадия", header_style),
        Paragraph("План. дата", header_style),
        Paragraph("Факт. дата", header_style),
        Paragraph("Дней", header_style),
        Paragraph("Поставщик", header_style),
        Paragraph("Статус", header_style),
        Paragraph("Примечания", header_style)
    ]]

    for entry in entries:
        table_data.append([
            Paragraph(entry.stage_name or "", cell_style),
            Paragraph(entry.plan_date or "", cell_style),
            Paragraph(entry.actual_date or "", cell_style),
            Paragraph(str(entry.actual_days or ""), cell_style),
            Paragraph(entry.supplier or "", cell_style),
            Paragraph(entry.status or "", cell_style),
            Paragraph(entry.notes or "", cell_style)
        ])

    col_widths = [140, 65, 65, 40, 100, 70, 120]
    table = Table(table_data, colWidths=col_widths)

    status_pdf_colors = {
        'В работе': '#FFF8E1',
        'Закуплено': '#E3F2FD',
        'Доставлено': '#E8F5E9',
        'Просрочено': '#FFEBEE',
    }

    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]

    for i, entry in enumerate(entries, 1):
        color_hex = status_pdf_colors.get(entry.status)
        if color_hex:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor(color_hex)))

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    doc.build(elements)
    output.seek(0)

    filename = f"supervision_timeline_{card_id}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =========================
# PROJECT TEMPLATES ENDPOINTS
# =========================

@app.post("/api/project-templates")
async def add_project_template(
    contract_id: int,
    template_url: str,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Добавить ссылку на шаблон проекта"""
    try:
        # Проверяем существование договора
        contract = db.query(Contract).filter(Contract.id == contract_id).first()
        if not contract:
            raise HTTPException(status_code=404, detail="Договор не найден")

        # Сохраняем шаблон как файл проекта
        file_record = ProjectFile(
            contract_id=contract_id,
            stage='template',
            file_type='template_url',
            public_link=template_url,
            yandex_path='',
            file_name=template_url
        )
        db.add(file_record)
        db.commit()
        db.refresh(file_record)

        return {'id': file_record.id, 'contract_id': contract_id, 'template_url': template_url}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка добавления шаблона: {str(e)}")


@app.get("/api/project-templates/{contract_id}")
async def get_project_templates(
    contract_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить все шаблоны для договора"""
    templates = db.query(ProjectFile).filter(
        ProjectFile.contract_id == contract_id,
        ProjectFile.stage == 'template'
    ).all()

    return [{
        'id': t.id,
        'contract_id': t.contract_id,
        'template_url': t.public_link,
        'created_at': t.upload_date.isoformat() if t.upload_date else None
    } for t in templates]


@app.delete("/api/project-templates/{template_id}")
async def delete_project_template(
    template_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Удалить шаблон проекта"""
    template = db.query(ProjectFile).filter(
        ProjectFile.id == template_id,
        ProjectFile.stage == 'template'
    ).first()

    if not template:
        raise HTTPException(status_code=404, detail="Шаблон не найден")

    db.delete(template)
    db.commit()

    return {"status": "success", "message": "Шаблон удален"}


# =========================
# CRM EXTENDED ENDPOINTS (Designer/Draftsman Reset)
# =========================

@app.post("/api/crm/cards/{card_id}/reset-designer")
async def reset_designer_completion(
    card_id: int,
    current_user: Employee = Depends(require_permission("crm_cards.reset_designer")),
    db: Session = Depends(get_db)
):
    """Сбросить отметку о завершении дизайнером"""
    try:
        # ИСПРАВЛЕНИЕ 06.02.2026: Добавлен поиск по '3д визуализация' для шаблонных проектов (#10)
        designer_executor = db.query(StageExecutor).filter(
            StageExecutor.crm_card_id == card_id,
            or_(
                StageExecutor.stage_name.ilike('%концепция%'),
                StageExecutor.stage_name.ilike('%визуализация%')
            )
        ).order_by(StageExecutor.id.desc()).first()

        if designer_executor:
            designer_executor.completed = False
            designer_executor.completed_date = None
            db.commit()

        return {"status": "success", "message": "Отметка дизайнера сброшена"}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка сброса: {str(e)}")


@app.post("/api/crm/cards/{card_id}/reset-draftsman")
async def reset_draftsman_completion(
    card_id: int,
    current_user: Employee = Depends(require_permission("crm_cards.reset_draftsman")),
    db: Session = Depends(get_db)
):
    """Сбросить отметку о завершении чертежником"""
    try:

        # Находим назначение чертежника
        draftsman_executor = db.query(StageExecutor).filter(
            StageExecutor.crm_card_id == card_id,
            or_(
                StageExecutor.stage_name.ilike('%чертежи%'),
                StageExecutor.stage_name.ilike('%планировочные%')
            )
        ).order_by(StageExecutor.id.desc()).first()

        if draftsman_executor:
            draftsman_executor.completed = False
            draftsman_executor.completed_date = None
            db.commit()

        return {"status": "success", "message": "Отметка чертежника сброшена"}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка сброса: {str(e)}")


@app.get("/api/crm/cards/{card_id}/approval-deadlines")
async def get_approval_stage_deadlines(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить дедлайны стадий согласования"""
    card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="CRM карточка не найдена")

    # Возвращаем данные о согласовании
    return {
        'card_id': card_id,
        'is_approved': card.is_approved,
        'approval_deadline': str(card.approval_deadline) if card.approval_deadline else None,
        'approval_stages': json.loads(card.approval_stages) if card.approval_stages else None
    }


@app.post("/api/crm/cards/{card_id}/complete-approval-stage")
async def complete_approval_stage(
    card_id: int,
    body: CompleteApprovalStageRequest,
    current_user: Employee = Depends(require_permission("crm_cards.complete_approval")),
    db: Session = Depends(get_db)
):
    """Завершить стадию согласования"""
    try:
        stage_name = body.stage_name

        card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="CRM карточка не найдена")

        # Обновляем approval_stages (JSON)
        current_stages = json.loads(card.approval_stages) if card.approval_stages else {}
        current_stages[stage_name] = {
            'completed': True,
            'completed_date': datetime.utcnow().isoformat(),
            'completed_by': current_user.id
        }
        card.approval_stages = json.dumps(current_stages)

        # Авто-установка is_approved когда текущая стадия согласована
        if stage_name == card.column_name:
            card.is_approved = True

        db.commit()

        return {
            "status": "success",
            "stage_name": stage_name,
            "completed": True,
            "is_approved": card.is_approved
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.patch("/api/crm/cards/{card_id}/stage-executor-deadline")
async def update_stage_executor_deadline(
    card_id: int,
    body: StageExecutorDeadlineRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить дедлайн исполнителя стадии"""
    try:
        stage_name = body.stage_name
        deadline = body.deadline

        stage_executor = db.query(StageExecutor).filter(
            StageExecutor.crm_card_id == card_id,
            StageExecutor.stage_name.ilike(f'%{stage_name}%')
        ).order_by(StageExecutor.id.desc()).first()

        if not stage_executor:
            raise HTTPException(status_code=404, detail="Назначение стадии не найдено")

        from datetime import datetime as dt
        stage_executor.deadline = dt.fromisoformat(deadline).date() if deadline else None
        db.commit()

        return {"status": "success", "stage_name": stage_executor.stage_name, "deadline": deadline}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.patch("/api/crm/cards/{card_id}/stage-executor/{stage_name}/complete")
async def complete_stage_for_executor(
    card_id: int,
    stage_name: str,
    body: CompleteStageExecutorRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Отметить стадию как выполненную для исполнителя"""
    try:
        executor_id = body.executor_id

        stage_executor = db.query(StageExecutor).filter(
            StageExecutor.crm_card_id == card_id,
            StageExecutor.stage_name == stage_name,
            StageExecutor.executor_id == executor_id
        ).first()

        # Fallback: поиск по подстроке stage_name
        if not stage_executor:
            stage_executor = db.query(StageExecutor).filter(
                StageExecutor.crm_card_id == card_id,
                StageExecutor.stage_name.ilike(f'%{stage_name}%'),
                StageExecutor.executor_id == executor_id
            ).order_by(StageExecutor.id.desc()).first()

        # Fallback 2: поиск только по card_id + executor_id (последнее назначение)
        if not stage_executor:
            stage_executor = db.query(StageExecutor).filter(
                StageExecutor.crm_card_id == card_id,
                StageExecutor.executor_id == executor_id
            ).order_by(StageExecutor.id.desc()).first()

        if not stage_executor:
            raise HTTPException(status_code=404, detail="Назначение стадии не найдено")

        stage_executor.completed = True
        stage_executor.completed_date = datetime.utcnow()
        db.commit()

        return {"status": "success", "stage_name": stage_name, "completed": True}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.get("/api/crm/cards/{card_id}/previous-executor")
async def get_previous_executor_by_position(
    card_id: int,
    position: str,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить предыдущего исполнителя по должности"""
    try:
        # Находим назначение для этой должности
        stage_executors = db.query(StageExecutor).filter(
            StageExecutor.crm_card_id == card_id
        ).order_by(StageExecutor.id.desc()).all()

        for se in stage_executors:
            employee = db.query(Employee).filter(Employee.id == se.executor_id).first()
            if employee and employee.position == position:
                return {'executor_id': se.executor_id, 'executor_name': employee.full_name}

        return {'executor_id': None}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.post("/api/crm/cards/{card_id}/manager-acceptance")
async def save_manager_acceptance(
    card_id: int,
    body: ManagerAcceptanceRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Сохранить принятие работы менеджером"""
    try:
        stage_name = body.stage_name
        executor_name = body.executor_name
        manager_id = body.manager_id

        card = db.query(CRMCard).filter(CRMCard.id == card_id).first()
        if not card:
            raise HTTPException(status_code=404, detail="CRM карточка не найдена")

        # Добавляем запись в историю действий
        history = ActionHistory(
            user_id=manager_id,
            action_type='acceptance',
            entity_type='stage',
            entity_id=card_id,
            description=f'Принятие работы: {stage_name} от {executor_name}'
        )
        db.add(history)
        db.commit()

        return {"status": "success", "stage_name": stage_name, "accepted_by": manager_id}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.get("/api/crm/cards/{card_id}/accepted-stages")
async def get_accepted_stages(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить список принятых стадий"""
    history = db.query(ActionHistory).filter(
        ActionHistory.entity_type == 'stage',
        ActionHistory.entity_id == card_id,
        ActionHistory.action_type == 'acceptance'
    ).all()

    return [{
        'id': h.id,
        'stage_name': h.description,
        'accepted_by': h.user_id,
        'accepted_date': h.action_date.isoformat() if h.action_date else None
    } for h in history]


# =========================
# PAYMENTS CRM ENDPOINTS
# =========================

@app.get("/api/payments/crm/{contract_id}")
async def get_payments_for_crm(
    contract_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить выплаты для CRM (не надзор)"""
    payments = db.query(Payment).filter(
        Payment.contract_id == contract_id,
        Payment.supervision_card_id == None
    ).all()

    result = []
    for p in payments:
        employee = db.query(Employee).filter(Employee.id == p.employee_id).first()
        result.append({
            'id': p.id,
            'contract_id': p.contract_id,
            'crm_card_id': p.crm_card_id,
            'employee_id': p.employee_id,
            'employee_name': employee.full_name if employee else 'Неизвестный',
            'position': employee.position if employee else '',
            'role': p.role,
            'stage_name': p.stage_name,
            'calculated_amount': float(p.calculated_amount) if p.calculated_amount else 0,
            'final_amount': float(p.final_amount) if p.final_amount else 0,
            'amount': float(p.final_amount) if p.final_amount else 0,
            'payment_type': p.payment_type,
            'report_month': p.report_month,
            'payment_status': p.payment_status,
            'is_paid': p.is_paid,
            'source': 'CRM'
        })

    return result


# =========================
# OPTIMIZED PAYMENTS ENDPOINTS
@app.get("/api/statistics/crm")
async def get_crm_statistics(
    project_type: str = "Индивидуальный",
    period: str = "all",
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить статистику CRM"""
    try:
        from sqlalchemy import extract

        if year is None:
            year = datetime.utcnow().year

        query = db.query(CRMCard).join(Contract).filter(
            Contract.project_type == project_type
        )

        if period == 'За год':
            query = query.filter(extract('year', CRMCard.created_at) == year)
        elif period == 'За месяц' and month:
            query = query.filter(
                extract('year', CRMCard.created_at) == year,
                extract('month', CRMCard.created_at) == month
            )

        cards = query.all()

        result = []
        for card in cards:
            result.append({
                'id': card.id,
                'contract_id': card.contract_id,
                'column_name': card.column_name,
                'contract_number': card.contract.contract_number,
                'address': card.contract.address,
                'area': float(card.contract.area) if card.contract.area else 0
            })

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.get("/api/statistics/crm/filtered")
async def get_crm_statistics_filtered(
    project_type: str,
    period: str,
    year: int,
    quarter: Optional[int] = None,
    month: Optional[int] = None,
    project_id: Optional[int] = None,
    executor_id: Optional[int] = None,
    stage_name: Optional[str] = None,
    status_filter: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить статистику CRM с фильтрами"""
    try:
        from sqlalchemy import extract

        query = db.query(CRMCard).join(Contract).filter(
            Contract.project_type == project_type
        )

        if period == 'За год':
            query = query.filter(extract('year', CRMCard.created_at) == year)
        elif period == 'За квартал' and quarter:
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            query = query.filter(
                extract('year', CRMCard.created_at) == year,
                extract('month', CRMCard.created_at).between(start_month, end_month)
            )
        elif period == 'За месяц' and month:
            query = query.filter(
                extract('year', CRMCard.created_at) == year,
                extract('month', CRMCard.created_at) == month
            )

        if project_id:
            query = query.filter(CRMCard.contract_id == project_id)

        if status_filter:
            query = query.filter(CRMCard.column_name == status_filter)

        cards = query.all()

        # Фильтрация по исполнителю или стадии
        if executor_id or stage_name:
            filtered_cards = []
            for card in cards:
                executors = db.query(StageExecutor).filter(
                    StageExecutor.crm_card_id == card.id
                ).all()

                if executor_id:
                    if any(e.executor_id == executor_id for e in executors):
                        filtered_cards.append(card)
                elif stage_name:
                    if any(stage_name in e.stage_name for e in executors):
                        filtered_cards.append(card)
            cards = filtered_cards

        result = []
        for card in cards:
            result.append({
                'id': card.id,
                'contract_id': card.contract_id,
                'column_name': card.column_name,
                'contract_number': card.contract.contract_number,
                'address': card.contract.address,
                'area': float(card.contract.area) if card.contract.area else 0,
                'is_approved': card.is_approved
            })

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.get("/api/statistics/approvals")
async def get_approval_statistics(
    project_type: str,
    period: str,
    year: int,
    quarter: Optional[int] = None,
    month: Optional[int] = None,
    project_id: Optional[int] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить статистику согласований"""
    try:
        from sqlalchemy import extract

        query = db.query(CRMCard).join(Contract).filter(
            Contract.project_type == project_type,
            CRMCard.is_approved == True
        )

        if period == 'За год':
            query = query.filter(extract('year', CRMCard.created_at) == year)
        elif period == 'За квартал' and quarter:
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            query = query.filter(
                extract('year', CRMCard.created_at) == year,
                extract('month', CRMCard.created_at).between(start_month, end_month)
            )
        elif period == 'За месяц' and month:
            query = query.filter(
                extract('year', CRMCard.created_at) == year,
                extract('month', CRMCard.created_at) == month
            )

        if project_id:
            query = query.filter(CRMCard.contract_id == project_id)

        cards = query.all()

        result = []
        for card in cards:
            result.append({
                'id': card.id,
                'contract_id': card.contract_id,
                'contract_number': card.contract.contract_number,
                'address': card.contract.address,
                'is_approved': card.is_approved,
                'approval_deadline': str(card.approval_deadline) if card.approval_deadline else None,
                'approval_stages': json.loads(card.approval_stages) if card.approval_stages else None
            })

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.get("/api/statistics/general")
async def get_general_statistics(
    year: int,
    quarter: Optional[int] = None,
    month: Optional[int] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить общую статистику"""
    try:
        from sqlalchemy import extract, func

        # Базовый запрос договоров
        query = db.query(Contract).filter(extract('year', Contract.created_at) == year)

        if quarter:
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            query = query.filter(extract('month', Contract.created_at).between(start_month, end_month))
        if month:
            query = query.filter(extract('month', Contract.created_at) == month)

        contracts = query.all()

        total_orders = len(contracts)
        total_area = sum(c.area or 0 for c in contracts)
        total_amount = sum(c.total_amount or 0 for c in contracts)

        individual_count = len([c for c in contracts if c.project_type == 'Индивидуальный'])
        template_count = len([c for c in contracts if c.project_type == 'Шаблонный'])

        active_count = len([c for c in contracts if c.status not in ['СДАН', 'РАСТОРГНУТ']])
        completed_count = len([c for c in contracts if c.status == 'СДАН'])
        cancelled_count = len([c for c in contracts if c.status == 'РАСТОРГНУТ'])

        # Сотрудники
        active_employees = db.query(Employee).filter(Employee.status == 'активный').count()

        # Платежи
        payments_query = db.query(Payment).filter(extract('year', Payment.created_at) == year)
        if quarter:
            payments_query = payments_query.filter(
                extract('month', Payment.created_at).between(start_month, end_month)
            )
        if month:
            payments_query = payments_query.filter(extract('month', Payment.created_at) == month)

        payments = payments_query.all()
        total_payments = sum(p.final_amount or 0 for p in payments)
        paid_payments = sum(p.final_amount or 0 for p in payments if p.is_paid)

        return {
            'total_orders': total_orders,
            'active': active_count,
            'completed': completed_count,
            'cancelled': cancelled_count,
            'individual_count': individual_count,
            'template_count': template_count,
            'total_area': total_area,
            'total_amount': total_amount,
            'active_employees': active_employees,
            'total_payments': total_payments,
            'paid_payments': paid_payments,
            'pending_payments': total_payments - paid_payments
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.get("/api/statistics/funnel")
async def get_funnel_statistics(
    year: Optional[int] = None,
    project_type: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Воронка проектов: количество CRM-карточек по колонкам Kanban"""
    try:
        from sqlalchemy import extract, func

        query = db.query(
            CRMCard.column_name,
            func.count(CRMCard.id).label("count")
        ).join(Contract, CRMCard.contract_id == Contract.id)

        if year:
            query = query.filter(extract('year', Contract.created_at) == year)
        if project_type:
            query = query.filter(Contract.project_type == project_type)

        rows = query.group_by(CRMCard.column_name).all()

        funnel = {row.column_name: row.count for row in rows}
        return {"funnel": funnel, "total": sum(funnel.values())}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.get("/api/statistics/executor-load")
async def get_executor_load(
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Нагрузка на исполнителей: количество активных стадий на каждого"""
    try:
        from sqlalchemy import extract, func

        query = db.query(
            Employee.full_name,
            func.count(StageExecutor.id).label("active_stages")
        ).join(
            StageExecutor, StageExecutor.employee_id == Employee.id
        ).join(
            CRMCard, CRMCard.id == StageExecutor.crm_card_id
        ).filter(
            CRMCard.column_name.notin_(['СДАН', 'РАСТОРГНУТ'])
        )

        if year or month:
            query = query.join(Contract, CRMCard.contract_id == Contract.id)
            if year:
                query = query.filter(extract('year', Contract.created_at) == year)
            if month:
                query = query.filter(extract('month', Contract.created_at) == month)

        rows = query.group_by(Employee.full_name).order_by(
            func.count(StageExecutor.id).desc()
        ).limit(15).all()

        return [{"name": row.full_name, "active_stages": row.active_stages} for row in rows]

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


# =========================
# AGENTS ENDPOINTS
# =========================

@app.post("/api/agents")
async def add_agent(
    name: str,
    color: str = "#FFFFFF",
    current_user: Employee = Depends(require_permission("agents.create")),
    db: Session = Depends(get_db)
):
    """Добавить нового агента"""
    try:
        # Проверяем, не существует ли уже агент с таким именем
        existing = db.query(Employee).filter(
            Employee.full_name == name,
            Employee.position == 'Агент'
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Агент с таким именем уже существует")

        # Создаём агента как сотрудника с обязательными полями
        agent = Employee(
            full_name=name,
            login=f"agent_{name.lower().replace(' ', '_')}",
            position='Агент',
            department='Агенты',
            phone='',
            password_hash='',
            agent_color=color,
            status='активный'
        )
        db.add(agent)
        db.commit()
        db.refresh(agent)

        return {"status": "success", "id": agent.id, "name": name, "color": color}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.patch("/api/agents/{name}/color")
async def update_agent_color(
    name: str,
    color: str,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить цвет агента"""
    try:
        agent = db.query(Employee).filter(
            Employee.full_name == name,
            or_(Employee.position == 'Агент', Employee.secondary_position == 'Агент')
        ).first()

        if not agent:
            raise HTTPException(status_code=404, detail="Агент не найден")

        agent.agent_color = color
        db.commit()

        return {"status": "success", "name": name, "color": color}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


# =========================
# HEARTBEAT (ONLINE STATUS)
# =========================

@app.post("/api/heartbeat")
async def send_heartbeat(
    employee_id: int = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Отправить heartbeat для поддержания онлайн-статуса.
    Возвращает список онлайн пользователей.
    """
    try:
        # Обновляем last_activity текущего пользователя
        current_user.last_activity = datetime.utcnow()
        current_user.is_online = True
        db.commit()

        # Определяем порог активности (5 минут)
        activity_threshold = datetime.utcnow() - timedelta(minutes=5)

        # Получаем список онлайн пользователей
        online_employees = db.query(Employee).filter(
            Employee.last_activity > activity_threshold,
            Employee.is_online == True,
            Employee.status == 'активный'
        ).all()

        online_users = [
            {
                'id': emp.id,
                'full_name': emp.full_name,
                'position': emp.position,
                'last_activity': emp.last_activity.isoformat() if emp.last_activity else None
            }
            for emp in online_employees
        ]

        return {
            'status': 'ok',
            'online_users': online_users,
            'online_count': len(online_users)
        }

    except Exception as e:
        logger.error(f"Heartbeat error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =========================
# CONCURRENT EDITING LOCKS
# =========================

@app.post("/api/locks")
async def create_lock(
    lock_data: LockRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Создать блокировку записи для редактирования.
    Принимает JSON body: {entity_type, entity_id, employee_id (опционально)}.
    Возвращает 409 если запись уже заблокирована другим пользователем.
    """
    entity_type = lock_data.entity_type
    entity_id = lock_data.entity_id
    employee_id = lock_data.employee_id or current_user.id
    try:
        # Проверяем существующую блокировку
        existing_lock = db.query(ConcurrentEdit).filter(
            ConcurrentEdit.entity_type == entity_type,
            ConcurrentEdit.entity_id == entity_id
        ).first()

        if existing_lock:
            # Проверяем, не истекла ли блокировка (2 минуты)
            if existing_lock.expires_at and existing_lock.expires_at < datetime.utcnow():
                # Блокировка истекла, удаляем её
                db.delete(existing_lock)
                db.commit()
            elif existing_lock.employee_id != employee_id:
                # Запись заблокирована другим пользователем
                locked_by = db.query(Employee).filter(
                    Employee.id == existing_lock.employee_id
                ).first()

                locked_by_name = locked_by.full_name if locked_by else 'другим пользователем'

                raise HTTPException(
                    status_code=409,
                    detail={
                        'message': 'Запись заблокирована',
                        'locked_by': locked_by_name,
                        'locked_at': existing_lock.locked_at.isoformat()
                    }
                )
            else:
                # Обновляем время блокировки
                existing_lock.locked_at = datetime.utcnow()
                existing_lock.expires_at = datetime.utcnow() + timedelta(minutes=2)
                db.commit()
                return {'status': 'renewed', 'entity_type': entity_type, 'entity_id': entity_id}

        # Получаем реальную сессию текущего пользователя
        user_session = db.query(UserSession).filter(
            UserSession.employee_id == current_user.id
        ).order_by(UserSession.last_activity.desc()).first()

        if not user_session:
            # Создаём сессию если нет
            user_session = UserSession(
                employee_id=current_user.id,
                session_token="lock-session",
                ip_address="0.0.0.0",
                last_activity=datetime.utcnow(),
                is_active=True
            )
            db.add(user_session)
            db.flush()

        # Создаём новую блокировку
        new_lock = ConcurrentEdit(
            entity_type=entity_type,
            entity_id=entity_id,
            employee_id=employee_id,
            session_id=user_session.id,
            locked_at=datetime.utcnow(),
            expires_at=datetime.utcnow() + timedelta(minutes=2)
        )
        db.add(new_lock)
        db.commit()

        return {'status': 'created', 'entity_type': entity_type, 'entity_id': entity_id}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Lock creation error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/locks/{entity_type}/{entity_id}")
async def check_lock(
    entity_type: str,
    entity_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Проверить блокировку записи"""
    try:
        lock = db.query(ConcurrentEdit).filter(
            ConcurrentEdit.entity_type == entity_type,
            ConcurrentEdit.entity_id == entity_id
        ).first()

        if not lock:
            return {'is_locked': False, 'locked_by': None}

        # Проверяем, не истекла ли блокировка
        if lock.expires_at and lock.expires_at < datetime.utcnow():
            db.delete(lock)
            db.commit()
            return {'is_locked': False, 'locked_by': None}

        # Получаем имя пользователя, заблокировавшего запись
        locked_by = db.query(Employee).filter(
            Employee.id == lock.employee_id
        ).first()

        return {
            'is_locked': True,
            'locked_by': locked_by.full_name if locked_by else 'неизвестный пользователь',
            'locked_at': lock.locked_at.isoformat(),
            'expires_at': lock.expires_at.isoformat() if lock.expires_at else None,
            'is_own_lock': lock.employee_id == current_user.id
        }

    except Exception as e:
        logger.error(f"Lock check error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/locks/user/{employee_id}")
async def release_user_locks(
    employee_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Снять все блокировки пользователя"""
    try:
        # Только сам пользователь или админ может снять блокировки
        if employee_id != current_user.id and current_user.position not in ['Руководитель студии', 'Старший менеджер проектов']:
            raise HTTPException(status_code=403, detail="Нет прав")

        locks = db.query(ConcurrentEdit).filter(
            ConcurrentEdit.employee_id == employee_id
        ).all()

        count = len(locks)
        for lock in locks:
            db.delete(lock)

        db.commit()

        return {'status': 'released', 'count': count}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"User locks release error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/locks/{entity_type}/{entity_id}")
async def release_lock(
    entity_type: str,
    entity_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Снять блокировку записи"""
    try:
        lock = db.query(ConcurrentEdit).filter(
            ConcurrentEdit.entity_type == entity_type,
            ConcurrentEdit.entity_id == entity_id
        ).first()

        if lock:
            # Только владелец или админ может снять блокировку
            if lock.employee_id == current_user.id or current_user.position in ['Руководитель студии', 'Старший менеджер проектов']:
                db.delete(lock)
                db.commit()
                return {'status': 'released'}
            else:
                raise HTTPException(status_code=403, detail="Нельзя снять чужую блокировку")

        return {'status': 'not_found'}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Lock release error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ========== DASHBOARD ENDPOINTS ==========

@app.get("/api/dashboard/clients")
async def get_clients_dashboard(
    year: Optional[int] = None,
    agent_type: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить статистику для дашборда страницы Клиенты"""
    try:
        # 1. Всего клиентов
        total_clients = db.query(Client).count()

        # 2. Всего физлиц
        total_individual = db.query(Client).filter(
            Client.client_type == 'Физическое лицо'
        ).count()

        # 3. Всего юрлиц
        total_legal = db.query(Client).filter(
            Client.client_type == 'Юридическое лицо'
        ).count()

        # 4. Клиенты за год
        # contract_date хранится как VARCHAR, используем LIKE для поиска года
        clients_by_year = 0
        if year:
            year_str = str(year)
            clients_by_year = db.query(Contract.client_id).filter(
                Contract.contract_date.like(f'%{year_str}%')
            ).distinct().count()

        # 5. Клиенты агента (всего)
        agent_clients_total = 0
        if agent_type:
            agent_clients_total = db.query(Contract.client_id).filter(
                Contract.agent_type == agent_type
            ).distinct().count()

        # 6. Клиенты агента за год
        agent_clients_by_year = 0
        if agent_type and year:
            year_str = str(year)
            agent_clients_by_year = db.query(Contract.client_id).filter(
                Contract.agent_type == agent_type,
                Contract.contract_date.like(f'%{year_str}%')
            ).distinct().count()

        return {
            'total_clients': total_clients,
            'total_individual': total_individual,
            'total_legal': total_legal,
            'clients_by_year': clients_by_year,
            'agent_clients_total': agent_clients_total,
            'agent_clients_by_year': agent_clients_by_year
        }

    except Exception as e:
        logger.error(f"Clients dashboard error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dashboard/contracts")
async def get_contracts_dashboard(
    year: Optional[int] = None,
    agent_type: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить статистику для дашборда страницы Договора"""
    try:
        # 1-2. Индивидуальные заказы и площадь
        individual_query = db.query(
            func.count(Contract.id),
            func.coalesce(func.sum(Contract.area), 0)
        ).filter(Contract.project_type == 'Индивидуальный')
        individual_orders, individual_area = individual_query.first()

        # 3-4. Шаблонные заказы и площадь
        template_query = db.query(
            func.count(Contract.id),
            func.coalesce(func.sum(Contract.area), 0)
        ).filter(Contract.project_type == 'Шаблонный')
        template_orders, template_area = template_query.first()

        # 5. Заказы агента за год
        # contract_date хранится как VARCHAR, используем LIKE для поиска года
        agent_orders_by_year = 0
        if agent_type and year:
            year_str = str(year)
            agent_orders_by_year = db.query(Contract).filter(
                Contract.agent_type == agent_type,
                Contract.contract_date.like(f'%{year_str}%')
            ).count()

        # 6. Площадь агента за год
        agent_area_by_year = 0
        if agent_type and year:
            year_str = str(year)
            result = db.query(
                func.coalesce(func.sum(Contract.area), 0)
            ).filter(
                Contract.agent_type == agent_type,
                Contract.contract_date.like(f'%{year_str}%')
            ).scalar()
            agent_area_by_year = float(result) if result else 0

        return {
            'individual_orders': individual_orders,
            'individual_area': float(individual_area),
            'template_orders': template_orders,
            'template_area': float(template_area),
            'agent_orders_by_year': agent_orders_by_year,
            'agent_area_by_year': agent_area_by_year
        }

    except Exception as e:
        logger.error(f"Contracts dashboard error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dashboard/crm")
async def get_crm_dashboard(
    project_type: str,
    agent_type: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить статистику для дашборда СРМ (Индивидуальные/Шаблонные/Надзор)"""
    try:
        # Определяем таблицу CRM
        if project_type == 'Авторский надзор':
            crm_table = SupervisionCard
            contract_condition = Contract.status == 'АВТОРСКИЙ НАДЗОР'
        elif project_type == 'Индивидуальный':
            crm_table = CRMCard
            contract_condition = Contract.project_type == 'Индивидуальный'
        else:  # Шаблонный
            crm_table = CRMCard
            contract_condition = Contract.project_type == 'Шаблонный'

        # 1-2. Всего заказов и площадь (из договоров)
        total_query = db.query(
            func.count(Contract.id),
            func.coalesce(func.sum(Contract.area), 0)
        ).filter(contract_condition)
        total_orders, total_area = total_query.first()

        # 3. Активные заказы в СРМ
        active_orders = db.query(crm_table).count()

        # 4. Архивные заказы (считаем из истории или архивной таблицы)
        # Для упрощения считаем как разницу между всеми договорами и активными
        archive_orders = total_orders - active_orders if total_orders > active_orders else 0

        # 5. Активные заказы агента
        agent_active_orders = 0
        if agent_type:
            agent_active_orders = db.query(crm_table).join(
                Contract, crm_table.contract_id == Contract.id
            ).filter(Contract.agent_type == agent_type).count()

        # 6. Архивные заказы агента
        agent_archive_orders = 0
        if agent_type:
            agent_total = db.query(Contract).filter(
                contract_condition,
                Contract.agent_type == agent_type
            ).count()
            agent_archive_orders = agent_total - agent_active_orders if agent_total > agent_active_orders else 0

        return {
            'total_orders': total_orders,
            'total_area': float(total_area),
            'active_orders': active_orders,
            'archive_orders': archive_orders,
            'agent_active_orders': agent_active_orders,
            'agent_archive_orders': agent_archive_orders
        }

    except Exception as e:
        logger.error(f"CRM dashboard error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dashboard/employees")
async def get_employees_dashboard(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить статистику для дашборда страницы Сотрудники"""
    try:
        # Используем func.lower для регистронезависимого сравнения
        # Значения в БД: 'активный', 'уволен', 'в резерве'
        # Отделы: 'Административный отдел', 'Проектный отдел', 'Исполнительный отдел'

        # 1. Активные сотрудники
        active_employees = db.query(Employee).filter(
            func.lower(Employee.status) == 'активный'
        ).count()

        # 2. Резерв
        reserve_employees = db.query(Employee).filter(
            func.lower(Employee.status) == 'в резерве'
        ).count()

        # 3. Руководящий состав (по должностям, как во вкладках UI)
        admin_positions = ['Руководитель студии', 'Старший менеджер проектов', 'СДП', 'ГАП']
        active_management = db.query(Employee).filter(
            func.lower(Employee.status) == 'активный',
            Employee.position.in_(admin_positions)
        ).count()

        # 4. Проектный отдел (по должностям)
        project_positions = ['Дизайнер', 'Чертёжник']
        active_projects_dept = db.query(Employee).filter(
            func.lower(Employee.status) == 'активный',
            Employee.position.in_(project_positions)
        ).count()

        # 5. Исполнительный отдел (по должностям)
        exec_positions = ['Менеджер', 'ДАН', 'Замерщик']
        active_execution_dept = db.query(Employee).filter(
            func.lower(Employee.status) == 'активный',
            Employee.position.in_(exec_positions)
        ).count()

        # 6. Дни рождения (ближайшие 30 дней)
        today = datetime.now()
        upcoming_birthdays = 0
        nearest_birthday = None
        min_days = 366

        employees = db.query(Employee).filter(
            func.lower(Employee.status) == 'активный'
        ).all()
        from datetime import date as date_cls
        birthday_names = []
        for emp in employees:
            if emp.birth_date:
                try:
                    # birth_date хранится как строка 'yyyy-MM-dd'
                    if isinstance(emp.birth_date, str):
                        bd = datetime.strptime(emp.birth_date, '%Y-%m-%d').date()
                    else:
                        bd = emp.birth_date if isinstance(emp.birth_date, date_cls) else emp.birth_date.date()

                    birth_this_year = bd.replace(year=today.year)
                    if birth_this_year < today.date():
                        birth_this_year = bd.replace(year=today.year + 1)

                    days_until = (birth_this_year - today.date()).days
                    if 0 <= days_until <= 30:
                        upcoming_birthdays += 1

                    if days_until < min_days:
                        min_days = days_until
                        birthday_names = [emp.full_name]
                    elif days_until == min_days:
                        birthday_names.append(emp.full_name)
                except Exception:
                    continue

        nearest_birthday_text = ", ".join(birthday_names) if birthday_names else "Нет данных"

        return {
            'active_employees': active_employees,
            'reserve_employees': reserve_employees,
            'active_management': active_management,
            'active_admin': active_management,
            'active_projects_dept': active_projects_dept,
            'active_project': active_projects_dept,
            'active_execution_dept': active_execution_dept,
            'active_execution': active_execution_dept,
            'upcoming_birthdays': upcoming_birthdays,
            'nearest_birthday': nearest_birthday_text
        }

    except Exception as e:
        logger.error(f"Employees dashboard error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dashboard/salaries")
async def get_salaries_dashboard(
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить статистику для дашборда страницы Зарплаты"""
    try:
        # 1. Всего выплачено (все время)
        total_paid = db.query(
            func.coalesce(func.sum(Payment.final_amount), 0)
        ).filter(Payment.payment_status == 'paid').scalar()

        # 2. Выплачено за год
        paid_by_year = 0
        if year:
            paid_by_year = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).filter(
                Payment.payment_status == 'paid',
                Payment.report_month.like(f'{year}-%')
            ).scalar()

        # 3. Выплачено за месяц
        paid_by_month = 0
        if year and month:
            paid_by_month = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).filter(
                Payment.payment_status == 'paid',
                Payment.report_month == f'{year}-{month:02d}'
            ).scalar()

        # 4. Индивидуальные за год
        individual_by_year = 0
        if year:
            individual_by_year = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).join(Contract).filter(
                Payment.payment_status == 'paid',
                Payment.report_month.like(f'{year}-%'),
                Contract.project_type == 'Индивидуальный'
            ).scalar()

        # 5. Шаблонные за год
        template_by_year = 0
        if year:
            template_by_year = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).join(Contract).filter(
                Payment.payment_status == 'paid',
                Payment.report_month.like(f'{year}-%'),
                Contract.project_type == 'Шаблонный'
            ).scalar()

        # 6. Надзор за год
        supervision_by_year = 0
        if year:
            supervision_by_year = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).join(Contract).filter(
                Payment.payment_status == 'paid',
                Payment.report_month.like(f'{year}-%'),
                Contract.status == 'АВТОРСКИЙ НАДЗОР'
            ).scalar()

        return {
            'total_paid': float(total_paid) if total_paid else 0,
            'paid_by_year': float(paid_by_year) if paid_by_year else 0,
            'paid_by_month': float(paid_by_month) if paid_by_month else 0,
            'individual_by_year': float(individual_by_year) if individual_by_year else 0,
            'template_by_year': float(template_by_year) if template_by_year else 0,
            'supervision_by_year': float(supervision_by_year) if supervision_by_year else 0
        }

    except Exception as e:
        logger.error(f"Salaries dashboard error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dashboard/salaries-by-type")
async def get_salaries_by_type_dashboard(
    payment_type: str,
    year: Optional[int] = None,
    month: Optional[int] = None,
    agent_type: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить статистику для дашборда вкладок зарплат по типу выплат

    Args:
        payment_type: Тип вкладки ('all', 'individual', 'template', 'salary', 'supervision')
        year: Год для фильтра
        month: Месяц для фильтра
        agent_type: Тип агента для фильтра

    Returns:
        Dict с ключами: total_paid, paid_by_year, paid_by_month, payments_count, to_pay_amount, by_agent
    """
    try:
        # Базовый запрос
        if payment_type == 'salary':
            # Оклады - из таблицы salaries
            # ВАЖНО: таблица salaries не имеет поля status/payment_status
            # Оклады считаются всегда выплаченными

            # Всего выплачено (все оклады)
            total_paid = db.query(
                func.coalesce(func.sum(Salary.amount), 0)
            ).scalar()

            # За год
            paid_by_year = 0
            if year:
                paid_by_year = db.query(
                    func.coalesce(func.sum(Salary.amount), 0)
                ).filter(
                    Salary.report_month.like(f'{year}-%')
                ).scalar()

            # За месяц
            paid_by_month = 0
            if year and month:
                paid_by_month = db.query(
                    func.coalesce(func.sum(Salary.amount), 0)
                ).filter(
                    Salary.report_month == f'{year}-{month:02d}'
                ).scalar()

            # Количество выплат (все оклады)
            payments_count = db.query(func.count(Salary.id)).scalar()

            # К оплате - для окладов всегда 0 (нет статуса)
            to_pay_amount = 0

            # По агенту - для окладов не применимо
            by_agent = 0

        else:
            # Условие фильтрации по типу
            query_filter = []
            if payment_type == 'individual':
                query_filter.append(Contract.project_type == 'Индивидуальный')
            elif payment_type == 'template':
                query_filter.append(Contract.project_type == 'Шаблонный')
            elif payment_type == 'supervision':
                query_filter.append(Contract.status == 'АВТОРСКИЙ НАДЗОР')
            # 'all' - без дополнительного фильтра

            # Всего выплачено
            q = db.query(func.coalesce(func.sum(Payment.final_amount), 0))
            if query_filter:
                q = q.join(Contract)
                for f in query_filter:
                    q = q.filter(f)
            total_paid = q.filter(Payment.payment_status == 'paid').scalar()

            # За год
            paid_by_year = 0
            if year:
                q = db.query(func.coalesce(func.sum(Payment.final_amount), 0))
                if query_filter:
                    q = q.join(Contract)
                    for f in query_filter:
                        q = q.filter(f)
                paid_by_year = q.filter(
                    Payment.payment_status == 'paid',
                    Payment.report_month.like(f'{year}-%')
                ).scalar()

            # За месяц
            paid_by_month = 0
            if year and month:
                q = db.query(func.coalesce(func.sum(Payment.final_amount), 0))
                if query_filter:
                    q = q.join(Contract)
                    for f in query_filter:
                        q = q.filter(f)
                paid_by_month = q.filter(
                    Payment.payment_status == 'paid',
                    Payment.report_month == f'{year}-{month:02d}'
                ).scalar()

            # Количество выплат (paid)
            q = db.query(func.count(Payment.id))
            if query_filter:
                q = q.join(Contract)
                for f in query_filter:
                    q = q.filter(f)
            payments_count = q.filter(Payment.payment_status == 'paid').scalar()

            # К оплате
            q = db.query(func.coalesce(func.sum(Payment.final_amount), 0))
            if query_filter:
                q = q.join(Contract)
                for f in query_filter:
                    q = q.filter(f)
            to_pay_amount = q.filter(Payment.payment_status == 'to_pay').scalar()

            # По агенту
            by_agent = 0
            if agent_type:
                q = db.query(func.coalesce(func.sum(Payment.final_amount), 0)).join(Contract)
                if query_filter:
                    for f in query_filter:
                        q = q.filter(f)
                by_agent = q.filter(
                    Payment.payment_status == 'paid',
                    Contract.agent_type == agent_type
                ).scalar()

        return {
            'total_paid': float(total_paid) if total_paid else 0,
            'paid_by_year': float(paid_by_year) if paid_by_year else 0,
            'paid_by_month': float(paid_by_month) if paid_by_month else 0,
            'payments_count': int(payments_count) if payments_count else 0,
            'to_pay_amount': float(to_pay_amount) if to_pay_amount else 0,
            'by_agent': float(by_agent) if by_agent else 0
        }

    except Exception as e:
        logger.error(f"Salaries by type dashboard error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dashboard/salaries-all")
async def get_salaries_all_dashboard(
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Дашборд 'Все выплаты': total_paid, paid_by_year, paid_by_month, individual/template/supervision_by_year"""
    try:
        # Всего выплачено
        total_paid = db.query(
            func.coalesce(func.sum(Payment.final_amount), 0)
        ).filter(Payment.payment_status == 'paid').scalar()

        # За год
        paid_by_year = 0
        individual_by_year = 0
        template_by_year = 0
        supervision_by_year = 0

        if year:
            paid_by_year = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).filter(
                Payment.payment_status == 'paid',
                Payment.report_month.like(f'{year}-%')
            ).scalar()

            # Индивидуальные за год
            individual_by_year = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).join(Contract).filter(
                Payment.payment_status == 'paid',
                Payment.report_month.like(f'{year}-%'),
                Contract.project_type == 'Индивидуальный'
            ).scalar()

            # Шаблонные за год
            template_by_year = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).join(Contract).filter(
                Payment.payment_status == 'paid',
                Payment.report_month.like(f'{year}-%'),
                Contract.project_type == 'Шаблонный'
            ).scalar()

            # Авторский надзор за год - платежи с supervision_card_id
            supervision_by_year = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).filter(
                Payment.payment_status == 'paid',
                Payment.report_month.like(f'{year}-%'),
                Payment.supervision_card_id.isnot(None)
            ).scalar()

        # За месяц
        paid_by_month = 0
        if year and month:
            paid_by_month = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).filter(
                Payment.payment_status == 'paid',
                Payment.report_month == f'{year}-{month:02d}'
            ).scalar()

        return {
            'total_paid': float(total_paid) if total_paid else 0,
            'paid_by_year': float(paid_by_year) if paid_by_year else 0,
            'paid_by_month': float(paid_by_month) if paid_by_month else 0,
            'individual_by_year': float(individual_by_year) if individual_by_year else 0,
            'template_by_year': float(template_by_year) if template_by_year else 0,
            'supervision_by_year': float(supervision_by_year) if supervision_by_year else 0
        }

    except Exception as e:
        logger.error(f"Salaries all dashboard error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dashboard/salaries-individual")
async def get_salaries_individual_dashboard(
    year: Optional[int] = None,
    month: Optional[int] = None,
    agent_type: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Дашборд 'Индивидуальные': total_paid, paid_by_year, paid_by_month, by_agent, avg_payment, payments_count"""
    try:
        base_q = db.query(Payment).join(Contract).filter(
            Payment.payment_status == 'paid',
            Contract.project_type == 'Индивидуальный'
        )

        # Всего выплачено и средний чек
        total_result = db.query(
            func.coalesce(func.sum(Payment.final_amount), 0),
            func.coalesce(func.avg(Payment.final_amount), 0)
        ).join(Contract).filter(
            Payment.payment_status == 'paid',
            Contract.project_type == 'Индивидуальный'
        ).first()
        total_paid = total_result[0]
        avg_payment = total_result[1]

        # За год и кол-во выплат
        paid_by_year = 0
        payments_count = 0
        if year:
            year_result = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0),
                func.count(Payment.id)
            ).join(Contract).filter(
                Payment.payment_status == 'paid',
                Contract.project_type == 'Индивидуальный',
                Payment.report_month.like(f'{year}-%')
            ).first()
            paid_by_year = year_result[0]
            payments_count = year_result[1]

        # За месяц
        paid_by_month = 0
        if year and month:
            paid_by_month = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).join(Contract).filter(
                Payment.payment_status == 'paid',
                Contract.project_type == 'Индивидуальный',
                Payment.report_month == f'{year}-{month:02d}'
            ).scalar()

        # По агенту
        by_agent = 0
        if agent_type:
            by_agent = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).join(Contract).filter(
                Payment.payment_status == 'paid',
                Contract.project_type == 'Индивидуальный',
                Contract.agent_type == agent_type
            ).scalar()

        result = {
            'total_paid': float(total_paid) if total_paid else 0,
            'paid_by_year': float(paid_by_year) if paid_by_year else 0,
            'paid_by_month': float(paid_by_month) if paid_by_month else 0,
            'by_agent': float(by_agent) if by_agent else 0,
            'avg_payment': float(avg_payment) if avg_payment else 0,
            'payments_count': int(payments_count) if payments_count else 0
        }
        logger.info(f"[DASHBOARD] salaries-individual: year={year}, month={month}, agent={agent_type} -> {result}")
        return result

    except Exception as e:
        logger.error(f"Salaries individual dashboard error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dashboard/salaries-template")
async def get_salaries_template_dashboard(
    year: Optional[int] = None,
    month: Optional[int] = None,
    agent_type: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Дашборд 'Шаблонные': total_paid, paid_by_year, paid_by_month, by_agent, avg_payment, payments_count"""
    try:
        # Всего выплачено и средний чек
        total_result = db.query(
            func.coalesce(func.sum(Payment.final_amount), 0),
            func.coalesce(func.avg(Payment.final_amount), 0)
        ).join(Contract).filter(
            Payment.payment_status == 'paid',
            Contract.project_type == 'Шаблонный'
        ).first()
        total_paid = total_result[0]
        avg_payment = total_result[1]

        # За год и кол-во выплат
        paid_by_year = 0
        payments_count = 0
        if year:
            year_result = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0),
                func.count(Payment.id)
            ).join(Contract).filter(
                Payment.payment_status == 'paid',
                Contract.project_type == 'Шаблонный',
                Payment.report_month.like(f'{year}-%')
            ).first()
            paid_by_year = year_result[0]
            payments_count = year_result[1]

        # За месяц
        paid_by_month = 0
        if year and month:
            paid_by_month = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).join(Contract).filter(
                Payment.payment_status == 'paid',
                Contract.project_type == 'Шаблонный',
                Payment.report_month == f'{year}-{month:02d}'
            ).scalar()

        # По агенту
        by_agent = 0
        if agent_type:
            by_agent = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).join(Contract).filter(
                Payment.payment_status == 'paid',
                Contract.project_type == 'Шаблонный',
                Contract.agent_type == agent_type
            ).scalar()

        return {
            'total_paid': float(total_paid) if total_paid else 0,
            'paid_by_year': float(paid_by_year) if paid_by_year else 0,
            'paid_by_month': float(paid_by_month) if paid_by_month else 0,
            'by_agent': float(by_agent) if by_agent else 0,
            'avg_payment': float(avg_payment) if avg_payment else 0,
            'payments_count': int(payments_count) if payments_count else 0
        }

    except Exception as e:
        logger.error(f"Salaries template dashboard error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dashboard/salaries-salary")
async def get_salaries_salary_dashboard(
    year: Optional[int] = None,
    month: Optional[int] = None,
    project_type: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Дашборд 'Оклады': total_paid, paid_by_year, paid_by_month, by_project_type, avg_salary, employees_count"""
    try:
        # Всего выплачено и средний оклад (только оплаченные)
        total_result = db.query(
            func.coalesce(func.sum(Salary.amount), 0),
            func.coalesce(func.avg(Salary.amount), 0)
        ).filter(
            Salary.payment_status == 'paid'
        ).first()
        total_paid = total_result[0]
        avg_salary = total_result[1]

        # За год и кол-во уникальных сотрудников (только оплаченные)
        paid_by_year = 0
        employees_count = 0
        if year:
            year_result = db.query(
                func.coalesce(func.sum(Salary.amount), 0),
                func.count(func.distinct(Salary.employee_id))
            ).filter(
                Salary.payment_status == 'paid',
                Salary.report_month.like(f'{year}-%')
            ).first()
            paid_by_year = year_result[0]
            employees_count = year_result[1]

        # За месяц (только оплаченные)
        paid_by_month = 0
        if year and month:
            paid_by_month = db.query(
                func.coalesce(func.sum(Salary.amount), 0)
            ).filter(
                Salary.payment_status == 'paid',
                Salary.report_month == f'{year}-{month:02d}'
            ).scalar()

        # По типу проекта (только оплаченные)
        by_project_type = 0
        if project_type:
            by_project_type = db.query(
                func.coalesce(func.sum(Salary.amount), 0)
            ).filter(
                Salary.payment_status == 'paid',
                Salary.project_type == project_type
            ).scalar()

        return {
            'total_paid': float(total_paid) if total_paid else 0,
            'paid_by_year': float(paid_by_year) if paid_by_year else 0,
            'paid_by_month': float(paid_by_month) if paid_by_month else 0,
            'by_project_type': float(by_project_type) if by_project_type else 0,
            'avg_salary': float(avg_salary) if avg_salary else 0,
            'employees_count': int(employees_count) if employees_count else 0
        }

    except Exception as e:
        logger.error(f"Salaries salary dashboard error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dashboard/salaries-supervision")
async def get_salaries_supervision_dashboard(
    year: Optional[int] = None,
    month: Optional[int] = None,
    agent_type: Optional[str] = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Дашборд 'Авторский надзор': total_paid, paid_by_year, paid_by_month, by_agent, avg_payment, payments_count"""
    try:
        # Всего выплачено и средний чек (надзорные = supervision_card_id IS NOT NULL)
        total_result = db.query(
            func.coalesce(func.sum(Payment.final_amount), 0),
            func.coalesce(func.avg(Payment.final_amount), 0)
        ).filter(
            Payment.payment_status == 'paid',
            Payment.supervision_card_id.isnot(None)
        ).first()
        total_paid = total_result[0]
        avg_payment = total_result[1]

        # За год и кол-во выплат
        paid_by_year = 0
        payments_count = 0
        if year:
            year_result = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0),
                func.count(Payment.id)
            ).filter(
                Payment.payment_status == 'paid',
                Payment.supervision_card_id.isnot(None),
                Payment.report_month.like(f'{year}-%')
            ).first()
            paid_by_year = year_result[0]
            payments_count = year_result[1]

        # За месяц
        paid_by_month = 0
        if year and month:
            paid_by_month = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).filter(
                Payment.payment_status == 'paid',
                Payment.supervision_card_id.isnot(None),
                Payment.report_month == f'{year}-{month:02d}'
            ).scalar()

        # По агенту (нужен JOIN через supervision_cards -> contracts)
        by_agent = 0
        if agent_type:
            by_agent = db.query(
                func.coalesce(func.sum(Payment.final_amount), 0)
            ).join(SupervisionCard, Payment.supervision_card_id == SupervisionCard.id
            ).join(Contract, SupervisionCard.contract_id == Contract.id
            ).filter(
                Payment.payment_status == 'paid',
                Payment.supervision_card_id.isnot(None),
                Contract.agent_type == agent_type
            ).scalar()

        return {
            'total_paid': float(total_paid) if total_paid else 0,
            'paid_by_year': float(paid_by_year) if paid_by_year else 0,
            'paid_by_month': float(paid_by_month) if paid_by_month else 0,
            'by_agent': float(by_agent) if by_agent else 0,
            'avg_payment': float(avg_payment) if avg_payment else 0,
            'payments_count': int(payments_count) if payments_count else 0
        }

    except Exception as e:
        logger.error(f"Salaries supervision dashboard error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dashboard/agent-types")
async def get_agent_types(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить список всех типов агентов"""
    try:
        agents = db.query(Contract.agent_type).filter(
            Contract.agent_type.isnot(None),
            Contract.agent_type != ''
        ).distinct().order_by(Contract.agent_type).all()

        return [agent[0] for agent in agents]

    except Exception as e:
        logger.error(f"Agent types error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dashboard/contract-years")
async def get_contract_years(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить список всех годов из договоров (для фильтров дашборда)

    Возвращает список годов в обратном порядке (от нового к старому),
    включая текущий год и следующий год.
    """
    try:
        from datetime import datetime
        from sqlalchemy import extract

        # Получаем все уникальные годы из дат договоров
        years_query = db.query(
            extract('year', Contract.contract_date).label('year')
        ).filter(
            Contract.contract_date.isnot(None)
        ).distinct().all()

        db_years = set()
        for row in years_query:
            if row.year:
                db_years.add(int(row.year))

        # Добавляем текущий год и следующий год
        current_year = datetime.now().year
        next_year = current_year + 1

        db_years.add(current_year)
        db_years.add(next_year)

        # Сортируем в обратном порядке (от нового к старому)
        years_list = sorted(db_years, reverse=True)

        return years_list

    except Exception as e:
        logger.error(f"Contract years error: {e}")
        # Возвращаем fallback
        from datetime import datetime
        current_year = datetime.now().year
        return list(range(current_year + 1, current_year - 10, -1))


# =========================
# SYNC DATA ENDPOINTS
# =========================

@app.get("/api/sync/stage-executors")
async def get_all_stage_executors(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить всех исполнителей стадий для синхронизации"""
    try:
        executors = db.query(StageExecutor).all()

        return [{
            'id': e.id,
            'crm_card_id': e.crm_card_id,
            'stage_name': e.stage_name,
            'executor_id': e.executor_id,
            'assigned_date': e.assigned_date.isoformat() if e.assigned_date else None,
            'assigned_by': e.assigned_by,
            'deadline': e.deadline.isoformat() if e.deadline else (e.deadline if isinstance(e.deadline, str) else None),
            'completed': e.completed,
            'completed_date': e.completed_date.isoformat() if e.completed_date else None,
            'submitted_date': e.submitted_date.isoformat() if e.submitted_date else None
        } for e in executors]

    except Exception as e:
        logger.error(f"Sync stage executors error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/sync/approval-deadlines")
async def get_all_approval_deadlines(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить все дедлайны согласования для синхронизации"""
    try:
        deadlines = db.query(ApprovalStageDeadline).all()

        return [{
            'id': d.id,
            'crm_card_id': d.crm_card_id,
            'stage_name': d.stage_name,
            'deadline': d.deadline if isinstance(d.deadline, str) else (d.deadline.isoformat() if d.deadline else None),
            'is_completed': d.is_completed,
            'completed_date': d.completed_date.isoformat() if d.completed_date else None,
            'created_at': d.created_at.isoformat() if d.created_at else None
        } for d in deadlines]

    except Exception as e:
        logger.error(f"Sync approval deadlines error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/sync/action-history")
async def get_all_action_history(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить всю историю действий для синхронизации"""
    try:
        history = db.query(ActionHistory).all()

        return [{
            'id': h.id,
            'user_id': h.user_id,
            'action_type': h.action_type,
            'entity_type': h.entity_type,
            'entity_id': h.entity_id,
            'description': h.description,
            'action_date': h.action_date.isoformat() if h.action_date else None
        } for h in history]

    except Exception as e:
        logger.error(f"Sync action history error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/sync/supervision-history")
async def get_all_supervision_history(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить всю историю проектов надзора для синхронизации"""
    try:
        history = db.query(SupervisionProjectHistory).all()

        return [{
            'id': h.id,
            'supervision_card_id': h.supervision_card_id,
            'entry_type': h.entry_type,
            'message': h.message,
            'created_by': h.created_by,
            'created_at': h.created_at.isoformat() if h.created_at else None
        } for h in history]

    except Exception as e:
        logger.error(f"Sync supervision history error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =========================
# MESSENGER ENDPOINTS
# =========================

# --- Вспомогательная функция: формирование названия чата ---
def _build_chat_title(contract, card) -> str:
    """Формирует название чата: ИН-Город-Адрес"""
    type_prefix = {
        'Индивидуальный': 'ИН',
        'Шаблонный': 'ШП',
    }
    prefix = type_prefix.get(contract.project_type, 'ПР')
    city = (contract.city or '').replace('_', '-')
    address = (contract.address or '').replace('_', '-')
    # Убираем дублирование города в адресе
    if city and address:
        city_map = {'спб': 'санкт-петербург', 'мск': 'москва', 'нск': 'новосибирск', 'екб': 'екатеринбург'}
        full_city = city_map.get(city.lower(), city.lower())
        addr_check = address.lower().replace('_', '-')
        for c in [full_city, city.lower()]:
            if addr_check.startswith(c):
                address = address[len(c):].lstrip('.,;:_ -')
                break
    return f"{prefix}-{city}-{address}"


def _build_script_context(db: Session, card, contract) -> dict:
    """Собрать контекст переменных для скриптов"""
    ctx = {
        'client_name': '',
        'stage_name': '',
        'deadline': '',
        'manager_name': '',
        'senior_manager': '',
        'sdp': '',
        'director': '',
        'address': contract.address or '',
        'city': contract.city or '',
        'project_type': contract.project_type or '',
        'contract_number': contract.contract_number or '',
        'area': str(contract.area or ''),
    }

    # Клиент
    if contract.client_id:
        client = db.query(Client).filter(Client.id == contract.client_id).first()
        if client:
            ctx['client_name'] = client.full_name or ''

    # Менеджер
    if card.manager_id:
        mgr = db.query(Employee).filter(Employee.id == card.manager_id).first()
        if mgr:
            ctx['manager_name'] = mgr.full_name or ''

    # Старший менеджер
    if card.senior_manager_id:
        sm = db.query(Employee).filter(Employee.id == card.senior_manager_id).first()
        if sm:
            ctx['senior_manager'] = sm.full_name or ''

    # СДП
    if card.sdp_id:
        sdp = db.query(Employee).filter(Employee.id == card.sdp_id).first()
        if sdp:
            ctx['sdp'] = sdp.full_name or ''

    # Руководитель студии
    director = db.query(Employee).filter(Employee.position == 'Руководитель студии').first()
    if director:
        ctx['director'] = director.full_name or ''

    return ctx


async def _send_invites_to_members(chat_id: int, db: Session):
    """Разослать invite-ссылки участникам чата"""
    chat = db.query(MessengerChat).filter(MessengerChat.id == chat_id).first()
    if not chat or not chat.invite_link:
        return

    members = db.query(MessengerChatMember).filter(
        MessengerChatMember.messenger_chat_id == chat_id,
        MessengerChatMember.invite_status == 'pending'
    ).all()

    tg = get_telegram_service()
    email_svc = get_email_service()

    for member in members:
        sent = False

        # Пробуем через Telegram бота (личное сообщение)
        if member.telegram_user_id and tg.bot_available:
            try:
                await tg.send_message(
                    member.telegram_user_id,
                    f"Вас пригласили в проектный чат: {chat.chat_title}\n"
                    f"Присоединяйтесь: {chat.invite_link}"
                )
                member.invite_status = 'sent'
                sent = True
            except Exception:
                pass

        # Если не получилось через Telegram — отправляем email
        if not sent and member.email and email_svc.available:
            # Получаем имя участника
            name = ""
            if member.member_type == 'employee':
                emp = db.query(Employee).filter(Employee.id == member.member_id).first()
                name = emp.full_name if emp else ""
            elif member.member_type == 'client':
                cl = db.query(Client).filter(Client.id == member.member_id).first()
                name = cl.full_name if cl else ""

            success = await email_svc.send_chat_invite(
                to_email=member.email,
                recipient_name=name,
                chat_title=chat.chat_title or "",
                invite_link=chat.invite_link,
            )
            if success:
                member.invite_status = 'email_sent'
                sent = True

        member.invited_at = datetime.utcnow() if sent else None

    db.commit()


async def _trigger_messenger_notification(
    db: Session,
    crm_card_id: int,
    script_type: str,
    stage_name: str = "",
):
    """
    Хук автоуведомлений: найти чат карточки → найти подходящий скрипт → отправить.
    Вызывается из workflow-эндпоинтов через asyncio.create_task().
    script_type: 'project_start' | 'stage_complete' | 'project_end'
    """
    try:
        # Найти активный чат для карточки
        chat = db.query(MessengerChat).filter(
            MessengerChat.crm_card_id == crm_card_id,
            MessengerChat.is_active == True
        ).first()
        if not chat or not chat.telegram_chat_id:
            return  # Чат не создан или не привязан к Telegram

        # Найти подходящий скрипт
        scripts_query = db.query(MessengerScript).filter(
            MessengerScript.script_type == script_type,
            MessengerScript.is_enabled == True
        )
        # Для stage_complete — ищем скрипт конкретной стадии или общий
        if script_type == 'stage_complete' and stage_name:
            specific = scripts_query.filter(
                MessengerScript.stage_name == stage_name
            ).first()
            if specific:
                script = specific
            else:
                script = scripts_query.filter(
                    (MessengerScript.stage_name == None) | (MessengerScript.stage_name == '')
                ).first()
        else:
            script = scripts_query.first()

        if not script:
            return  # Нет включённого скрипта для этого события

        # Собрать контекст
        card = db.query(CRMCard).filter(CRMCard.id == crm_card_id).first()
        if not card:
            return
        contract = db.query(Contract).filter(Contract.id == card.contract_id).first()
        if not contract:
            return

        ctx = _build_script_context(db, card, contract)
        ctx['stage_name'] = stage_name or card.column_name or ''

        # Отправить через Telegram
        tg = get_telegram_service()
        msg_id = await tg.send_script_message(
            chat_id=chat.telegram_chat_id,
            template=script.message_template,
            context=ctx,
        )

        # Записать в лог
        if msg_id:
            log_entry = MessengerMessageLog(
                messenger_chat_id=chat.id,
                message_type=f'auto_{script_type}',
                message_text=tg.render_template(script.message_template, ctx),
                sent_by=None,
                telegram_message_id=msg_id,
                delivery_status='sent',
            )
            db.add(log_entry)
            db.commit()
            logger.info(
                f"Автоуведомление отправлено: card={crm_card_id}, "
                f"type={script_type}, stage={stage_name}"
            )
    except Exception as e:
        logger.error(f"Ошибка автоуведомления (card={crm_card_id}): {e}")


@app.post("/api/messenger/chats", response_model=MessengerChatDetailResponse)
async def create_messenger_chat(
    data: MessengerChatCreate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Создать чат автоматически (MTProto) для CRM-карточки"""
    # Перечитываем настройки (для консистентности между воркерами)
    messenger_settings = {}
    for row in db.query(MessengerSetting).all():
        messenger_settings[row.setting_key] = row.setting_value or ""
    tg_svc = get_telegram_service()
    tg_svc.configure(messenger_settings)

    # Проверка: уже есть активный чат
    existing = db.query(MessengerChat).filter(
        MessengerChat.crm_card_id == data.crm_card_id,
        MessengerChat.is_active == True
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Чат для этой карточки уже существует")

    # Получаем карточку и контракт
    card = db.query(CRMCard).filter(CRMCard.id == data.crm_card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="CRM-карточка не найдена")
    contract = db.query(Contract).filter(Contract.id == card.contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")

    tg = get_telegram_service()
    if not tg.mtproto_available:
        raise HTTPException(status_code=503, detail="MTProto не настроен. Используйте привязку чата.")

    # Формируем название чата
    chat_title = _build_chat_title(contract, card)

    # Определяем фото
    avatar_type = (contract.agent_type or '').lower()
    if 'фестиваль' in avatar_type or 'festival' in avatar_type:
        avatar_type = 'festival'
    elif 'петрович' in avatar_type or 'petrovich' in avatar_type:
        avatar_type = 'petrovich'
    else:
        avatar_type = 'festival'

    photo_path = os.path.join(os.path.dirname(__file__), '..', 'resources', f'{avatar_type}_logo.png')
    if not os.path.exists(photo_path):
        photo_path = None

    # Получаем username бота
    bot_username = None
    if tg.bot_available:
        try:
            bot_info = await tg._bot.get_me()
            bot_username = bot_info.username
        except Exception:
            pass

    # Создаём группу
    result = await tg.create_group(
        title=chat_title,
        photo_path=photo_path,
        bot_username=bot_username,
    )

    # Сохраняем в БД
    chat = MessengerChat(
        contract_id=contract.id,
        crm_card_id=data.crm_card_id,
        messenger_type=data.messenger_type,
        telegram_chat_id=result["chat_id"],
        chat_title=result["title"],
        invite_link=result["invite_link"],
        avatar_type=avatar_type,
        creation_method="auto",
        created_by=current_user.id,
        is_active=True,
    )
    db.add(chat)
    db.flush()

    # Добавляем участников
    members_resp = _add_chat_members(db, chat, data.members, contract, card)

    db.commit()

    # Рассылаем invite-ссылки асинхронно
    asyncio.create_task(_send_invites_to_members(chat.id, db))

    return MessengerChatDetailResponse(
        chat=MessengerChatResponse.model_validate(chat),
        members=members_resp
    )


@app.post("/api/messenger/chats/bind", response_model=MessengerChatDetailResponse)
async def bind_messenger_chat(
    data: MessengerChatBind,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Привязать существующий чат по invite-ссылке"""
    # Проверка: уже есть активный чат
    existing = db.query(MessengerChat).filter(
        MessengerChat.crm_card_id == data.crm_card_id,
        MessengerChat.is_active == True
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Чат для этой карточки уже существует")

    card = db.query(CRMCard).filter(CRMCard.id == data.crm_card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="CRM-карточка не найдена")
    contract = db.query(Contract).filter(Contract.id == card.contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")

    chat_title = _build_chat_title(contract, card)

    # Определяем avatar_type
    avatar_type = (contract.agent_type or '').lower()
    if 'фестиваль' in avatar_type or 'festival' in avatar_type:
        avatar_type = 'festival'
    else:
        avatar_type = 'petrovich'

    # Пробуем получить chat_id через бота (бот должен быть в группе)
    tg = get_telegram_service()
    telegram_chat_id = None

    if tg.bot_available:
        resolved = await tg.resolve_invite_link(data.invite_link)
        if resolved:
            telegram_chat_id = resolved

    chat = MessengerChat(
        contract_id=contract.id,
        crm_card_id=data.crm_card_id,
        messenger_type=data.messenger_type,
        telegram_chat_id=telegram_chat_id,
        chat_title=chat_title,
        invite_link=data.invite_link,
        avatar_type=avatar_type,
        creation_method="manual",
        created_by=current_user.id,
        is_active=True,
    )
    db.add(chat)
    db.flush()

    # Участники
    members_resp = _add_chat_members(db, chat, data.members, contract, card)

    db.commit()

    # Рассылаем invite-ссылки
    asyncio.create_task(_send_invites_to_members(chat.id, db))

    return MessengerChatDetailResponse(
        chat=MessengerChatResponse.model_validate(chat),
        members=members_resp
    )


def _add_chat_members(
    db: Session, chat: MessengerChat, members_input: list,
    contract: Contract, card: CRMCard
) -> list:
    """Добавить участников в чат"""
    members_resp = []

    for m in members_input:
        phone = None
        email = None

        if m.member_type == 'employee':
            emp = db.query(Employee).filter(Employee.id == m.member_id).first()
            if emp:
                phone = emp.phone
                email = emp.email
        elif m.member_type == 'client':
            cl = db.query(Client).filter(Client.id == m.member_id).first()
            if cl:
                phone = cl.phone
                email = cl.email

        member = MessengerChatMember(
            messenger_chat_id=chat.id,
            member_type=m.member_type,
            member_id=m.member_id,
            role_in_project=m.role_in_project,
            is_mandatory=m.is_mandatory,
            phone=phone,
            email=email,
            invite_status='pending',
        )
        db.add(member)
        db.flush()
        members_resp.append(ChatMemberResponse.model_validate(member))

    return members_resp


@app.get("/api/messenger/chats/by-card/{card_id}", response_model=MessengerChatDetailResponse)
async def get_messenger_chat_by_card(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить чат по CRM-карточке"""
    chat = db.query(MessengerChat).filter(
        MessengerChat.crm_card_id == card_id,
        MessengerChat.is_active == True
    ).first()
    if not chat:
        raise HTTPException(status_code=404, detail="Чат не найден")

    members = db.query(MessengerChatMember).filter(
        MessengerChatMember.messenger_chat_id == chat.id
    ).all()

    return MessengerChatDetailResponse(
        chat=MessengerChatResponse.model_validate(chat),
        members=[ChatMemberResponse.model_validate(m) for m in members]
    )


@app.delete("/api/messenger/chats/{chat_id}")
async def delete_messenger_chat(
    chat_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Удалить/отвязать чат"""
    chat = db.query(MessengerChat).filter(MessengerChat.id == chat_id).first()
    if not chat:
        raise HTTPException(status_code=404, detail="Чат не найден")

    tg = get_telegram_service()

    # Если чат был создан автоматически — пробуем удалить группу
    if chat.creation_method == 'auto' and chat.telegram_chat_id:
        await tg.delete_group(chat.telegram_chat_id)
    elif chat.telegram_chat_id and tg.bot_available:
        # Для привязанного чата — бот просто покидает
        await tg.leave_chat(chat.telegram_chat_id)

    # Помечаем как неактивный
    chat.is_active = False
    db.commit()

    return {"status": "deleted", "chat_id": chat_id}


@app.post("/api/messenger/chats/{chat_id}/message")
async def send_chat_message(
    chat_id: int,
    data: SendMessageRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Отправить сообщение в чат"""
    chat = db.query(MessengerChat).filter(
        MessengerChat.id == chat_id, MessengerChat.is_active == True
    ).first()
    if not chat or not chat.telegram_chat_id:
        raise HTTPException(status_code=404, detail="Чат не найден или не привязан")

    tg = get_telegram_service()
    msg_id = await tg.send_message(chat.telegram_chat_id, data.text)

    # Логируем
    log = MessengerMessageLog(
        messenger_chat_id=chat.id,
        message_type='manual',
        message_text=data.text,
        sent_by=current_user.id,
        telegram_message_id=msg_id,
        delivery_status='sent' if msg_id else 'failed',
    )
    db.add(log)
    db.commit()

    return {"status": "sent" if msg_id else "failed", "telegram_message_id": msg_id}


@app.post("/api/messenger/chats/{chat_id}/send-invites")
async def send_chat_invites(
    chat_id: int,
    data: SendInvitesRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Разослать invite-ссылки участникам"""
    chat = db.query(MessengerChat).filter(MessengerChat.id == chat_id).first()
    if not chat:
        raise HTTPException(status_code=404, detail="Чат не найден")

    await _send_invites_to_members(chat.id, db)
    return {"status": "invites_sent"}


@app.post("/api/messenger/chats/{chat_id}/files")
async def send_files_to_chat(
    chat_id: int,
    data: SendFilesRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Отправить файлы в чат (с Яндекс.Диска)"""
    import tempfile

    chat = db.query(MessengerChat).filter(
        MessengerChat.id == chat_id, MessengerChat.is_active == True
    ).first()
    if not chat or not chat.telegram_chat_id:
        raise HTTPException(status_code=404, detail="Чат не найден или не привязан")

    tg = get_telegram_service()
    yd = get_yandex_disk_service()
    sent_ids = []

    # Собираем Yandex пути: из file_ids + из прямых yandex_paths
    yandex_files = []
    for file_id in (data.file_ids or []):
        pf = db.query(ProjectFile).filter(ProjectFile.id == file_id).first()
        if pf and pf.yandex_path:
            yandex_files.append({
                "yandex_path": pf.yandex_path,
                "file_name": pf.file_name or os.path.basename(pf.yandex_path),
                "file_type": pf.file_type or "file",
            })

    for yp in (data.yandex_paths or []):
        yandex_files.append({
            "yandex_path": yp,
            "file_name": os.path.basename(yp),
            "file_type": "image" if any(yp.lower().endswith(ext) for ext in ['.jpg', '.jpeg', '.png', '.webp']) else "file",
        })

    if not yandex_files:
        raise HTTPException(status_code=400, detail="Нет файлов для отправки")

    # Определяем тип отправки: галерея (изображения) или документы
    image_exts = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}

    if data.as_gallery:
        # Отправка изображений галереей
        images_for_gallery = []
        docs_to_send = []

        for yf in yandex_files:
            ext = os.path.splitext(yf["file_name"])[1].lower()
            if ext in image_exts:
                images_for_gallery.append(yf)
            else:
                docs_to_send.append(yf)

        # Скачиваем и отправляем изображения галереей
        if images_for_gallery:
            photo_bytes_list = []
            for img in images_for_gallery[:10]:  # Telegram ограничение: 10 фото в галерее
                try:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(img["file_name"])[1]) as tmp:
                        yd.download_file(img["yandex_path"], tmp.name)
                        with open(tmp.name, 'rb') as f:
                            photo_bytes_list.append({
                                "bytes": f.read(),
                                "filename": img["file_name"],
                            })
                    os.unlink(tmp.name)
                except Exception as e:
                    logger.warning(f"Ошибка скачивания {img['yandex_path']}: {e}")

            if photo_bytes_list:
                msg_ids = await tg.send_media_group_from_bytes(
                    chat.telegram_chat_id, photo_bytes_list, caption=data.caption
                )
                if msg_ids:
                    sent_ids.extend(msg_ids)

        # Документы отправляем отдельно
        for doc in docs_to_send:
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(doc["file_name"])[1]) as tmp:
                    yd.download_file(doc["yandex_path"], tmp.name)
                    msg_id = await tg.send_document(
                        chat.telegram_chat_id, tmp.name, caption=doc["file_name"]
                    )
                    if msg_id:
                        sent_ids.append(msg_id)
                os.unlink(tmp.name)
            except Exception as e:
                logger.warning(f"Ошибка отправки документа {doc['file_name']}: {e}")
    else:
        # Все файлы как документы (со ссылками)
        links = []
        for yf in yandex_files:
            try:
                public_link = yd.get_public_link(yf["yandex_path"])
                links.append(f'<a href="{public_link}">{yf["file_name"]}</a>')
            except Exception:
                links.append(yf["file_name"])

        if links:
            text = data.caption + "\n\n" if data.caption else ""
            text += "\n".join(links)
            msg_id = await tg.send_message(chat.telegram_chat_id, text, parse_mode="HTML")
            if msg_id:
                sent_ids.append(msg_id)

    # Логируем
    file_names = [yf["file_name"] for yf in yandex_files]
    log = MessengerMessageLog(
        messenger_chat_id=chat.id,
        message_type='files',
        message_text=data.caption or "",
        file_links=",".join(file_names),
        sent_by=current_user.id,
        telegram_message_id=sent_ids[0] if sent_ids else None,
        delivery_status='sent' if sent_ids else 'failed',
    )
    db.add(log)
    db.commit()

    return {
        "status": "sent" if sent_ids else "failed",
        "files_count": len(yandex_files),
        "telegram_message_ids": sent_ids,
    }


# =========================
# MESSENGER SCRIPTS ENDPOINTS
# =========================

@app.get("/api/messenger/scripts", response_model=list[MessengerScriptResponse])
async def get_messenger_scripts(
    project_type: str = None,
    script_type: str = None,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить скрипты (с фильтрацией)"""
    query = db.query(MessengerScript)
    if project_type:
        query = query.filter(
            (MessengerScript.project_type == project_type) | (MessengerScript.project_type == None)
        )
    if script_type:
        query = query.filter(MessengerScript.script_type == script_type)

    scripts = query.order_by(MessengerScript.sort_order).all()
    return [MessengerScriptResponse.model_validate(s) for s in scripts]


@app.post("/api/messenger/scripts", response_model=MessengerScriptResponse)
async def create_messenger_script(
    data: MessengerScriptCreate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Создать скрипт"""
    script = MessengerScript(**data.model_dump())
    db.add(script)
    db.commit()
    db.refresh(script)
    return MessengerScriptResponse.model_validate(script)


@app.put("/api/messenger/scripts/{script_id}", response_model=MessengerScriptResponse)
async def update_messenger_script(
    script_id: int,
    data: MessengerScriptUpdate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить скрипт"""
    script = db.query(MessengerScript).filter(MessengerScript.id == script_id).first()
    if not script:
        raise HTTPException(status_code=404, detail="Скрипт не найден")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(script, key, value)
    script.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(script)
    return MessengerScriptResponse.model_validate(script)


@app.delete("/api/messenger/scripts/{script_id}")
async def delete_messenger_script(
    script_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Удалить скрипт"""
    script = db.query(MessengerScript).filter(MessengerScript.id == script_id).first()
    if not script:
        raise HTTPException(status_code=404, detail="Скрипт не найден")
    db.delete(script)
    db.commit()
    return {"status": "deleted"}


@app.patch("/api/messenger/scripts/{script_id}/toggle")
async def toggle_messenger_script(
    script_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Включить/выключить скрипт"""
    script = db.query(MessengerScript).filter(MessengerScript.id == script_id).first()
    if not script:
        raise HTTPException(status_code=404, detail="Скрипт не найден")
    script.is_enabled = not script.is_enabled
    script.updated_at = datetime.utcnow()
    db.commit()
    return {"id": script.id, "is_enabled": script.is_enabled}


# =========================
# MESSENGER SETTINGS ENDPOINTS
# =========================

@app.get("/api/messenger/settings", response_model=list[MessengerSettingResponse])
async def get_messenger_settings(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить все настройки мессенджера"""
    settings_list = db.query(MessengerSetting).all()
    return [MessengerSettingResponse.model_validate(s) for s in settings_list]


@app.put("/api/messenger/settings")
async def update_messenger_settings(
    data: MessengerSettingsBulkUpdate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить настройки мессенджера (массовое обновление)"""
    for item in data.settings:
        setting = db.query(MessengerSetting).filter(
            MessengerSetting.setting_key == item.setting_key
        ).first()
        if setting:
            setting.setting_value = item.setting_value
            setting.updated_at = datetime.utcnow()
            setting.updated_by = current_user.id
        else:
            setting = MessengerSetting(
                setting_key=item.setting_key,
                setting_value=item.setting_value,
                updated_by=current_user.id,
            )
            db.add(setting)

    db.commit()

    # Переинициализируем сервисы с новыми настройками
    messenger_settings = {}
    for row in db.query(MessengerSetting).all():
        messenger_settings[row.setting_key] = row.setting_value or ""

    tg = get_telegram_service()
    tg.configure(messenger_settings)

    email_svc = get_email_service()
    email_svc.configure(messenger_settings)

    return {"status": "updated", "bot_available": tg.bot_available, "email_available": email_svc.available}


@app.get("/api/messenger/status")
async def get_messenger_status(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Статус сервисов мессенджера (перечитывает настройки из БД для консистентности между воркерами)"""
    # Перечитываем настройки из БД чтобы учесть изменения от другого воркера
    messenger_settings = {}
    for row in db.query(MessengerSetting).all():
        messenger_settings[row.setting_key] = row.setting_value or ""

    tg = get_telegram_service()
    tg.configure(messenger_settings)

    email_svc = get_email_service()
    email_svc.configure(messenger_settings)

    return {
        "telegram_bot_available": tg.bot_available,
        "telegram_mtproto_available": tg.mtproto_available,
        "email_available": email_svc.available,
    }


# =========================
# MESSENGER MTPROTO AUTHORIZATION
# =========================

@app.post("/api/messenger/mtproto/send-code")
async def mtproto_send_code(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Шаг 1: Отправить код подтверждения на телефон для MTProto авторизации"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Только администратор")

    # Перечитываем настройки из БД
    messenger_settings = {}
    for row in db.query(MessengerSetting).all():
        messenger_settings[row.setting_key] = row.setting_value or ""

    tg = get_telegram_service()
    tg.configure(messenger_settings)

    if not PYROGRAM_AVAILABLE:
        raise HTTPException(status_code=503, detail="Pyrogram не установлен на сервере")

    if not messenger_settings.get("telegram_api_id") or not messenger_settings.get("telegram_api_hash"):
        raise HTTPException(status_code=400, detail="API ID и API Hash не заполнены")
    if not messenger_settings.get("telegram_phone"):
        raise HTTPException(status_code=400, detail="Телефон не указан")

    try:
        phone_code_hash = await tg.send_auth_code()
        # Сохраняем hash в БД чтобы любой воркер мог его прочитать
        existing = db.query(MessengerSetting).filter_by(setting_key="telegram_phone_code_hash").first()
        if existing:
            existing.setting_value = phone_code_hash
        else:
            db.add(MessengerSetting(setting_key="telegram_phone_code_hash", setting_value=phone_code_hash))
        db.commit()
        return {"status": "code_sent", "phone": messenger_settings["telegram_phone"]}
    except Exception as e:
        logger.error(f"Ошибка отправки кода MTProto: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/messenger/mtproto/verify-code")
async def mtproto_verify_code(
    data: dict,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Шаг 2: Подтвердить код и активировать MTProto сессию"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Только администратор")

    code = str(data.get("code", "")).strip()
    if not code:
        raise HTTPException(status_code=400, detail="Код не указан")

    # Перечитываем настройки и hash из БД
    messenger_settings = {}
    for row in db.query(MessengerSetting).all():
        messenger_settings[row.setting_key] = row.setting_value or ""

    phone_code_hash = messenger_settings.get("telegram_phone_code_hash", "")
    if not phone_code_hash:
        raise HTTPException(status_code=400, detail="Сначала запросите код через send-code")

    tg = get_telegram_service()
    tg.configure(messenger_settings)

    try:
        user_info = await tg.verify_auth_code(phone_code_hash, code)
        # Удаляем hash — больше не нужен
        hash_row = db.query(MessengerSetting).filter_by(setting_key="telegram_phone_code_hash").first()
        if hash_row:
            db.delete(hash_row)
            db.commit()
        return {"status": "success", "user": user_info}
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Ошибка верификации MTProto: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/messenger/mtproto/session-status")
async def mtproto_session_status(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Проверить статус Pyrogram-сессии"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Только администратор")

    messenger_settings = {}
    for row in db.query(MessengerSetting).all():
        messenger_settings[row.setting_key] = row.setting_value or ""

    tg = get_telegram_service()
    tg.configure(messenger_settings)

    try:
        result = await tg.check_session_valid()
        return result
    except Exception as e:
        logger.error(f"Ошибка проверки MTProto сессии: {e}")
        return {"valid": False, "error": str(e)}


# =========================
# SYNC MESSENGER DATA
# =========================

@app.get("/api/sync/messenger-chats")
async def sync_messenger_chats(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Синхронизация чатов"""
    chats = db.query(MessengerChat).all()
    return [{
        'id': c.id,
        'contract_id': c.contract_id,
        'crm_card_id': c.crm_card_id,
        'messenger_type': c.messenger_type,
        'telegram_chat_id': c.telegram_chat_id,
        'chat_title': c.chat_title,
        'invite_link': c.invite_link,
        'avatar_type': c.avatar_type,
        'creation_method': c.creation_method,
        'created_by': c.created_by,
        'created_at': c.created_at.isoformat() if c.created_at else None,
        'is_active': c.is_active,
    } for c in chats]


@app.get("/api/sync/messenger-scripts")
async def sync_messenger_scripts(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Синхронизация скриптов"""
    scripts = db.query(MessengerScript).all()
    return [{
        'id': s.id,
        'script_type': s.script_type,
        'project_type': s.project_type,
        'stage_name': s.stage_name,
        'message_template': s.message_template,
        'use_auto_deadline': s.use_auto_deadline,
        'is_enabled': s.is_enabled,
        'sort_order': s.sort_order,
        'created_at': s.created_at.isoformat() if s.created_at else None,
        'updated_at': s.updated_at.isoformat() if s.updated_at else None,
    } for s in scripts]


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
