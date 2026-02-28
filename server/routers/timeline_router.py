"""
Роутер таблицы сроков CRM проектов (timeline).
Подключается в main.py через app.include_router(timeline_router, prefix="/api/timeline").
"""
import io
import logging
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from database import get_db, Employee, Contract, ProjectTimelineEntry
from auth import get_current_user
from schemas import TimelineEntryUpdate, TimelineInitRequest
from services.timeline_service import (
    build_project_timeline_template,
    build_template_project_timeline,
)

logger = logging.getLogger(__name__)
router = APIRouter(tags=["timeline"])


@router.get("/")
async def get_all_timelines(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить все записи таймлайна проектов"""
    entries = db.query(ProjectTimelineEntry).order_by(
        ProjectTimelineEntry.contract_id,
        ProjectTimelineEntry.sort_order
    ).all()
    return [
        {c.name: getattr(e, c.name) for c in e.__table__.columns}
        for e in entries
    ]


@router.get("/{contract_id}")
async def get_project_timeline(
    contract_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить таблицу сроков проекта"""
    entries = db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id
    ).order_by(ProjectTimelineEntry.sort_order).all()
    return [
        {c.name: getattr(e, c.name) for c in e.__table__.columns}
        for e in entries
    ]


@router.post("/{contract_id}/init")
async def init_project_timeline(
    contract_id: int,
    request: TimelineInitRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Инициализация таблицы сроков проекта из шаблона"""
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")

    # Проверяем, есть ли уже записи
    existing = db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id
    ).count()
    if existing > 0:
        return {"status": "already_initialized", "count": existing}

    # Используем agent_type из договора для выбора кастомного шаблона нормо-дней
    contract_agent = contract.agent_type or 'Все агенты'

    if request.project_type == 'Шаблонный':
        template_subtype = request.project_subtype or 'Стандарт'
        floors = getattr(request, 'floors', 1) or 1
        entries, contract_term, K = build_template_project_timeline(
            template_subtype, request.area, floors, agent_type=contract_agent
        )
    else:
        entries, contract_term, K = build_project_timeline_template(
            request.project_type, request.area, request.project_subtype, agent_type=contract_agent
        )

    for e in entries:
        record = ProjectTimelineEntry(
            contract_id=contract_id,
            stage_code=e['stage_code'],
            stage_name=e['stage_name'],
            stage_group=e['stage_group'],
            substage_group=e.get('substage_group', ''),
            executor_role=e['executor_role'],
            is_in_contract_scope=e['is_in_contract_scope'],
            sort_order=e['sort_order'],
            raw_norm_days=e.get('raw_norm_days', 0),
            cumulative_days=e.get('cumulative_days', 0),
            norm_days=e.get('norm_days', 0)
        )
        db.add(record)

    db.commit()
    return {
        "status": "initialized",
        "count": len(entries),
        "contract_term": contract_term,
        "area_coefficient": K
    }


@router.post("/{contract_id}/reinit")
async def reinit_project_timeline(
    contract_id: int,
    request: TimelineInitRequest,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Пересоздать таблицу сроков проекта (удалить старые записи и создать заново)"""
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")

    # K10: Проверяем наличие заполненных actual_date перед reinit
    filled_entries = db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id,
        ProjectTimelineEntry.actual_date.isnot(None),
        ProjectTimelineEntry.actual_date != ''
    ).count()
    if filled_entries > 0 and not request.force:
        raise HTTPException(
            status_code=409,
            detail=f"Невозможно пересоздать таблицу сроков: {filled_entries} записей имеют фактические даты. Используйте force=true для принудительного сброса."
        )

    # Удаляем старые записи
    db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id
    ).delete()
    db.flush()

    # Используем agent_type из договора для выбора кастомного шаблона нормо-дней
    contract_agent = contract.agent_type or 'Все агенты'

    if request.project_type == 'Шаблонный':
        template_subtype = request.project_subtype or 'Стандарт'
        floors = getattr(request, 'floors', 1) or 1
        entries, contract_term, K = build_template_project_timeline(
            template_subtype, request.area, floors, agent_type=contract_agent
        )
    else:
        entries, contract_term, K = build_project_timeline_template(
            request.project_type, request.area, request.project_subtype, agent_type=contract_agent
        )

    for e in entries:
        record = ProjectTimelineEntry(
            contract_id=contract_id,
            stage_code=e['stage_code'],
            stage_name=e['stage_name'],
            stage_group=e['stage_group'],
            substage_group=e.get('substage_group', ''),
            executor_role=e['executor_role'],
            is_in_contract_scope=e['is_in_contract_scope'],
            sort_order=e['sort_order'],
            raw_norm_days=e.get('raw_norm_days', 0),
            cumulative_days=e.get('cumulative_days', 0),
            norm_days=e.get('norm_days', 0)
        )
        db.add(record)

    db.commit()
    return {
        "status": "reinitialized",
        "count": len(entries),
        "contract_term": contract_term,
        "area_coefficient": K
    }


@router.put("/{contract_id}/entry/{stage_code}")
async def update_timeline_entry(
    contract_id: int,
    stage_code: str,
    update: TimelineEntryUpdate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить запись таблицы сроков"""
    entry = db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id,
        ProjectTimelineEntry.stage_code == stage_code
    ).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    update_data = update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(entry, key, value)
    entry.updated_at = datetime.utcnow()

    db.commit()
    return {c.name: getattr(entry, c.name) for c in entry.__table__.columns}


@router.get("/{contract_id}/summary")
async def get_timeline_summary(
    contract_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Сводка по таблице сроков"""
    entries = db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id,
        ProjectTimelineEntry.executor_role != 'header'
    ).all()

    total_norm = sum(e.norm_days for e in entries if e.is_in_contract_scope)
    total_actual = sum(e.actual_days for e in entries if e.is_in_contract_scope and e.actual_days > 0)
    overdue_count = sum(1 for e in entries if e.actual_days > e.norm_days > 0)

    return {
        "total_norm_days": total_norm,
        "total_actual_days": total_actual,
        "overdue_count": overdue_count,
        "entries_count": len(entries)
    }


@router.get("/{contract_id}/export/excel")
async def export_timeline_excel(
    contract_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Экспорт таблицы сроков CRM в Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")

    entries = db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id
    ).order_by(ProjectTimelineEntry.sort_order).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица сроков"

    # Шапка
    ws.append(["Адрес:", contract.address or ""])
    ws.append(["Тип проекта:", contract.project_type or ""])
    ws.append(["Площадь:", f"{contract.area or 0} м²"])
    ws.append([])

    # Заголовки таблицы
    headers = ["Действия по этапам", "Дата", "Кол-во дней", "Норма дней", "Статус", "Исполнитель"]
    ws.append(headers)

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Данные
    stage_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    stage_font = Font(bold=True, color="FFFFFF")
    substage_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    substage_font = Font(bold=True)
    overdue_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

    for entry in entries:
        row_data = [
            entry.stage_name,
            entry.actual_date or "",
            entry.actual_days or 0,
            entry.norm_days or 0,
            entry.status or "",
            entry.executor_role or ""
        ]
        ws.append(row_data)
        row_idx = ws.max_row

        # Стили для заголовков этапов
        if entry.executor_role == 'header':
            for col_idx in range(1, 7):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = stage_fill
                cell.font = stage_font
                cell.border = thin_border
        elif entry.executor_role == 'subheader':
            for col_idx in range(1, 7):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = substage_fill
                cell.font = substage_font
                cell.border = thin_border
        else:
            for col_idx in range(1, 7):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
            # Подсветка просроченных
            if entry.actual_days and entry.norm_days and entry.actual_days > entry.norm_days:
                for col_idx in range(1, 7):
                    ws.cell(row=row_idx, column=col_idx).fill = overdue_fill

    # Ширина колонок
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"timeline_{contract_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{contract_id}/export/pdf")
async def export_timeline_pdf(
    contract_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Экспорт таблицы сроков CRM в PDF (фирменный стиль)."""
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import ParagraphStyle
    from pdf_helper import build_timeline_pdf, _font

    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="Договор не найден")

    entries = db.query(ProjectTimelineEntry).filter(
        ProjectTimelineEntry.contract_id == contract_id
    ).order_by(ProjectTimelineEntry.sort_order).all()

    fn = _font()
    cell_style = ParagraphStyle('Cell', fontName=fn, fontSize=8, leading=10)

    headers = ["Этап", "Дата", "Дней", "Норма", "Статус", "Исполнитель"]
    rows = []
    row_styles = []

    for i, entry in enumerate(entries):
        rows.append([
            Paragraph(entry.stage_name or "", cell_style),
            Paragraph(entry.actual_date or "", cell_style),
            Paragraph(str(entry.actual_days or ""), cell_style),
            Paragraph(str(entry.norm_days or ""), cell_style),
            Paragraph(entry.status or "", cell_style),
            Paragraph(
                entry.executor_role
                if entry.executor_role not in ('header', 'subheader')
                else "",
                cell_style,
            ),
        ])
        row_idx = i + 1  # +1 т.к. row 0 — заголовок таблицы
        if entry.executor_role == 'header':
            row_styles.append({'row_idx': row_idx, 'bg': '#2C3E50', 'fg': '#FFFFFF'})
        elif entry.executor_role == 'subheader':
            row_styles.append({'row_idx': row_idx, 'bg': '#D6E4F0'})
        elif entry.actual_days and entry.norm_days and entry.actual_days > entry.norm_days:
            row_styles.append({'row_idx': row_idx, 'bg': '#F2DCDB'})

    pdf_bytes = build_timeline_pdf(
        title="Таблица сроков проекта",
        contract=contract,
        headers=headers,
        rows=rows,
        col_widths=[180, 60, 50, 50, 50, 80],
        row_styles=row_styles,
    )

    from datetime import date
    from urllib.parse import quote
    today = date.today().strftime("%d.%m.%Y")
    address = contract.address or f"договор_{contract_id}"
    ru_name = f'Отчет Таблица сроков {address} от {today}.pdf'
    encoded = quote(ru_name)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                f"attachment; filename*=UTF-8''{encoded}; "
                f"filename=timeline_{contract_id}.pdf"
        },
    )
