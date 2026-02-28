"""
Роутер таблицы сроков надзора (supervision-timeline).
Подключается в main.py через app.include_router(supervision_timeline_router, prefix="/api/supervision-timeline").
"""
import io
import asyncio
import logging
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from database import get_db, Employee, Contract, SupervisionCard, SupervisionTimelineEntry
from auth import get_current_user
from schemas import SupervisionTimelineUpdate
from services.notification_service import trigger_supervision_notification

logger = logging.getLogger(__name__)
router = APIRouter(tags=["supervision-timeline"])

SUPERVISION_STAGES = [
    ('STAGE_1_CERAMIC', 'Стадия 1: Закупка керамогранита'),
    ('STAGE_2_PLUMBING', 'Стадия 2: Закупка сантехники'),
    ('STAGE_3_EQUIPMENT', 'Стадия 3: Закупка оборудования'),
    ('STAGE_4_DOORS', 'Стадия 4: Закупка дверей и окон'),
    ('STAGE_5_WALL', 'Стадия 5: Закупка настенных материалов'),
    ('STAGE_6_FLOOR', 'Стадия 6: Закупка напольных материалов'),
    ('STAGE_7_STUCCO', 'Стадия 7: Лепной декор'),
    ('STAGE_8_LIGHTING', 'Стадия 8: Освещение'),
    ('STAGE_9_APPLIANCES', 'Стадия 9: Бытовая техника'),
    ('STAGE_10_CUSTOM_FURNITURE', 'Стадия 10: Закупка заказной мебели'),
    ('STAGE_11_FACTORY_FURNITURE', 'Стадия 11: Закупка фабричной мебели'),
    ('STAGE_12_DECOR', 'Стадия 12: Закупка декора'),
]


@router.get("/")
async def get_all_supervision_timelines(
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить все записи таймлайна надзора"""
    entries = db.query(SupervisionTimelineEntry).order_by(
        SupervisionTimelineEntry.supervision_card_id,
        SupervisionTimelineEntry.sort_order
    ).all()
    return {
        "entries": [
            {c.name: getattr(e, c.name) for c in e.__table__.columns}
            for e in entries
        ],
        "total": len(entries)
    }


@router.get("/{card_id}")
async def get_supervision_timeline(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить таблицу сроков надзора"""
    entries = db.query(SupervisionTimelineEntry).filter(
        SupervisionTimelineEntry.supervision_card_id == card_id
    ).order_by(SupervisionTimelineEntry.sort_order).all()

    totals = {
        "budget_planned": sum(e.budget_planned or 0 for e in entries),
        "budget_actual": sum(e.budget_actual or 0 for e in entries),
        "budget_savings": sum(e.budget_savings or 0 for e in entries),
        "commission": sum(e.commission or 0 for e in entries),
    }

    return {
        "entries": [
            {c.name: getattr(e, c.name) for c in e.__table__.columns}
            for e in entries
        ],
        "totals": totals,
    }


@router.post("/{card_id}/init")
async def init_supervision_timeline(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Инициализация таблицы сроков надзора (12 стадий)"""
    sv_card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
    if not sv_card:
        raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

    existing = db.query(SupervisionTimelineEntry).filter(
        SupervisionTimelineEntry.supervision_card_id == card_id
    ).count()
    if existing > 0:
        return {"status": "already_initialized", "count": existing}

    for i, (code, name) in enumerate(SUPERVISION_STAGES):
        record = SupervisionTimelineEntry(
            supervision_card_id=card_id,
            stage_code=code,
            stage_name=name,
            sort_order=i + 1,
            status='Не начато'
        )
        db.add(record)

    db.commit()
    return {"status": "initialized", "count": len(SUPERVISION_STAGES)}


@router.put("/{card_id}/entry/{stage_code}")
async def update_supervision_timeline_entry(
    card_id: int,
    stage_code: str,
    update: SupervisionTimelineUpdate,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновить запись таблицы сроков надзора"""
    entry = db.query(SupervisionTimelineEntry).filter(
        SupervisionTimelineEntry.supervision_card_id == card_id,
        SupervisionTimelineEntry.stage_code == stage_code
    ).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    old_status = entry.status
    update_data = update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(entry, key, value)

    # Автоматический расчёт экономии
    if entry.budget_planned is not None and entry.budget_actual is not None:
        entry.budget_savings = entry.budget_planned - entry.budget_actual

    entry.updated_at = datetime.utcnow()
    db.commit()

    # Хук: уведомление в чат при завершении стадии
    new_status = entry.status
    if new_status and new_status != old_status and new_status.lower() in ('выполнено', 'завершено'):
        asyncio.create_task(trigger_supervision_notification(
            db, card_id, 'supervision_stage_complete', stage_name=entry.stage_name
        ))

    return {c.name: getattr(entry, c.name) for c in entry.__table__.columns}


@router.get("/{card_id}/summary")
async def get_supervision_timeline_summary(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Сводка по таблице сроков надзора"""
    entries = db.query(SupervisionTimelineEntry).filter(
        SupervisionTimelineEntry.supervision_card_id == card_id
    ).all()

    total_budget_planned = sum(e.budget_planned or 0 for e in entries)
    total_budget_actual = sum(e.budget_actual or 0 for e in entries)
    total_savings = sum(e.budget_savings or 0 for e in entries)
    total_defects = sum(e.defects_found or 0 for e in entries)
    total_resolved = sum(e.defects_resolved or 0 for e in entries)
    total_visits = sum(e.site_visits or 0 for e in entries)

    status_counts = {}
    for e in entries:
        status_counts[e.status] = status_counts.get(e.status, 0) + 1

    return {
        "total_budget_planned": total_budget_planned,
        "total_budget_actual": total_budget_actual,
        "total_savings": total_savings,
        "total_defects_found": total_defects,
        "total_defects_resolved": total_resolved,
        "total_site_visits": total_visits,
        "status_counts": status_counts,
        "entries_count": len(entries)
    }


@router.get("/{card_id}/export/excel")
async def export_supervision_timeline_excel(
    card_id: int,
    include_commission: bool = True,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Экспорт таблицы сроков надзора в Excel (с/без комиссии)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

    contract = db.query(Contract).filter(Contract.id == card.contract_id).first()
    entries = db.query(SupervisionTimelineEntry).filter(
        SupervisionTimelineEntry.supervision_card_id == card_id
    ).order_by(SupervisionTimelineEntry.sort_order).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица сроков надзора"

    # Шапка
    ws.append(["Адрес:", contract.address if contract else ""])
    ws.append([])

    # Заголовки (с/без комиссии)
    if include_commission:
        headers = ["Стадия", "План. дата", "Факт. дата", "Дней", "Исполнитель",
                   "Бюджет план", "Бюджет факт", "Экономия", "Поставщик",
                   "Комиссия", "Статус", "Примечания"]
        widths = [30, 14, 14, 10, 18, 14, 14, 14, 18, 14, 14, 25]
    else:
        headers = ["Стадия", "План. дата", "Факт. дата", "Дней", "Исполнитель",
                   "Бюджет план", "Бюджет факт", "Экономия", "Поставщик",
                   "Статус", "Примечания"]
        widths = [30, 14, 14, 10, 18, 14, 14, 14, 18, 14, 25]
    ws.append(headers)

    header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Цвета статусов
    status_colors = {
        'В работе': 'FFF8E1',
        'Закуплено': 'E3F2FD',
        'Доставлено': 'E8F5E9',
        'Просрочено': 'FFEBEE',
    }

    def _fmt_date(d):
        """Конвертировать дату YYYY-MM-DD → DD.MM.YYYY для экспорта"""
        if not d:
            return ""
        try:
            from datetime import datetime as dt
            return dt.strptime(d, "%Y-%m-%d").strftime("%d.%m.%Y")
        except (ValueError, TypeError):
            return d

    for entry in entries:
        if include_commission:
            row_data = [
                entry.stage_name,
                _fmt_date(entry.plan_date),
                _fmt_date(entry.actual_date),
                entry.actual_days or 0,
                entry.executor or "",
                entry.budget_planned or 0,
                entry.budget_actual or 0,
                entry.budget_savings or 0,
                entry.supplier or "",
                entry.commission or 0,
                entry.status or "",
                entry.notes or "",
            ]
        else:
            row_data = [
                entry.stage_name,
                _fmt_date(entry.plan_date),
                _fmt_date(entry.actual_date),
                entry.actual_days or 0,
                entry.executor or "",
                entry.budget_planned or 0,
                entry.budget_actual or 0,
                entry.budget_savings or 0,
                entry.supplier or "",
                entry.status or "",
                entry.notes or "",
            ]
        ws.append(row_data)
        row_idx = ws.max_row

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).border = thin_border

        # Цвет по статусу
        color_hex = status_colors.get(entry.status)
        if color_hex:
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Сводка ИТОГО
    ws.append([])
    total_planned = sum(e.budget_planned or 0 for e in entries)
    total_actual = sum(e.budget_actual or 0 for e in entries)
    total_savings = sum(e.budget_savings or 0 for e in entries)
    total_commission = sum(e.commission or 0 for e in entries)

    if include_commission:
        total_row_data = ["ИТОГО:", "", "", "", "", total_planned, total_actual,
                          total_savings, "", total_commission, "", ""]
    else:
        total_row_data = ["ИТОГО:", "", "", "", "", total_planned, total_actual,
                          total_savings, "", "", ""]
    ws.append(total_row_data)
    total_row = ws.max_row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.border = thin_border

    # Ширина колонок
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"supervision_timeline_{card_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{card_id}/export/pdf")
async def export_supervision_timeline_pdf(
    card_id: int,
    include_commission: bool = False,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Экспорт таблицы сроков надзора в PDF (с/без комиссии, с ИТОГО)."""
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import ParagraphStyle
    from pdf_helper import build_timeline_pdf, _font, _font_bold, COLOR_GRANDTOTAL

    card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

    contract = db.query(Contract).filter(Contract.id == card.contract_id).first()
    entries = db.query(SupervisionTimelineEntry).filter(
        SupervisionTimelineEntry.supervision_card_id == card_id
    ).order_by(SupervisionTimelineEntry.sort_order).all()

    fn = _font()
    fb = _font_bold()
    cell_style = ParagraphStyle('Cell', fontName=fn, fontSize=8, leading=10)
    bold_style = ParagraphStyle('CellB', fontName=fb, fontSize=8, leading=10)

    STATUS_COLORS = {
        'В работе': '#FFF8E1',
        'Закуплено': '#E3F2FD',
        'Доставлено': '#E8F5E9',
        'Просрочено': '#FFEBEE',
    }

    if include_commission:
        headers = ["Стадия", "План. дата", "Факт. дата", "Дней", "Исполнитель",
                   "Бюджет план", "Бюджет факт", "Поставщик", "Комиссия",
                   "Статус", "Примечания"]
        col_widths = [120, 55, 55, 35, 80, 55, 55, 80, 50, 55, 80]
    else:
        headers = ["Стадия", "План. дата", "Факт. дата", "Дней", "Исполнитель",
                   "Бюджет план", "Бюджет факт", "Поставщик",
                   "Статус", "Примечания"]
        col_widths = [130, 60, 60, 40, 85, 60, 60, 90, 60, 90]

    rows = []
    row_styles = []

    total_planned = 0
    total_actual = 0
    total_commission = 0

    def _fmt_date_pdf(d):
        """Конвертировать дату YYYY-MM-DD → DD.MM.YYYY для экспорта"""
        if not d:
            return ""
        try:
            from datetime import datetime as dt
            return dt.strptime(d, "%Y-%m-%d").strftime("%d.%m.%Y")
        except (ValueError, TypeError):
            return d

    for entry in entries:
        bp = entry.budget_planned or 0
        ba = entry.budget_actual or 0
        cm = entry.commission or 0
        total_planned += bp
        total_actual += ba
        total_commission += cm

        base_cells = [
            Paragraph(entry.stage_name or "", cell_style),
            Paragraph(_fmt_date_pdf(entry.plan_date), cell_style),
            Paragraph(_fmt_date_pdf(entry.actual_date), cell_style),
            Paragraph(str(entry.actual_days or ""), cell_style),
            Paragraph(entry.executor or "", cell_style),
            Paragraph(f"{bp:,.0f}" if bp else "", cell_style),
            Paragraph(f"{ba:,.0f}" if ba else "", cell_style),
            Paragraph(entry.supplier or "", cell_style),
        ]
        if include_commission:
            base_cells.append(Paragraph(f"{cm:,.0f}" if cm else "", cell_style))
        base_cells.extend([
            Paragraph(entry.status or "", cell_style),
            Paragraph(entry.notes or "", cell_style),
        ])
        rows.append(base_cells)

        color_hex = STATUS_COLORS.get(entry.status)
        if color_hex:
            row_styles.append({'row_idx': len(rows), 'bg': color_hex})

    # Строка ИТОГО
    grand_idx = len(rows) + 1
    num_cols = len(headers)
    total_cells = [Paragraph('<b>ИТОГО:</b>', bold_style)]
    total_cells.extend([Paragraph('', cell_style)] * 4)  # даты, дней, исполнитель
    total_cells.append(Paragraph(f'<b>{total_planned:,.0f}</b>', bold_style))
    total_cells.append(Paragraph(f'<b>{total_actual:,.0f}</b>', bold_style))
    total_cells.append(Paragraph('', cell_style))  # поставщик
    if include_commission:
        total_cells.append(Paragraph(f'<b>{total_commission:,.0f}</b>', bold_style))
    total_cells.extend([Paragraph('', cell_style)] * 2)  # статус, примечания
    rows.append(total_cells)
    row_styles.append({'row_idx': grand_idx, 'bg': COLOR_GRANDTOTAL, 'bold': True})

    pdf_bytes = build_timeline_pdf(
        title="Таблица сроков авторского надзора",
        contract=contract,
        headers=headers,
        rows=rows,
        col_widths=col_widths,
        row_styles=row_styles,
    )

    from datetime import date
    from urllib.parse import quote
    today = date.today().strftime("%d.%m.%Y")
    addr = contract.address if contract else f"надзор_{card_id}"
    suffix = " с комиссией" if include_commission else ""
    ru_name = f'Отчет Авторский надзор{suffix} {addr} от {today}.pdf'
    encoded = quote(ru_name)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                f"attachment; filename*=UTF-8''{encoded}; "
                f"filename=supervision_timeline_{card_id}.pdf"
        },
    )
