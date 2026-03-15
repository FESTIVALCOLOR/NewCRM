"""
Роутер выездов и дефектов авторского надзора (supervision-visits).
CRUD + итого по месяцам + экспорт Excel/PDF.
"""
import io
import logging
from datetime import datetime, date
from collections import defaultdict
from typing import List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from urllib.parse import quote

from database import get_db, Employee, Contract, SupervisionCard, SupervisionVisit
from auth import get_current_user
from permissions import require_permission
from schemas import SupervisionVisitCreate, SupervisionVisitUpdate, SupervisionVisitResponse
from services.notification_service import trigger_supervision_notification

logger = logging.getLogger(__name__)
router = APIRouter(tags=["supervision-visits"])

SUPERVISION_STAGES = [
    ('STAGE_1_CERAMIC', 'Стадия 1: Закупка керамогранита'),
    ('STAGE_2_PLUMBING', 'Стадия 2: Закупка сантехники'),
    ('STAGE_3_EQUIPMENT', 'Стадия 3: Закупка оборудования'),
    ('STAGE_4_DOORS', 'Стадия 4: Закупка дверей и окон'),
    ('STAGE_5_WALL', 'Стадия 5: Закупка настенных материалов'),
    ('STAGE_6_FLOOR', 'Стадия 6: Закупка напольных материалов'),
    ('STAGE_7_STUCCO', 'Стадия 7: Лепной декор'),
    ('STAGE_8_LIGHTING', 'Стадия 8: Освещение'),
    ('STAGE_9_APPLIANCES', 'Стадия 9: Бытовая техника'),
    ('STAGE_10_CUSTOM_FURNITURE', 'Стадия 10: Закупка заказной мебели'),
    ('STAGE_11_FACTORY_FURNITURE', 'Стадия 11: Закупка фабричной мебели'),
    ('STAGE_12_DECOR', 'Стадия 12: Закупка декора'),
]


# ── Статические endpoints ПЕРЕД динамическими ──


@router.get("/{card_id}/visits/summary")
async def get_visits_summary(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Итого выездов по месяцам."""
    card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

    visits = db.query(SupervisionVisit).filter(
        SupervisionVisit.supervision_card_id == card_id
    ).order_by(SupervisionVisit.visit_date).all()

    # Группировка по месяцам (YYYY-MM)
    by_month = defaultdict(int)
    for v in visits:
        month_key = v.visit_date[:7] if v.visit_date and len(v.visit_date) >= 7 else "Без даты"
        by_month[month_key] += 1

    return {
        "card_id": card_id,
        "total_visits": len(visits),
        "by_month": dict(sorted(by_month.items())),
    }


@router.get("/{card_id}/visits/export/excel")
async def export_visits_excel(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Экспорт выездов в Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

    contract = db.query(Contract).filter(Contract.id == card.contract_id).first()
    visits = db.query(SupervisionVisit).filter(
        SupervisionVisit.supervision_card_id == card_id
    ).order_by(SupervisionVisit.sort_order, SupervisionVisit.visit_date).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Выезды и дефекты"

    # Шапка
    address = contract.address if contract else f"Карточка {card_id}"
    ws.append([f"Выезды и дефекты: {address}"])
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    # Заголовки
    headers = ["Стадия", "Выезд на объект", "ФИО исполнителя (ДАН)", "Примечание"]
    ws.append(headers)
    header_fill = PatternFill(start_color='444444', end_color='444444', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    MONTH_NAMES_RU = ['', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                      'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']

    def _fmt_date(d):
        if not d:
            return ""
        try:
            from datetime import datetime as dt
            return dt.strptime(d, "%Y-%m-%d").strftime("%d.%m.%Y")
        except (ValueError, TypeError):
            return d

    def _fmt_month(ym):
        """YYYY-MM → Январь 2026"""
        try:
            parts = ym.split('-')
            return f"{MONTH_NAMES_RU[int(parts[1])]} {parts[0]}"
        except (IndexError, ValueError):
            return ym

    # Данные
    for visit in visits:
        ws.append([
            visit.stage_name or "",
            _fmt_date(visit.visit_date),
            visit.executor_name or "",
            visit.notes or "",
        ])

    # Итого
    by_month = defaultdict(int)
    for v in visits:
        month_key = v.visit_date[:7] if v.visit_date and len(v.visit_date) >= 7 else "Без даты"
        by_month[month_key] += 1

    ws.append([])
    ws.append(["ИТОГО ВЫЕЗДОВ:", str(len(visits))])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for month_key in sorted(by_month.keys()):
        ws.append([f"  {_fmt_month(month_key)}", str(by_month[month_key])])

    # Ширины
    widths = [35, 18, 25, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"supervision_visits_{card_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{card_id}/visits/export/pdf")
async def export_visits_pdf(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Экспорт выездов в PDF."""
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import ParagraphStyle
    from pdf_helper import (
        build_timeline_pdf, _font, _font_bold,
        COLOR_SUBTOTAL, COLOR_GRANDTOTAL,
    )

    card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

    contract = db.query(Contract).filter(Contract.id == card.contract_id).first()
    visits = db.query(SupervisionVisit).filter(
        SupervisionVisit.supervision_card_id == card_id
    ).order_by(SupervisionVisit.sort_order, SupervisionVisit.visit_date).all()

    fn = _font()
    fb = _font_bold()
    cell_style = ParagraphStyle('Cell', fontName=fn, fontSize=8, leading=10)
    bold_style = ParagraphStyle('CellB', fontName=fb, fontSize=8, leading=10)

    MONTH_NAMES_RU = ['', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                      'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']

    def _fmt_date(d):
        if not d:
            return ""
        try:
            from datetime import datetime as dt
            return dt.strptime(d, "%Y-%m-%d").strftime("%d.%m.%Y")
        except (ValueError, TypeError):
            return d

    def _fmt_month(ym):
        try:
            parts = ym.split('-')
            return f"{MONTH_NAMES_RU[int(parts[1])]} {parts[0]}"
        except (IndexError, ValueError):
            return ym

    headers = ["Стадия", "Выезд на объект", "ФИО исполнителя (ДАН)", "Примечание"]
    rows = []
    row_styles = []

    # Группировка итого по месяцам
    by_month = defaultdict(int)

    for visit in visits:
        rows.append([
            Paragraph(visit.stage_name or "", cell_style),
            Paragraph(_fmt_date(visit.visit_date), cell_style),
            Paragraph(visit.executor_name or "", cell_style),
            Paragraph(visit.notes or "", cell_style),
        ])
        if visit.visit_date and len(visit.visit_date) >= 7:
            by_month[visit.visit_date[:7]] += 1

    # Итого по месяцам
    for month_key in sorted(by_month.keys()):
        row_idx = len(rows) + 1
        rows.append([
            Paragraph(f'<b>{_fmt_month(month_key)}:</b>', bold_style),
            Paragraph(f'<b>{by_month[month_key]} выездов</b>', bold_style),
            Paragraph('', cell_style),
            Paragraph('', cell_style),
        ])
        row_styles.append({'row_idx': row_idx, 'bg': COLOR_SUBTOTAL, 'bold': True})

    # Общий итог
    grand_idx = len(rows) + 1
    rows.append([
        Paragraph('<b>ИТОГО ВЫЕЗДОВ:</b>', bold_style),
        Paragraph(f'<b>{len(visits)}</b>', bold_style),
        Paragraph('', cell_style),
        Paragraph('', cell_style),
    ])
    row_styles.append({'row_idx': grand_idx, 'bg': COLOR_GRANDTOTAL, 'bold': True})

    pdf_bytes = build_timeline_pdf(
        title="Таблица выездов и дефектов",
        contract=contract,
        headers=headers,
        rows=rows,
        col_widths=[200, 100, 150, 200],
        row_styles=row_styles,
    )

    today = date.today().strftime("%d.%m.%Y")
    addr = contract.address if contract else f"надзор_{card_id}"
    ru_name = f'Отчет Выезды и дефекты {addr} от {today}.pdf'
    encoded = quote(ru_name)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                f"attachment; filename*=UTF-8''{encoded}; "
                f"filename=supervision_visits_{card_id}.pdf"
        },
    )


# ── CRUD endpoints (динамические) ──


@router.get("/{card_id}/visits", response_model=List[SupervisionVisitResponse])
async def get_visits(
    card_id: int,
    current_user: Employee = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Получить все выезды по карточке надзора."""
    card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

    visits = db.query(SupervisionVisit).filter(
        SupervisionVisit.supervision_card_id == card_id
    ).order_by(SupervisionVisit.sort_order, SupervisionVisit.visit_date).all()

    return visits


@router.post("/{card_id}/visits", response_model=SupervisionVisitResponse, status_code=201)
async def create_visit(
    card_id: int,
    data: SupervisionVisitCreate,
    current_user: Employee = Depends(require_permission("supervision.update")),
    db: Session = Depends(get_db),
):
    """Создать запись выезда."""
    card = db.query(SupervisionCard).filter(SupervisionCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка надзора не найдена")

    # Определить sort_order
    max_order = db.query(SupervisionVisit).filter(
        SupervisionVisit.supervision_card_id == card_id
    ).count()

    visit = SupervisionVisit(
        supervision_card_id=card_id,
        stage_code=data.stage_code,
        stage_name=data.stage_name,
        visit_date=data.visit_date,
        executor_name=data.executor_name,
        notes=data.notes,
        sort_order=max_order + 1,
    )
    db.add(visit)
    db.commit()
    db.refresh(visit)

    # N4: Автотриггер supervision_visit + уведомление ДАН
    try:
        import asyncio
        visit_date_str = str(data.visit_date) if data.visit_date else ''
        asyncio.create_task(trigger_supervision_notification(
            db, card_id, 'supervision_visit',
            stage_name=data.stage_name or '',
            extra_context={'visit_date': visit_date_str},
        ))
        # Уведомление ДАН о создании выезда
        if card.dan_id:
            from services.notification_dispatcher import dispatch_notification
            contract = db.query(Contract).filter(Contract.id == card.contract_id).first()
            address = contract.address if contract else ''
            asyncio.create_task(dispatch_notification(
                db=db, employee_id=card.dan_id,
                event_type='supervision',
                title=f'Выезд: {address}',
                message=f'Запланирован выезд по надзору {address} на {visit_date_str}.',
                related_entity_type='supervision_card',
                related_entity_id=card_id,
                project_type='supervision',
            ))
    except Exception as e:
        logging.getLogger(__name__).warning(f"supervision_visit trigger: {e}")

    return visit


@router.put("/{card_id}/visits/{visit_id}", response_model=SupervisionVisitResponse)
async def update_visit(
    card_id: int,
    visit_id: int,
    data: SupervisionVisitUpdate,
    current_user: Employee = Depends(require_permission("supervision.update")),
    db: Session = Depends(get_db),
):
    """Обновить запись выезда."""
    visit = db.query(SupervisionVisit).filter(
        SupervisionVisit.id == visit_id,
        SupervisionVisit.supervision_card_id == card_id,
    ).first()
    if not visit:
        raise HTTPException(status_code=404, detail="Запись выезда не найдена")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(visit, key, value)

    visit.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(visit)
    return visit


@router.delete("/{card_id}/visits/{visit_id}")
async def delete_visit(
    card_id: int,
    visit_id: int,
    current_user: Employee = Depends(require_permission("supervision.update")),
    db: Session = Depends(get_db),
):
    """Удалить запись выезда."""
    visit = db.query(SupervisionVisit).filter(
        SupervisionVisit.id == visit_id,
        SupervisionVisit.supervision_card_id == card_id,
    ).first()
    if not visit:
        raise HTTPException(status_code=404, detail="Запись выезда не найдена")

    db.delete(visit)
    db.commit()
    return {"detail": "Запись выезда удалена"}
